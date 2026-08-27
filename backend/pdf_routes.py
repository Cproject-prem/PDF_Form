"""
PDF Form Builder routes.

Adds endpoints for:
  - Uploading PDF templates (stored on local disk under /app/backend/uploads/pdf/)
  - Listing/getting/updating/deleting PDF templates
  - Publishing PDF templates so they can be filled publicly
  - Submitting filled values, generating completed PDFs (stored under /uploads/completed/)
  - Listing & downloading submissions

Field coordinates are stored in PDF points (origin top-left) so we can
later map them onto reportlab/pypdf overlays without rasterising the PDF.
"""
from __future__ import annotations

import base64
import io
import logging
import os
import re
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

from fastapi import (
    APIRouter,
    Depends,
    File,
    Form,
    HTTPException,
    Request,
    Response,
    UploadFile,
)
from pydantic import BaseModel, ConfigDict, Field
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas as rl_canvas

logger = logging.getLogger("jotform.pdf")

# --------------------------------------------------------------------- config
# --------------------------------------------------------------------- config
# All storage locations are overridable via env so operators can point them
# at Docker volumes or shared network mounts.  Missing folders are created
# on boot so the container starts cleanly on a fresh disk.
UPLOAD_DIR = Path(os.environ.get("UPLOAD_DIR", "/app/backend/uploads"))
PDF_DIR = Path(os.environ.get("LOCAL_PDF_TEMPLATES_ROOT", str(UPLOAD_DIR / "pdf")))
COMPLETED_DIR = Path(os.environ.get("LOCAL_COMPLETED_PDF_ROOT", str(UPLOAD_DIR / "completed")))
ASSET_DIR = Path(os.environ.get("LOCAL_ASSETS_ROOT", str(UPLOAD_DIR / "assets")))
for _d in (PDF_DIR, COMPLETED_DIR, ASSET_DIR):
    _d.mkdir(parents=True, exist_ok=True)

MAX_PDF_MB = int(os.environ.get("MAX_PDF_MB", "50"))

# Field types supported by the PDF Form Builder
PDF_FIELD_TYPES = {
    "short_text", "long_text", "number", "date", "time", "email", "phone",
    "dropdown", "checkbox", "radio", "tick", "signature", "initial", "image",
    "file", "qr_code", "barcode", "heading", "paragraph", "static_text",
    "divider", "auto_number", "calculation", "hidden",
}

# --------------------------------------------------------------------- models
class PDFField(BaseModel):
    # Allow extra so PDF fields can carry the same data_source/lookup/formula
    # blobs that normal form fields use — full parity with the Form Builder.
    model_config = ConfigDict(extra="allow")
    id: str
    page: int = 1
    # Coordinates as percentages of the page (0..1) for resolution independence
    x: float = 0.1
    y: float = 0.1
    width: float = 0.2
    height: float = 0.04
    rotation: float = 0.0
    z_index: int = 0
    type: str = "short_text"
    name: str = ""
    label: str = ""
    placeholder: str = ""
    default_value: Any = ""
    static_text: str = ""
    required: bool = False
    read_only: bool = False
    locked: bool = False
    visible: bool = True
    options: List[str] = []
    # Per-option positions for checkbox / radio fields. When present, each
    # option is rendered at its own {x,y,w,h} on the page (independent of
    # the parent field's bounding box). Coords are page-normalized (0..1).
    option_positions: Optional[List[Dict[str, float]]] = None
    no_border: bool = False
    validation: Dict[str, Any] = Field(default_factory=dict)
    font_size: int = 12
    font_auto_fit: bool = True
    font_family: str = "Helvetica"
    is_bold: bool = False
    is_italic: bool = False
    font_color: str = "#111827"
    border_color: str = "#2563EB"
    background_color: str = "#DBEAFE"
    opacity: float = 0.4
    alignment: str = "left"
    conditional_logic: Optional[Dict[str, Any]] = None
    db_mapping: str = ""
    # ---- Parity tabs with Form Builder ----
    data_source: Optional[Dict[str, Any]] = None
    lookup: Optional[Dict[str, Any]] = None
    formula: Optional[Dict[str, Any]] = None
    description: str = ""
    help_text: str = ""


class PDFPage(BaseModel):
    page: int
    width: float    # in PDF points
    height: float


class PDFTemplateIn(BaseModel):
    title: str = "Untitled PDF Form"
    description: str = ""
    fields: List[PDFField] = []
    settings: Dict[str, Any] = Field(default_factory=dict)
    status: str = "draft"
    pages: List[PDFPage] = []
    version: int = 1
    # ---- Assignment fields (row-level security) ----
    assigned_site_ids: List[str] = Field(default_factory=list)
    assigned_vendor_ids: List[str] = Field(default_factory=list)
    assigned_vendor_user_ids: List[str] = Field(default_factory=list)
    assigned_admin_ids: List[str] = Field(default_factory=list)
    assigned_member_ids: List[str] = Field(default_factory=list)
    assigned_department_ids: List[str] = Field(default_factory=list)
    assigned_team_ids: List[str] = Field(default_factory=list)
    assigned_cluster_managers: List[str] = Field(default_factory=list)
    # Optional per-template filename template — see /app/backend/filename_resolver.py.
    filename_template: Optional[str] = ""
    # Optional auto-save document vault path (e.g. /Reports/TBT)
    doc_vault_path: Optional[str] = ""


class PDFTemplate(PDFTemplateIn):
    template_id: str
    slug: str
    owner_id: str
    original_filename: str
    storage_filename: str   # unique on-disk name (uuid.pdf)
    file_size: int
    created_at: str
    updated_at: str
    is_deleted: bool = False
    is_archived: bool = False


class PDFTemplatePatch(BaseModel):
    is_archived: Optional[bool] = None
    status: Optional[str] = None
    title: Optional[str] = None
    settings: Optional[Dict[str, Any]] = None
    filename_template: Optional[str] = None
    doc_vault_path: Optional[str] = None


class PDFSubmissionIn(BaseModel):
    values: Dict[str, Any] = Field(default_factory=dict)


class PDFSubmission(BaseModel):
    submission_id: str
    template_id: str
    template_version: int = 1
    values: Dict[str, Any]
    submitted_by: Optional[str] = None
    submitted_by_email: Optional[str] = None
    submitted_by_name: Optional[str] = None
    ip: Optional[str] = None
    user_agent: Optional[str] = None
    completed_filename: Optional[str] = None  # filename inside completed/
    status: str = "submitted"
    created_at: str


# --------------------------------------------------------------------- utils
def _slug(title: str) -> str:
    base = re.sub(r"[^a-z0-9]+", "-", title.lower()).strip("-")[:40] or "pdf-form"
    return f"{base}-{uuid.uuid4().hex[:6]}"


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _read_pdf_pages(path: Path) -> List[PDFPage]:
    """Return page dimensions (PDF points) for each page."""
    reader = PdfReader(str(path))
    pages: List[PDFPage] = []
    for i, p in enumerate(reader.pages):
        box = p.mediabox
        pages.append(PDFPage(page=i + 1, width=float(box.width), height=float(box.height)))
    return pages


def _auto_detect_fields(path: Path) -> List[Dict[str, Any]]:
    """Parse PDF text and auto-generate form fields based on {{ ... }} markers."""
    from pypdf import PdfReader
    import uuid
    import re
    
    reader = PdfReader(str(path))
    fields = []
    
    pattern = re.compile(r'([-_—–]*)\s*\{\{\s*([a-zA-Z0-9_\s]+?)\s*\}\}\s*([-_—–]*)')

    for i, p in enumerate(reader.pages):
        page_num = i + 1
        page_height = float(p.mediabox.height)
        
        raw_frags = []
        
        def visitor_text(text, cm, tm, fontDict, fontSize):
            if not text.strip():
                return
            abs_x = tm[4] * cm[0] + tm[5] * cm[2] + cm[4]
            abs_y = tm[4] * cm[1] + tm[5] * cm[3] + cm[5]
            raw_frags.append({"text": text.replace("\n", ""), "x": abs_x, "y": abs_y, "fs": fontSize})
            
        try:
            p.extract_text(visitor_text=visitor_text)
        except Exception:
            pass
            
        raw_frags.sort(key=lambda f: (-f["y"], f["x"]))
        
        lines_list = []
        current_line = []
        current_y = None
        for f in raw_frags:
            if current_y is None or abs(f["y"] - current_y) < 3.0:
                current_line.append(f)
                if current_y is None:
                    current_y = f["y"]
            else:
                lines_list.append(current_line)
                current_line = [f]
                current_y = f["y"]
        if current_line:
            lines_list.append(current_line)
            
        for frags in lines_list:
            frags.sort(key=lambda f: f["x"])
            full_text = "".join(f["text"] for f in frags)
            
            with open("auto_detect_debug.txt", "a", encoding="utf-8") as dbg:
                dbg.write(f"LINE Y={current_y}: {repr(full_text)}\n")
            
            for m in pattern.finditer(full_text):
                with open("auto_detect_debug.txt", "a", encoding="utf-8") as dbg:
                    dbg.write(f"MATCH FOUND: {m.group(0)}\n")
                if not frags: continue
                base_f = frags[0]
                
                match_start_idx = m.start()
                match_end_idx = m.end() - 1
                
                start_x = None
                end_x = None
                curr_len = 0
                
                # Iterate through fragments to map the string indices to physical X coordinates
                for f in frags:
                    f_len = len(f["text"])
                    # Use a standard 0.55 multiplier just for interpolation inside a single fragment
                    frag_char_w = f["fs"] * 0.55 if f["fs"] > 0 else 6.0
                    
                    if start_x is None and curr_len <= match_start_idx < curr_len + f_len:
                        char_offset = match_start_idx - curr_len
                        start_x = f["x"] + (char_offset * frag_char_w)
                        
                    if curr_len <= match_end_idx < curr_len + f_len:
                        char_offset = match_end_idx - curr_len
                        end_x = f["x"] + ((char_offset + 1) * frag_char_w)
                        
                    curr_len += f_len

                if start_x is None: start_x = base_f["x"]
                if end_x is None: end_x = start_x + (len(m.group(0)) * (base_f["fs"] * 0.55 if base_f["fs"] > 0 else 6.0))
                
                match_x = start_x
                width = max(end_x - start_x, 40.0)
                
                # Slightly adjust Y so it perfectly centers over the text baseline
                match_y = page_height - base_f["y"] - base_f["fs"] * 1.05
                height = max(base_f["fs"] * 1.5, 20.0)
                
                # Convert to percentages for the frontend
                x_pct = match_x / float(p.mediabox.width) if float(p.mediabox.width) else 0
                y_pct = match_y / page_height if page_height else 0
                w_pct = width / float(p.mediabox.width) if float(p.mediabox.width) else 0.1
                h_pct = height / page_height if page_height else 0.04
                
                field_type_raw = m.group(2).lower().strip()
                
                type_map = {
                    "text": "short_text", "str": "short_text", "txt": "short_text", "short text": "short_text", "long text": "long_text",
                    "dropdown": "dropdown", "select": "dropdown", "drop": "dropdown",
                    "checkbox": "checkbox", "chk": "checkbox", "check": "checkbox",
                    "signature": "signature", "sign": "signature", "sig": "signature",
                    "date": "date", "dt": "date",
                    "number": "number", "num": "number",
                    "radio": "radio", "rad": "radio"
                }
                
                mapped_type = type_map.get(field_type_raw, "short_text")
                
                fields.append({
                    "id": f"f_{uuid.uuid4().hex[:8]}",
                    "type": mapped_type,
                    "label": field_type_raw.title(),
                    "page": page_num,
                    "x": x_pct,
                    "y": y_pct,
                    "width": w_pct,
                    "height": h_pct,
                    "required": False,
                })
            
    return fields


def _safe_filename(orig: str) -> str:
    name = re.sub(r"[^\w.\-]+", "_", orig)
    return name[:120] or "file"


def _resolve_pdf_name(tpl: Optional[Dict[str, Any]], sub: Optional[Dict[str, Any]] = None) -> str:
    if sub and sub.get("completed_filename"):
        return _safe_filename(sub["completed_filename"])
    if tpl and tpl.get("title"):
        stem = re.sub(r"[^\w.\-]+", "_", tpl["title"].strip()).strip("_") or "completed"
        return f"{stem}.pdf"
    if tpl and tpl.get("original_filename"):
        return _safe_filename(tpl["original_filename"])
    return "completed.pdf"


# --------------------------------------------------------------------- PDF generation
def _hex_to_rgb(hex_color: str) -> tuple:
    h = hex_color.lstrip("#")
    if len(h) == 3:
        h = "".join(c * 2 for c in h)
    try:
        return tuple(int(h[i:i + 2], 16) / 255 for i in (0, 2, 4))
    except Exception:
        return (0, 0, 0)


def _resolve_font(family: str, bold: bool, italic: bool) -> str:
    family = family or "Helvetica"
    if family == "Times-Roman":
        if bold and italic: return "Times-BoldItalic"
        if bold: return "Times-Bold"
        if italic: return "Times-Italic"
        return "Times-Roman"
    if family == "Courier":
        if bold and italic: return "Courier-BoldOblique"
        if bold: return "Courier-Bold"
        if italic: return "Courier-Oblique"
        return "Courier"
    # Default to Helvetica
    if bold and italic: return "Helvetica-BoldOblique"
    if bold: return "Helvetica-Bold"
    if italic: return "Helvetica-Oblique"
    return "Helvetica"


def _draw_text(c: rl_canvas.Canvas, text: str, x: float, y: float, w: float, h: float,
               font: str, size: int, color_hex: str, alignment: str = "left",
               is_bold: bool = False, is_italic: bool = False, auto_fit: bool = True) -> None:
    if not text:
        return
    resolved_font = _resolve_font(font, is_bold, is_italic)
    try:
        c.setFont(resolved_font, size)
    except Exception:
        c.setFont("Helvetica", size)
    r, g, b = _hex_to_rgb(color_hex)
    c.setFillColorRGB(r, g, b)
    text = str(text)
    
    if auto_fit:
        while c.stringWidth(text, c._fontname, size) > (w - 4) and size > 4:
            size -= 1
            c.setFont(c._fontname, size)

    text_w = c.stringWidth(text, c._fontname, size)
    if alignment == "center":
        tx = x + (w - text_w) / 2
    elif alignment == "right":
        tx = x + w - text_w - 2
    else:
        tx = x + 2
    # vertically center (PDF y origin = bottom-left)
    baseline = y - h + (h - size) / 2 + size * 0.2
    c.drawString(tx, baseline, text)


def _draw_multiline(c: rl_canvas.Canvas, text: str, x: float, y: float, w: float, h: float,
                    font: str, size: int, color_hex: str,
                    is_bold: bool = False, is_italic: bool = False, auto_fit: bool = True) -> None:
    if not text:
        return
    resolved_font = _resolve_font(font, is_bold, is_italic)
    try:
        c.setFont(resolved_font, size)
    except Exception:
        c.setFont("Helvetica", size)
    r, g, b = _hex_to_rgb(color_hex)
    c.setFillColorRGB(r, g, b)
    
    lines: List[str] = []
    while size > 4:
        line_h = size * 1.2
        lines = []
        fits = True
        for paragraph in str(text).split("\n"):
            words = paragraph.split(" ")
            cur = ""
            for word in words:
                test = (cur + " " + word).strip()
                if c.stringWidth(test, c._fontname, size) <= w - 4:
                    cur = test
                else:
                    if cur:
                        lines.append(cur)
                    cur = word
                    # if a single word is wider than the box, we definitely need to shrink
                    if c.stringWidth(word, c._fontname, size) > w - 4:
                        fits = False
            if cur:
                lines.append(cur)
        
        # Check if the total height of lines fits in the box
        if not auto_fit or (fits and len(lines) * line_h <= h - 4):
            break
        
        size -= 1
        c.setFont(c._fontname, size)

    top = y - 2
    line_h = size * 1.2
    for i, line in enumerate(lines):
        ly = top - (i + 1) * line_h
        if ly < y - h:
            break
        c.drawString(x + 2, ly, line)


def _draw_signature(c: rl_canvas.Canvas, val: Any, x: float, y: float, w: float, h: float) -> None:
    if not val:
        return
    data_url = ""
    if isinstance(val, dict):
        data_url = val.get("data_url") or val.get("url") or ""
    elif isinstance(val, str):
        data_url = val

    if not data_url:
        return

    try:
        if "," in data_url:
            b64 = data_url.split(",", 1)[1]
            raw = base64.b64decode(b64)
        else:
            raw = base64.b64decode(data_url)
        from reportlab.lib.utils import ImageReader
        img = ImageReader(io.BytesIO(raw))
        c.drawImage(img, x, y - h, width=w, height=h, mask="auto", preserveAspectRatio=True, anchor="sw")
    except Exception as e:
        logger.warning(f"signature render failed: {e}")


def _draw_qr(c: rl_canvas.Canvas, text: str, x: float, y: float, w: float, h: float) -> None:
    if not text:
        return
    try:
        import qrcode as _qr
        img = _qr.make(str(text))
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        buf.seek(0)
        from reportlab.lib.utils import ImageReader
        c.drawImage(ImageReader(buf), x, y - h, width=min(w, h), height=min(w, h), mask="auto")
    except Exception as e:
        logger.warning(f"qr render failed: {e}")


def _draw_barcode(c: rl_canvas.Canvas, text: str, x: float, y: float, w: float, h: float) -> None:
    if not text:
        return
    try:
        import barcode
        from barcode.writer import ImageWriter
        cls = barcode.get_barcode_class("code128")
        b = cls(str(text), writer=ImageWriter())
        buf = io.BytesIO()
        b.write(buf, options={"write_text": False})
        buf.seek(0)
        from reportlab.lib.utils import ImageReader
        c.drawImage(ImageReader(buf), x, y - h, width=w, height=h, mask="auto")
    except Exception as e:
        logger.warning(f"barcode render failed: {e}")


def _draw_checkbox(c: rl_canvas.Canvas, checked: bool, x: float, y: float, size: float,
                   color_hex: str = "#111827", no_border: bool = False) -> None:
    r, g, b = _hex_to_rgb(color_hex)
    c.setStrokeColorRGB(r, g, b)
    c.setLineWidth(1)
    box = min(size, 14)
    bx = x + 2
    by = y - box - 2
    
    if not no_border:
        c.rect(bx, by, box, box, stroke=1, fill=0)
        
    if checked:
        c.setLineWidth(1.5)
        c.line(bx + 2, by + box / 2, bx + box / 2 - 1, by + 2)
        c.line(bx + box / 2 - 1, by + 2, bx + box - 2, by + box - 2)


def _draw_radio(c: rl_canvas.Canvas, selected: bool, x: float, y: float, size: float,
                color_hex: str = "#111827") -> None:
    """Draw a radio button (empty circle, filled dot when selected)."""
    r, g, b = _hex_to_rgb(color_hex)
    c.setStrokeColorRGB(r, g, b)
    c.setLineWidth(1)
    diameter = min(size, 14)
    radius = diameter / 2
    cx = x + 2 + radius
    cy = y - radius - 2
    c.circle(cx, cy, radius, stroke=1, fill=0)
    if selected:
        c.setFillColorRGB(r, g, b)
        c.circle(cx, cy, radius * 0.5, stroke=0, fill=1)


def _draw_tick_only(c: rl_canvas.Canvas, x: float, y: float, w: float, h: float,
                    color_hex: str = "#111827") -> None:
    """Draw ONLY a check mark (no surrounding box) sized to the field bounds.
    Used when the source PDF already has an unchecked box printed on the
    page — we just stamp a tick on top instead of drawing a second box."""
    r, g, b = _hex_to_rgb(color_hex)
    c.setFillColorRGB(r, g, b)
    # ReportLab remaps Unicode -> ZapfDingbats glyph codes internally, so we
    # MUST pass the Unicode heavy-check-mark ("\u2714") — passing the raw
    # ASCII byte 0x34 makes ReportLab emit a fallback "n" glyph that renders
    # as an invisible mark (root cause of the 'tick not seen' bug).
    fs = max(6, min(w, h) * 0.95)
    c.setFont("ZapfDingbats", fs)
    # Centre inside the field's box (baseline y; nudge for glyph metrics)
    cx = x + w / 2 - fs * 0.35
    cy = y - h + h / 2 - fs * 0.30
    c.drawString(cx, cy, "\u2714")  # ✔ heavy check mark (ZapfDingbats code 4)


def generate_completed_pdf(template_path: Path, fields: List[PDFField], values: Dict[str, Any],
                           output_path: Path, uploads_root: Optional[Path] = None,
                           submission_id: Optional[str] = None) -> None:
    """Overlay field values on the original PDF and save to output_path.
    Original PDF is never modified; we merge a per-page overlay layer.

    If `uploads_root` + `submission_id` are given, any file/image field whose
    value is `{file_id, filename, ...}` and points to an image is embedded
    into the PDF at the field's position (instead of just its filename).
    """
    reader = PdfReader(str(template_path))
    writer = PdfWriter()

    def _resolve_image_bytes(val: Any) -> Optional[bytes]:
        """Return raw image bytes for supported inputs (data-URL or uploaded
        file ref); else None."""
        # 1. data-URL (signatures)
        if isinstance(val, str) and val.startswith("data:image"):
            try:
                return base64.b64decode(val.split(",", 1)[1])
            except Exception:
                return None
        # 2. uploaded file ref
        if isinstance(val, dict) and val.get("file_id"):
            fname = (val.get("filename") or "").lower()
            ct = (val.get("content_type") or "").lower()
            ext = fname.rsplit(".", 1)[-1] if "." in fname else ""
            if not (ct.startswith("image/") or ext in {"png", "jpg", "jpeg", "gif", "webp", "bmp"}):
                return None
            if not uploads_root:
                return None
            candidates = []
            if submission_id:
                candidates.append(uploads_root / "submissions" / submission_id / (val.get("filename") or ""))
            if ext:
                candidates.append(uploads_root / "tmp" / f"{val['file_id']}.{ext}")
            for p in candidates:
                if p and p.exists():
                    try:
                        return p.read_bytes()
                    except Exception:
                        pass
            # Fallback: search by basename anywhere under uploads/
            try:
                fallback_root = uploads_root.parent
                for cand in fallback_root.rglob(val.get("filename") or "__none__"):
                    if cand.is_file():
                        return cand.read_bytes()
            except Exception:
                pass
        return None

    # group fields per page
    by_page: Dict[int, List[PDFField]] = {}
    for f in fields:
        if not f.visible or f.type in ("divider", "hidden"):
            continue
        by_page.setdefault(int(f.page), []).append(f)

    for idx, page in enumerate(reader.pages, start=1):
        media = page.mediabox
        pw, ph = float(media.width), float(media.height)
        if idx in by_page:
            overlay_buf = io.BytesIO()
            c = rl_canvas.Canvas(overlay_buf, pagesize=(pw, ph))
            for f in by_page[idx]:
                # convert normalized top-left coords -> PDF points bottom-left
                x = f.x * pw
                top_y = ph - (f.y * ph)
                w = f.width * pw
                h = f.height * ph
                val = values.get(f.id, f.default_value if f.default_value is not None else "")
                if f.type in ("heading", "paragraph", "static_text"):
                    val = f.static_text or val or f.label
                ftype = f.type
                if ftype in ("short_text", "number", "email", "phone", "date", "time",
                             "url", "dropdown", "initial", "auto_number",
                             "calculation"):
                    _draw_text(c, val, x, top_y, w, h, f.font_family, f.font_size,
                               f.font_color, f.alignment, getattr(f, "is_bold", False), getattr(f, "is_italic", False))
                elif ftype in ("long_text", "paragraph", "static_text", "heading"):
                    _draw_multiline(c, val, x, top_y, w, h,
                                    f.font_family if ftype != "heading" else f.font_family,
                                    f.font_size if ftype != "heading" else max(f.font_size, 14),
                                    f.font_color, getattr(f, "is_bold", False), getattr(f, "is_italic", False))
                elif ftype == "signature":
                    if (isinstance(val, str) and (val.startswith("data:image") or len(val) > 50)) or (isinstance(val, dict) and (val.get("data_url") or val.get("url"))):
                        _draw_signature(c, val, x, top_y, w, h)
                    elif val:
                        _draw_text(c, str(val), x, top_y, w, h, f.font_family, f.font_size,
                                   f.font_color, f.alignment, getattr(f, "is_bold", False), getattr(f, "is_italic", False))
                elif ftype == "checkbox":
                    # val may be bool or list of selected options.
                    # NEW: when `option_positions` is set, each option has its
                    # own {x, y, w, h} on the page → draw a checkbox+label at
                    # each position independently. Falls back to the legacy
                    # stacked layout when positions aren't set.
                    positions = getattr(f, "option_positions", None) or []
                    if f.options and len(f.options) > 1 and positions:
                        selected = set(val if isinstance(val, list) else [])
                        for i, opt in enumerate(f.options):
                            if i >= len(positions):
                                break
                            p = positions[i] or {}
                            ox = float(p.get("x", f.x)) * pw
                            oy = ph - (float(p.get("y", f.y)) * ph)
                            ow = float(p.get("w", f.width)) * pw
                            oh = float(p.get("h", f.height)) * ph
                            _draw_checkbox(c, opt in selected, ox, oy, oh, f.font_color, no_border=getattr(f, "no_border", False))
                            # Label to the right of the box, vertically centered
                            box_side = min(oh, 14)
                            _draw_text(c, opt, ox + box_side + 6, oy,
                                       ow - box_side - 8, oh,
                                       f.font_family, f.font_size, f.font_color, "left", getattr(f, "is_bold", False), getattr(f, "is_italic", False))
                    elif f.options and len(f.options) > 1:
                        selected = set(val if isinstance(val, list) else [])
                        line_h = max(f.font_size * 1.4, 14)
                        for i, opt in enumerate(f.options):
                            oy = top_y - i * line_h
                            _draw_checkbox(c, opt in selected, x, oy, line_h, f.font_color, no_border=getattr(f, "no_border", False))
                            _draw_text(c, opt, x + line_h + 4, oy, w - line_h - 8,
                                       line_h, f.font_family, f.font_size, f.font_color, "left", getattr(f, "is_bold", False), getattr(f, "is_italic", False))
                    else:
                        # Single-option (or Yes/No boolean) checkbox: the source
                        # PDF usually already prints its own square in the same
                        # spot, so we ONLY stamp the tick mark when checked —
                        # never our own box. Handles all common submission shapes:
                        #   • `True` / `"true"` / `1`
                        #   • `["Option A"]` when `f.options == ["Option A"]`
                        #   • the raw option string itself
                        checked = False
                        if val is True or val == "true" or val == 1 or val == "1":
                            checked = True
                        elif isinstance(val, list) and val:
                            checked = True
                        elif isinstance(val, str) and val.strip():
                            checked = True
                        if checked:
                            _draw_tick_only(c, x, top_y, w, h, f.font_color)
                elif ftype == "radio":
                    # Radio with per-option positions: draw a small circle +
                    # label at each option's position; fill the selected one.
                    positions = getattr(f, "option_positions", None) or []
                    if f.options and positions:
                        selected_val = val if isinstance(val, str) else ""
                        for i, opt in enumerate(f.options):
                            if i >= len(positions):
                                break
                            p = positions[i] or {}
                            ox = float(p.get("x", f.x)) * pw
                            oy = ph - (float(p.get("y", f.y)) * ph)
                            ow = float(p.get("w", f.width)) * pw
                            oh = float(p.get("h", f.height)) * ph
                            _draw_radio(c, opt == selected_val, ox, oy, oh, f.font_color)
                            box_side = min(oh, 14)
                            _draw_text(c, opt, ox + box_side + 6, oy,
                                       ow - box_side - 8, oh,
                                       f.font_family, f.font_size, f.font_color, "left", getattr(f, "is_bold", False), getattr(f, "is_italic", False))
                    else:
                        # Legacy behavior: single label drawn as text
                        _draw_text(c, val, x, top_y, w, h, f.font_family, f.font_size,
                                   f.font_color, f.alignment, getattr(f, "is_bold", False), getattr(f, "is_italic", False))
                elif ftype == "tick":
                    # Single yes/no — same treatment as single checkbox: stamp
                    # only the tick, no box (so we don't overlap the PDF's
                    # own printed square).
                    checked = val is True or val == "true" or val == 1 or val == "1"
                    if checked:
                        _draw_tick_only(c, x, top_y, w, h, f.font_color)
                    # Optional inline label next to the tick, if the field
                    # width is wide enough for it.
                    lbl = getattr(f, "tick_label", None) or ""
                    if lbl and w > h * 2:
                        _draw_text(c, lbl, x + h + 4, top_y, w - h - 8, h,
                                   f.font_family, f.font_size, f.font_color, "left", getattr(f, "is_bold", False), getattr(f, "is_italic", False))
                elif ftype == "qr_code":
                    _draw_qr(c, val, x, top_y, w, h)
                elif ftype == "barcode":
                    _draw_barcode(c, val, x, top_y, w, h)
                elif ftype == "image":
                    img_bytes = _resolve_image_bytes(val)
                    if img_bytes:
                        try:
                            from reportlab.lib.utils import ImageReader
                            img = ImageReader(io.BytesIO(img_bytes))
                            c.drawImage(img, x, top_y - h, width=w, height=h,
                                        mask="auto", preserveAspectRatio=True, anchor="sw")
                        except Exception as e:
                            logger.warning(f"image render failed: {e}")
                elif ftype == "file":
                    img_bytes = _resolve_image_bytes(val)
                    if img_bytes:
                        try:
                            from reportlab.lib.utils import ImageReader
                            img = ImageReader(io.BytesIO(img_bytes))
                            c.drawImage(img, x, top_y - h, width=w, height=h,
                                        mask="auto", preserveAspectRatio=True, anchor="sw")
                        except Exception as e:
                            logger.warning(f"file image render failed: {e}")
                    else:
                        fname = val.get("filename") if isinstance(val, dict) else (val or "")
                        _draw_text(c, f"[file] {fname}" if fname else "", x, top_y, w, h,
                                   f.font_family, f.font_size, f.font_color, f.alignment, getattr(f, "is_bold", False), getattr(f, "is_italic", False))
                elif ftype == "hidden":
                    continue
                else:
                    _draw_text(c, str(val) if val is not None else "", x, top_y, w, h,
                               f.font_family, f.font_size, f.font_color, f.alignment, getattr(f, "is_bold", False), getattr(f, "is_italic", False))
            c.showPage()
            c.save()
            overlay_buf.seek(0)
            overlay = PdfReader(overlay_buf)
            page.merge_page(overlay.pages[0])
        writer.add_page(page)

    with open(output_path, "wb") as fh:
        writer.write(fh)


# --------------------------------------------------------------------- router factory
def build_pdf_router(db, get_current_user, get_optional_user,
                     make_download_token=None, verify_download_token=None,
                     organize_submission_files=None,
                     uploads_root: Optional[Path] = None,
                     _api_prefix="/api"):
    """Build the router; requires DB + auth deps from the main app.
    make_download_token / verify_download_token are injected from server.py so
    anonymous submitters can download their filled PDF via a short-lived token.
    organize_submission_files (async) moves referenced uploaded files into a
    per-submission folder on disk.
    """
    router = APIRouter(prefix="/pdf-forms")
    public = APIRouter(prefix="/public/pdf-forms")
    subs = APIRouter(prefix="/pdf-submissions")
    pub_subs = APIRouter(prefix="/public/pdf-submissions")

    def _owner_query(user) -> Dict[str, Any]:
        from permissions import form_filter, is_super_admin
        q: Dict[str, Any] = {"is_deleted": False}
        if is_super_admin(user):
            return q
        # Re-use the central form_filter — pdf_templates use the same
        # assignment columns (form_id == template_id semantically).
        rls = form_filter(user)
        # rename keys: form_filter uses form_id; pdf collection uses template_id
        rls_str = str(rls)
        if "'form_id'" in rls_str:
            import json as _json
            rls = _json.loads(_json.dumps(rls).replace('"form_id"', '"template_id"'))
        return {"$and": [q, rls]}

    async def _get_template_for_user(template_id: str, user, *, write: bool = False) -> dict:
        from permissions import can_edit_form, can_view_form, is_super_admin
        doc = await db.pdf_templates.find_one({"template_id": template_id, "is_deleted": False},
                                              {"_id": 0})
        if not doc:
            raise HTTPException(404, "Template not found")
        # Adapt doc to look like a Form doc for permission check (form_id alias)
        probe = {**doc, "form_id": doc.get("template_id")}
        if write:
            if not (is_super_admin(user) or can_edit_form(user, probe)):
                raise HTTPException(403, "You do not have permission to edit this PDF form")
        else:
            if not (is_super_admin(user) or can_view_form(user, probe)):
                raise HTTPException(403, "You do not have permission to view this PDF form")
        return doc

    # --- Upload (creates a new template from a PDF) -----------------------
    @router.post("/upload", response_model=PDFTemplate)
    async def upload_pdf(file: UploadFile = File(...),
                         title: Optional[str] = Form(None),
                         auto_detect: bool = Form(True),
                         user=Depends(get_current_user)):
        from permissions import require_can_create_form
        require_can_create_form(user)
        data = await file.read()
        if not data:
            raise HTTPException(400, "Empty file")
        if len(data) > MAX_PDF_MB * 1024 * 1024:
            raise HTTPException(413, f"PDF exceeds {MAX_PDF_MB}MB")
        if not (file.filename or "").lower().endswith(".pdf"):
            raise HTTPException(400, "Only .pdf files are accepted")
        if not data[:4] == b"%PDF":
            raise HTTPException(400, "File is not a valid PDF")
        # save
        storage_name = f"{uuid.uuid4().hex}.pdf"
        target = PDF_DIR / storage_name
        target.write_bytes(data)
        try:
            pages = _read_pdf_pages(target)
            auto_fields = _auto_detect_fields(target) if auto_detect else []
        except Exception as e:
            target.unlink(missing_ok=True)
            raise HTTPException(400, f"Could not read PDF: {e}")
        tid = f"pdftpl_{uuid.uuid4().hex[:12]}"
        t = title or Path(file.filename).stem or "Untitled PDF Form"
        now = _now()
        doc = {
            "template_id": tid,
            "slug": _slug(t),
            "owner_id": user.user_id,
            "title": t,
            "description": "",
            "fields": auto_fields,
            "pages": [p.model_dump() for p in pages],
            "settings": {},
            "status": "draft",
            "version": 1,
            "original_filename": _safe_filename(file.filename),
            "storage_filename": storage_name,
            "file_size": len(data),
            "created_at": now,
            "updated_at": now,
            "is_deleted": False,
            "is_archived": False,
        }
        await db.pdf_templates.insert_one(doc)
        doc.pop("_id", None)
        return PDFTemplate(**doc)

    # --- List ------------------------------------------------------------
    @router.get("", response_model=List[PDFTemplate])
    async def list_templates(archived: bool = False, q: Optional[str] = None,
                             user=Depends(get_current_user)):
        query = _owner_query(user)
        query["is_archived"] = archived
        if q:
            query["title"] = {"$regex": q, "$options": "i"}
        rows = await db.pdf_templates.find(query, {"_id": 0}).sort("updated_at", -1).to_list(500)
        return [PDFTemplate(**r) for r in rows]

    @router.get("/{template_id}", response_model=PDFTemplate)
    async def get_template(template_id: str, user=Depends(get_current_user)):
        return PDFTemplate(**await _get_template_for_user(template_id, user))

    @router.put("/{template_id}", response_model=PDFTemplate)
    async def update_template(template_id: str, body: PDFTemplateIn, user=Depends(get_current_user)):
        existing = await _get_template_for_user(template_id, user, write=True)
        updates = body.model_dump()
        # validate field types
        for f in updates.get("fields", []):
            if f.get("type") not in PDF_FIELD_TYPES:
                raise HTTPException(400, f"Invalid field type: {f.get('type')}")
        updates["updated_at"] = _now()
        # protect pages from being wiped out
        if not updates.get("pages") and existing.get("pages"):
            updates["pages"] = existing["pages"]
        elif not updates.get("pages"):
            storage = existing.get("storage_filename")
            if storage and (PDF_DIR / storage).exists():
                try:
                    updates["pages"] = [p.model_dump() for p in _read_pdf_pages(PDF_DIR / storage)]
                except Exception:
                    pass
        # bump version if fields changed
        if updates.get("fields") != existing.get("fields"):
            updates["version"] = int(existing.get("version", 1)) + 1
        await db.pdf_templates.update_one({"template_id": template_id}, {"$set": updates})
        existing.update(updates)
        return PDFTemplate(**existing)
    @router.post("/{template_id}/auto-detect", response_model=PDFTemplate)
    async def auto_detect_template_fields(template_id: str, user=Depends(get_current_user)):
        existing = await _get_template_for_user(template_id, user, write=True)
        storage = existing.get("storage_filename")
        if not storage:
            raise HTTPException(400, "No original PDF attached to this template")
            
        target = PDF_DIR / storage
        if not target.exists():
            raise HTTPException(404, "Original PDF file not found on server")
            
        try:
            detected = _auto_detect_fields(target)
        except Exception as e:
            raise HTTPException(500, f"Auto-detect failed: {e}")
            
        if not detected:
            return PDFTemplate(**existing)
            
        # Append new fields
        old_fields = existing.get("fields") or []
        new_fields = old_fields + detected
        updates = {
            "fields": new_fields,
            "version": int(existing.get("version", 1)) + 1,
            "updated_at": _now()
        }
        await db.pdf_templates.update_one({"template_id": template_id}, {"$set": updates})
        existing.update(updates)
        return PDFTemplate(**existing)

    @router.post("/{template_id}/replace-pdf", response_model=PDFTemplate)
    async def replace_template_pdf(template_id: str,
                                   file: UploadFile = File(...),
                                   auto_detect: bool = Form(False),
                                   user=Depends(get_current_user)):
        existing = await _get_template_for_user(template_id, user, write=True)
        data = await file.read()
        if not data or not data[:4] == b"%PDF":
            raise HTTPException(400, "Invalid PDF file")
        storage_name = f"{uuid.uuid4().hex}.pdf"
        target = PDF_DIR / storage_name
        target.write_bytes(data)
        pages = _read_pdf_pages(target)
        auto_fields = _auto_detect_fields(target) if auto_detect else existing.get("fields", [])

        now = _now()
        new_version = int(existing.get("version", 1)) + 1
        updates = {
            "storage_filename": storage_name,
            "original_filename": _safe_filename(file.filename),
            "file_size": len(data),
            "pages": [p.model_dump() for p in pages],
            "version": new_version,
            "updated_at": now,
        }
        if auto_detect:
            updates["fields"] = auto_fields
        await db.pdf_templates.update_one({"template_id": template_id}, {"$set": updates})
        existing.update(updates)
        return PDFTemplate(**existing)

    @router.patch("/{template_id}", response_model=PDFTemplate)
    async def patch_template(template_id: str, body: PDFTemplatePatch, user=Depends(get_current_user)):
        existing = await _get_template_for_user(template_id, user, write=True)
        upd = {k: v for k, v in body.model_dump().items() if v is not None}
        if upd.get("status") and upd["status"] not in ("draft", "published", "archived"):
            raise HTTPException(400, "Invalid status")
        upd["updated_at"] = _now()
        await db.pdf_templates.update_one({"template_id": template_id}, {"$set": upd})
        existing.update(upd)
        return PDFTemplate(**existing)

    @router.delete("/{template_id}")
    async def delete_template(template_id: str, user=Depends(get_current_user)):
        await _get_template_for_user(template_id, user, write=True)
        await db.pdf_templates.update_one({"template_id": template_id},
                                          {"$set": {"is_deleted": True, "updated_at": _now()}})
        return {"ok": True}

    @router.get("/{template_id}/file")
    async def download_template_file(template_id: str, user=Depends(get_current_user)):
        doc = await _get_template_for_user(template_id, user)
        path = PDF_DIR / doc["storage_filename"]
        if not path.exists():
            raise HTTPException(404, "Original PDF missing")
        data = path.read_bytes()
        return Response(content=data, media_type="application/pdf",
                        headers={"Content-Disposition":
                                 f'inline; filename="{doc["original_filename"]}"'})

    @router.post("/{template_id}/duplicate", response_model=PDFTemplate)
    async def duplicate_template(template_id: str, user=Depends(get_current_user)):
        from permissions import require_can_create_form
        require_can_create_form(user)
        existing = await _get_template_for_user(template_id, user)
        new_id = f"pdftpl_{uuid.uuid4().hex[:12]}"
        new_storage = f"{uuid.uuid4().hex}.pdf"
        src = PDF_DIR / existing["storage_filename"]
        if src.exists():
            (PDF_DIR / new_storage).write_bytes(src.read_bytes())
        now = _now()
        new_doc = {**existing,
                   "template_id": new_id,
                   "slug": _slug(existing["title"]),
                   "title": existing["title"] + " (Copy)",
                   "owner_id": user.user_id,
                   "storage_filename": new_storage,
                   "created_at": now, "updated_at": now,
                   "is_archived": False, "is_deleted": False,
                   "status": "draft", "version": 1}
        new_doc.pop("_id", None)
        await db.pdf_templates.insert_one(new_doc)
        new_doc.pop("_id", None)
        return PDFTemplate(**new_doc)

    # --- Public submit ----------------------------------------------------
    @public.get("/{slug}")
    async def public_get(slug: str):
        doc = await db.pdf_templates.find_one({"slug": slug, "is_deleted": False},
                                              {"_id": 0, "owner_id": 0})
        if not doc:
            raise HTTPException(404, "Form not found")
        if doc.get("status") != "published":
            raise HTTPException(403, "Form is not published")
        return doc

    @public.get("/{slug}/file")
    async def public_get_file(slug: str):
        doc = await db.pdf_templates.find_one({"slug": slug, "is_deleted": False}, {"_id": 0})
        if not doc or doc.get("status") != "published":
            raise HTTPException(404, "Form not available")
        path = PDF_DIR / doc["storage_filename"]
        if not path.exists():
            raise HTTPException(404, "Original PDF missing")
        return Response(content=path.read_bytes(), media_type="application/pdf",
                        headers={"Content-Disposition":
                                 f'inline; filename="{doc["original_filename"]}"'})

    @public.post("/{slug}/submit")
    async def public_submit(slug: str, body: PDFSubmissionIn, request: Request,
                            viewer=Depends(get_current_user)):
        tpl = await db.pdf_templates.find_one({"slug": slug, "is_deleted": False}, {"_id": 0})
        if not tpl:
            raise HTTPException(404, "Form not found")
        if tpl.get("status") != "published":
            raise HTTPException(403, "Form not accepting submissions")
        # validate required
        for f in tpl.get("fields", []):
            if f.get("required") and f.get("type") not in (
                    "heading", "paragraph", "static_text", "divider", "hidden"):
                v = body.values.get(f["id"])
                if v is None or v == "" or (isinstance(v, list) and len(v) == 0):
                    raise HTTPException(400, f"'{f.get('label') or f['id']}' is required")
        # generate completed PDF
        src = PDF_DIR / tpl["storage_filename"]
        sid = f"pdfsub_{uuid.uuid4().hex[:12]}"
        out_name = f"{sid}.pdf"
        out_path = COMPLETED_DIR / out_name
        try:
            field_models = [PDFField(**f) for f in tpl.get("fields", [])]
            generate_completed_pdf(src, field_models, body.values, out_path,
                                   uploads_root=uploads_root, submission_id=sid)
        except Exception as e:
            logger.exception("PDF generation failed")
            raise HTTPException(500, f"PDF generation failed: {e}")
        # Resolve site_name from submitted values or plant override
        _site_name_val = body.values.get("_overridden_site_name") or body.values.get("site_name") or body.values.get("plant")
        if not _site_name_val:
            _SITE_LABELS = {"site name", "plant name", "site code", "asset id", "site_name", "plant_name", "site_code", "asset_id"}
            for _f in tpl.get("fields", []):
                _label = (_f.get("label") or "").strip().lower()
                _fid   = (_f.get("id") or "")
                if _label in _SITE_LABELS or _fid in {"site_name", "site_code", "plant_name", "asset_id"}:
                    _v = body.values.get(_fid)
                    if _v:
                        _site_name_val = str(_v)
                        break

        doc = {
            "submission_id": sid,
            "template_id": tpl["template_id"],
            "template_version": int(tpl.get("version", 1)),
            "values": body.values,
            "site_name": _site_name_val,
            "submitted_by": viewer.user_id,
            "submitted_by_email": getattr(viewer, "email", None),
            "submitted_by_name": getattr(viewer, "name", None),
            "ip": request.client.host if request.client else None,
            "user_agent": request.headers.get("user-agent"),
            "completed_filename": out_name,
            "status": "submitted",
            "created_at": _now(),
        }
        await db.pdf_submissions.insert_one(doc)
        doc.pop("_id", None)
        # Move any referenced uploads into /submissions/{sid}/ on disk
        if organize_submission_files:
            try:
                await organize_submission_files(sid, body.values)
            except Exception as _e:  # noqa: BLE001
                logger.warning(f"file organization failed for {sid}: {_e}")
        # fire workflow trigger
        try:
            from workflow_routes import fire_trigger as _ft
            # Resolve site_name from submitted values by scanning field labels
            _site_name_val = None
            _SITE_LABELS = {"site name", "plant name", "site code", "asset id", "site_name",
                            "plant_name", "site_code", "asset_id"}
            for _f in tpl.get("fields", []):
                _label = (_f.get("label") or "").strip().lower()
                _fid   = (_f.get("id") or "")
                if _label in _SITE_LABELS or _fid in {"site_name", "site_code", "plant_name", "asset_id"}:
                    _v = body.values.get(_fid)
                    if _v:
                        _site_name_val = str(_v)
                        break
            await _ft(db, "pdf_submitted",
                      {"submission_id": sid, "template_id": tpl["template_id"],
                       "form_name": tpl.get("title"), "values": body.values,
                       "submission_kind": "pdf",
                       "user_id": viewer.user_id,
                       "user_email": viewer.email,
                       "site_name": _site_name_val,
                       "ip": doc.get("ip")})
        except Exception as _e:
            logger.warning(f"workflow trigger pdf_submitted failed: {_e}")

        # Auto-upload generated PDF to Plant Docs Vault if site mapped
        try:
            from filename_resolver import resolve_filename as _rf, _pick, _pick_by_label
            site_identifier = _pick(body.values, ["site_code", "site_id", "asset_id", "plant_code", "plantId", "asset_code", "site"])
            # Fallback: scan PDF template fields by label (e.g. field labeled "Site Code" or "Plant Code")
            if not site_identifier:
                site_identifier = _pick_by_label(tpl.get("fields", []), body.values)
            if site_identifier:
                site_doc = await db.sites.find_one({"$or": [
                    {"site_id": site_identifier},
                    {"site_code": site_identifier},
                    {"asset_id": site_identifier},
                    {"site_name": {"$regex": f"^{site_identifier}", "$options": "i"}},
                ]}, {"_id": 0})
                if site_doc:
                    pdf_bytes = out_path.read_bytes()
                    # rename form object keys to match what resolve_filename expects
                    form_adapter = {"title": tpl.get("title", ""), "filename_template": tpl.get("filename_template", "")}
                    fname = _rf(tpl.get("filename_template"), form=form_adapter, submission=doc)
                    if not fname.lower().endswith(".pdf"):
                        fname += ".pdf"
                    from plant_docs_routes import save_internal_plant_doc, parse_vault_path
                    vault_path = tpl.get("doc_vault_path") or "/PDF Form"
                    folder, subfolder = parse_vault_path(vault_path)
                    if not folder:
                        folder = "PDF Form"
                    save_internal_plant_doc(site_doc["site_id"], folder, fname, pdf_bytes, subfolder_name=subfolder)
        except Exception as _e:
            logger.warning(f"failed to auto-sync PDF Form to vault: {_e}")

        download_token = make_download_token(sid, kind="pdf") if make_download_token else None
        payload = PDFSubmission(**doc).model_dump()
        if download_token:
            payload["download_token"] = download_token
        return payload

    # --- Public download of the filled PDF (token-scoped, anonymous-safe) ---
    def _render_completed_pdf_bytes(tpl: Dict[str, Any], sub: Dict[str, Any]) -> bytes:
        """Dynamically render the completed PDF bytes on-the-fly using the LATEST
        template PDF background file and the submission's stored values.

        This ensures:
          1. Downloads ALWAYS use the updated PDF template version.
          2. Disks do not accumulate static files for every submission (saves storage).
          3. Fallback to existing on-disk file if template file missing.
        """
        src = PDF_DIR / tpl.get("storage_filename", "")
        if src.exists():
            out_temp = COMPLETED_DIR / f"_temp_{uuid.uuid4().hex}.pdf"
            try:
                field_models = [PDFField(**f) for f in tpl.get("fields", [])]
                generate_completed_pdf(src, field_models, sub.get("values") or {},
                                       out_temp, uploads_root=uploads_root,
                                       submission_id=sub.get("submission_id"))
                if out_temp.exists():
                    data = out_temp.read_bytes()
                    return data
            except Exception as _e:
                logger.warning(f"on-the-fly PDF render failed, falling back to static file: {_e}")
            finally:
                out_temp.unlink(missing_ok=True)

        # Fallback to pre-rendered file if original template missing
        path = COMPLETED_DIR / (sub.get("completed_filename") or "")
        if path.exists():
            return path.read_bytes()
        raise HTTPException(404, "Completed PDF missing or template unavailable")

    @pub_subs.get("/{submission_id}/completed")
    async def public_download_completed(submission_id: str, token: str):
        if not verify_download_token:
            raise HTTPException(500, "Download token verifier not configured")
        verify_download_token(token, submission_id, kind="pdf")
        sub = await db.pdf_submissions.find_one({"submission_id": submission_id}, {"_id": 0})
        if not sub:
            raise HTTPException(404, "Not found")
        tpl = await db.pdf_templates.find_one(
            {"template_id": sub["template_id"], "is_deleted": False}, {"_id": 0},
        )
        if not tpl:
            raise HTTPException(404, "Parent template missing")
        pdf_bytes = _render_completed_pdf_bytes(tpl, sub)
        return Response(content=pdf_bytes, media_type="application/pdf",
                        headers={"Content-Disposition":
                                 f'attachment; filename="{_resolve_pdf_name(tpl, sub)}"'})

    # --- Submissions (owner) ---------------------------------------------
    @router.get("/{template_id}/submissions", response_model=List[PDFSubmission])
    async def list_subs(template_id: str, start_date: Optional[str] = None, end_date: Optional[str] = None, user=Depends(get_current_user)):
        await _get_template_for_user(template_id, user)
        scope = await _pdf_submission_scope_query(user)
        q = {"$and": [{"template_id": template_id}]}
        if scope:
            q["$and"].append(scope)
        if start_date or end_date:
            date_flt = {}
            if start_date: date_flt["$gte"] = start_date
            if end_date: date_flt["$lte"] = end_date
            q["$and"].append({"created_at": date_flt})
        rows = await db.pdf_submissions.find(q, {"_id": 0}) \
            .sort("created_at", -1).to_list(2000)
        return [PDFSubmission(**r) for r in rows]

    @subs.get("/{submission_id}", response_model=PDFSubmission)
    async def get_sub(submission_id: str, user=Depends(get_current_user)):
        sub = await db.pdf_submissions.find_one({"submission_id": submission_id}, {"_id": 0})
        if not sub:
            raise HTTPException(404, "Submission not found")
        await _get_template_for_user(sub["template_id"], user)
        if not await _pdf_submission_in_user_scope(user, sub):
            raise HTTPException(403, "You do not have permission to view this submission")
        return PDFSubmission(**sub)

    @subs.delete("/{submission_id}")
    async def del_sub(submission_id: str, user=Depends(get_current_user)):
        sub = await db.pdf_submissions.find_one({"submission_id": submission_id}, {"_id": 0})
        if not sub:
            raise HTTPException(404, "Not found")
        tpl = await _get_template_for_user(sub["template_id"], user)
        if not await _pdf_submission_in_user_scope(user, sub):
            raise HTTPException(403, "You do not have permission to delete this submission")
        if sub.get("completed_filename"):
            (COMPLETED_DIR / sub["completed_filename"]).unlink(missing_ok=True)

        # Delete corresponding file(s) from Plant Document vault
        try:
            from plant_docs_routes import delete_internal_plant_doc, parse_vault_path, delete_plant_docs_by_pattern
            from filename_resolver import resolve_filename as _rf, _pick, _pick_by_label

            site_identifier = _pick(sub.get("values") or {}, ["site_code", "site_id", "asset_id", "plant_code", "plantId", "asset_code", "site"])
            if not site_identifier:
                site_identifier = _pick_by_label(tpl.get("fields", []), sub.get("values") or {})

            site_id = None
            fname = None
            if site_identifier:
                site_doc = await db.sites.find_one({"$or": [
                    {"site_id": site_identifier},
                    {"site_code": site_identifier},
                    {"asset_id": site_identifier},
                    {"site_name": {"$regex": f"^{site_identifier}", "$options": "i"}},
                ]}, {"_id": 0})
                if site_doc:
                    site_id = site_doc.get("site_id")
                    form_adapter = {"title": tpl.get("title", ""), "filename_template": tpl.get("filename_template", "")}
                    fname = _rf(tpl.get("filename_template"), form=form_adapter, submission=sub)
                    if not fname.lower().endswith(".pdf"):
                        fname += ".pdf"
                    vault_path = tpl.get("doc_vault_path") or "/PDF Form"
                    folder, subfolder = parse_vault_path(vault_path)
                    if not folder:
                        folder = "PDF Form"
                    delete_internal_plant_doc(site_id, folder, fname, subfolder_name=subfolder)

            # Also pattern match by completed_filename, fname, and submission_id to clean any variant
            if sub.get("completed_filename"):
                delete_plant_docs_by_pattern(sub["completed_filename"], site_id=site_id)
            if fname:
                delete_plant_docs_by_pattern(fname, site_id=site_id)
            delete_plant_docs_by_pattern(submission_id, site_id=site_id)
        except Exception as _e:
            logger.warning(f"Failed to auto-delete PDF Form from vault on submission delete: {_e}")

        await db.pdf_submissions.delete_one({"submission_id": submission_id})
        return {"ok": True}

    async def _pdf_submission_scope_query(user) -> Dict[str, Any]:
        """Row-level filter for pdf_submissions. Mirrors the standard-form
        `_submission_scope_query` helper in server.py so vendors only see
        their own team's PDF submissions, admins see region/cluster scope,
        super-admin / access_override see all."""
        from permissions import (
            normalize_role, is_super_admin, has_access_override,
            async_submission_filter,
        )
        if is_super_admin(user) or has_access_override(user):
            return {}
        role = normalize_role(getattr(user, "role", ""))
        if role == "admin":
            return await async_submission_filter(db, user)
        vid = getattr(user, "vendor_id", None)
        uid = getattr(user, "user_id", None)
        if role == "vendor_admin" and vid:
            team = await db.users.find(
                {"vendor_id": vid}, {"_id": 0, "user_id": 1},
            ).to_list(2000)
            return {"submitted_by": {"$in": [u["user_id"] for u in team]}}
        if role == "vendor_user" and uid:
            return {"submitted_by": uid}
        return {"submission_id": "__none__"}

    async def _pdf_submission_in_user_scope(user, sub: Dict[str, Any]) -> bool:
        from permissions import normalize_role, is_super_admin, has_access_override
        if is_super_admin(user) or has_access_override(user):
            return True
        role = normalize_role(getattr(user, "role", ""))
        if role == "admin":
            return True
        uid = getattr(user, "user_id", None)
        vid = getattr(user, "vendor_id", None)
        if role == "vendor_user":
            return sub.get("submitted_by") == uid
        if role == "vendor_admin" and vid:
            if sub.get("submitted_by") == uid:
                return True
            submitter = await db.users.find_one(
                {"user_id": sub.get("submitted_by")}, {"_id": 0, "vendor_id": 1},
            )
            return bool(submitter and submitter.get("vendor_id") == vid)
        return False

    async def _can_view_submission(user, sub: Dict[str, Any]) -> bool:
        """Submitter + their vendor_admin can view a PDF submission's completed
        PDF; anyone with RLS access to the parent template can also view."""
        from permissions import normalize_role, is_super_admin
        if is_super_admin(user):
            return True
        if sub.get("submitted_by") == getattr(user, "user_id", None):
            return True
        role = normalize_role(getattr(user, "role", ""))
        if role == "vendor_admin" and getattr(user, "vendor_id", None):
            submitter = await db.users.find_one(
                {"user_id": sub.get("submitted_by")},
                {"_id": 0, "vendor_id": 1},
            )
            if submitter and submitter.get("vendor_id") == user.vendor_id:
                return True
        return False

    @subs.get("/{submission_id}/completed")
    async def download_completed(submission_id: str, user=Depends(get_current_user)):
        sub = await db.pdf_submissions.find_one({"submission_id": submission_id}, {"_id": 0})
        if not sub:
            raise HTTPException(404, "Not found")
        # Allow submitter + vendor_admin fall-through in addition to RLS
        tpl = None
        if await _can_view_submission(user, sub):
            tpl = await db.pdf_templates.find_one(
                {"template_id": sub["template_id"], "is_deleted": False}, {"_id": 0},
            )
        if not tpl:
            tpl = await _get_template_for_user(sub["template_id"], user)
        pdf_bytes = _render_completed_pdf_bytes(tpl, sub)
        return Response(content=pdf_bytes, media_type="application/pdf",
                        headers={"Content-Disposition":
                                 f'attachment; filename="{_resolve_pdf_name(tpl, sub)}"'})

    @subs.get("/{submission_id}/original")
    async def download_original(submission_id: str, user=Depends(get_current_user)):
        sub = await db.pdf_submissions.find_one({"submission_id": submission_id}, {"_id": 0})
        if not sub:
            raise HTTPException(404, "Not found")
        tpl = await _get_template_for_user(sub["template_id"], user)
        path = PDF_DIR / tpl["storage_filename"]
        if not path.exists():
            raise HTTPException(404, "Original PDF missing")
        return Response(content=path.read_bytes(), media_type="application/pdf",
                        headers={"Content-Disposition":
                                 f'attachment; filename="{tpl["original_filename"]}"'})

    # --- Excel export (per PDF template) ---------------------------------
    @router.get("/{template_id}/submissions/export.xlsx")
    async def export_pdf_subs_xlsx(template_id: str, start_date: Optional[str] = None, end_date: Optional[str] = None, user=Depends(get_current_user)):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        tpl = await _get_template_for_user(template_id, user)
        scope = await _pdf_submission_scope_query(user)
        xls_q = {"$and": [{"template_id": template_id}]}
        if scope:
            xls_q["$and"].append(scope)
        if start_date or end_date:
            date_flt = {}
            if start_date: date_flt["$gte"] = start_date
            if end_date: date_flt["$lte"] = end_date
            xls_q["$and"].append({"created_at": date_flt})
        rows = await db.pdf_submissions.find(xls_q, {"_id": 0}) \
            .sort("created_at", 1).to_list(5000)
        skip_types = ("heading", "paragraph", "static_text", "divider", "hidden")
        fields = [f for f in (tpl.get("fields") or []) if f.get("type") not in skip_types]
        field_ids = [f["id"] for f in fields]
        field_labels = {f["id"]: (f.get("label") or f.get("name") or f["id"]) for f in fields}

        wb = Workbook()
        ws = wb.active
        ws.title = (tpl.get("title") or "Submissions")[:31]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="7C3AED")
        sub_font = Font(italic=True, color="475569")
        label_row = ["Submission ID", "Status", "Submitted At", "Submitted By"] + \
                    [field_labels[fid] for fid in field_ids]
        key_row = ["submission_id", "status", "created_at", "submitted_by"] + field_ids
        ws.append(label_row)
        ws.append(key_row)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for cell in ws[2]:
            cell.font = sub_font

        def _val(v):
            if v is None:
                return ""
            if isinstance(v, (list, tuple)):
                return ", ".join(str(x) for x in v)
            if isinstance(v, dict):
                if "filename" in v:
                    return str(v["filename"])
                return str(v)
            if isinstance(v, str) and v.startswith("data:image"):
                return "[signature image]"
            return v

        for r in rows:
            vals = r.get("values") or {}
            ws.append(
                [r.get("submission_id"), r.get("status"), r.get("created_at"),
                 r.get("submitted_by_name") or r.get("submitted_by_email") or r.get("submitted_by") or ""] +
                [_val(vals.get(fid, "")) for fid in field_ids],
            )
        for col in ws.columns:
            w = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 12), 48)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_slug = tpl.get("slug") or template_id
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{safe_slug}-submissions.xlsx"'},
        )

    # --- Image asset upload (for image/signature fields in builder) -------
    @router.post("/assets/upload")
    async def upload_asset(file: UploadFile = File(...), user=Depends(get_current_user)):
        data = await file.read()
        if len(data) > 8 * 1024 * 1024:
            raise HTTPException(413, "Asset exceeds 8MB")
        ext = (file.filename.rsplit(".", 1)[-1] if "." in file.filename else "bin").lower()
        fid = f"{uuid.uuid4().hex}.{ext}"
        (ASSET_DIR / fid).write_bytes(data)
        return {"url": f"/api/pdf-forms/assets/{fid}", "filename": file.filename}

    @router.get("/assets/{fid}")
    async def get_asset(fid: str):
        path = ASSET_DIR / _safe_filename(fid)
        if not path.exists():
            raise HTTPException(404, "Asset not found")
        ct = "image/png" if fid.lower().endswith(".png") else "image/jpeg"
        return Response(content=path.read_bytes(), media_type=ct)

    # --- Edit an existing PDF submission -----------------------------------
    class PDFSubmissionEditIn(BaseModel):
        values: Dict[str, Any]
        edit_reason: str = ""

    @subs.put("/{submission_id}")
    async def edit_submission(submission_id: str, body: PDFSubmissionEditIn,
                              request: Request, user=Depends(get_current_user)):
        """Allow editing a submitted PDF. Permission rules:
        - super_admin / admin: always allowed
        - vendor_admin / vendor_user / submitter: only when the submission's
          workflow approval has been rejected (approval_status == 'rejected' OR
          status == 'rejected').
        Regenerates the completed PDF and writes an audit-log entry.
        After edit, re-fires the pdf_submitted workflow trigger so approval
        restarts from the beginning.
        """
        from permissions import normalize_role, is_super_admin
        sub = await db.pdf_submissions.find_one({"submission_id": submission_id}, {"_id": 0})
        if not sub:
            raise HTTPException(404, "Submission not found")

        role = normalize_role(getattr(user, "role", ""))
        is_admin = is_super_admin(user) or role == "admin"

        if not is_admin:
            # vendor / submitter may only edit after rejection
            sub_status = (sub.get("status") or "").lower()
            approval_status = (sub.get("approval_status") or "").lower()
            is_rejected = sub_status == "rejected" or approval_status == "rejected"

            # Check if user is the original submitter or a vendor_admin of same vendor
            is_submitter = sub.get("submitted_by") == getattr(user, "user_id", None)
            is_vendor_admin_of_submitter = False
            if role == "vendor_admin" and getattr(user, "vendor_id", None):
                submitter_user = await db.users.find_one(
                    {"user_id": sub.get("submitted_by")}, {"_id": 0, "vendor_id": 1}
                )
                if submitter_user and submitter_user.get("vendor_id") == user.vendor_id:
                    is_vendor_admin_of_submitter = True

            if not (is_submitter or is_vendor_admin_of_submitter):
                raise HTTPException(403, "You do not have permission to edit this submission")
            if not is_rejected:
                raise HTTPException(403, "Editing is only allowed after a rejection")

        # Fetch the template
        tpl = await db.pdf_templates.find_one(
            {"template_id": sub["template_id"], "is_deleted": False}, {"_id": 0}
        )
        if not tpl:
            raise HTTPException(404, "Form template not found")

        # Regenerate completed PDF with new values
        src = PDF_DIR / tpl["storage_filename"]
        out_path = COMPLETED_DIR / (sub.get("completed_filename") or f"{submission_id}.pdf")
        try:
            field_models = [PDFField(**f) for f in tpl.get("fields", [])]
            generate_completed_pdf(src, field_models, body.values, out_path,
                                   uploads_root=uploads_root, submission_id=submission_id)
        except Exception as e:
            logger.exception("PDF regeneration failed for edit")
            raise HTTPException(500, f"PDF generation failed: {e}")

        # Track changed fields
        old_values = sub.get("values") or {}
        changed_fields = [
            k for k in set(list(old_values.keys()) + list(body.values.keys()))
            if old_values.get(k) != body.values.get(k)
        ]

        now = _now()
        edit_history_entry = {
            "edited_by": getattr(user, "user_id", None),
            "edited_by_email": getattr(user, "email", None),
            "edited_by_name": getattr(user, "name", None),
            "edit_reason": body.edit_reason,
            "changed_fields": changed_fields,
            "edited_at": now,
        }

        # Update submission document
        await db.pdf_submissions.update_one(
            {"submission_id": submission_id},
            {"$set": {
                "values": body.values,
                "updated_at": now,
                "status": "submitted",          # reset status so workflow can re-evaluate
                "approval_status": None,
            }, "$inc": {"edit_count": 1},
             "$push": {"edit_history": edit_history_entry}},
        )

        # Write audit log
        await db.audit_logs.insert_one({
            "audit_id": _gen("aud"),
            "actor_id": getattr(user, "user_id", None),
            "actor_email": getattr(user, "email", None),
            "action": "submission.edit",
            "target_type": "pdf_submission",
            "target_id": submission_id,
            "details": {
                "edit_reason": body.edit_reason,
                "changed_fields": changed_fields,
                "form_name": tpl.get("title"),
                "submission_id": submission_id,
            },
            "ip": request.client.host if request.client else None,
            "created_at": now,
        })

        # Re-fire workflow trigger so approval restarts from beginning
        try:
            from workflow_routes import fire_trigger as _ft
            _site_name_val = None
            _SITE_LABELS = {"site name", "plant name", "site code", "asset id", "site_name",
                            "plant_name", "site_code", "asset_id"}
            for _f in tpl.get("fields", []):
                _label = (_f.get("label") or "").strip().lower()
                _fid = (_f.get("id") or "")
                if _label in _SITE_LABELS or _fid in {"site_name", "site_code", "plant_name", "asset_id"}:
                    _v = body.values.get(_fid)
                    if _v:
                        _site_name_val = str(_v)
                        break
            base_url = os.environ.get("PUBLIC_BASE_URL") or ""
            edit_url = f"{base_url}/p/{tpl.get('slug')}/edit/{submission_id}" if base_url else f"/p/{tpl.get('slug')}/edit/{submission_id}"
            await _ft(db, "pdf_submitted",
                      {"submission_id": submission_id,
                       "template_id": tpl["template_id"],
                       "form_name": tpl.get("title"),
                       "values": body.values,
                       "submission_kind": "pdf",
                       "user_id": getattr(user, "user_id", None),
                       "user_email": getattr(user, "email", None),
                       "site_name": _site_name_val,
                       "edit_submission": edit_url,
                       "is_resubmission": True,
                       "edit_reason": body.edit_reason,
                       "ip": request.client.host if request.client else None})
        except Exception as _e:
            logger.warning(f"workflow re-trigger after edit failed: {_e}")

        updated = await db.pdf_submissions.find_one({"submission_id": submission_id}, {"_id": 0})
        return updated

    return router, public, subs, pub_subs
