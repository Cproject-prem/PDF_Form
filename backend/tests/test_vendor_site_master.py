"""Backend tests for Vendor Management, Site Master, Master Data, Lookup Engine, RLS, and Workflow email attachments."""
import io
import os
import time
import pytest
import requests
from openpyxl import load_workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL", "https://pdf-merge-forms.preview.emergentagent.com").rstrip("/")
ADMIN_EMAIL = "admin@example.com"
ADMIN_PASSWORD = "Admin@12345"


@pytest.fixture(scope="session")
def admin_token():
    r = requests.post(f"{BASE_URL}/api/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD}, timeout=15)
    assert r.status_code == 200, r.text
    return r.json()["token"]


@pytest.fixture(scope="session")
def admin_headers(admin_token):
    return {"Authorization": f"Bearer {admin_token}", "Content-Type": "application/json"}


# ------------------------ Login regression ------------------------

class TestAuth:
    def test_admin_login(self):
        r = requests.post(f"{BASE_URL}/api/auth/login", json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD})
        assert r.status_code == 200
        d = r.json()
        assert d["user"]["email"] == ADMIN_EMAIL
        assert d["user"]["role"] == "super_admin"


# ------------------------ Site Master ------------------------

class TestSiteMaster:
    def test_list_sites_seeded(self, admin_headers):
        r = requests.get(f"{BASE_URL}/api/sites", headers=admin_headers)
        assert r.status_code == 200
        rows = r.json()
        names = {row.get("site_name") for row in rows}
        assert {"Alpha Solar 50MW", "Bravo Wind 30MW", "Charlie Hybrid 25MW"}.issubset(names), names

    def test_list_columns_23(self, admin_headers):
        r = requests.get(f"{BASE_URL}/api/sites/columns", headers=admin_headers)
        assert r.status_code == 200
        cols = r.json()
        # must contain all 23 core columns
        core_keys = [c["key"] for c in cols if c.get("core")]
        expected = ["site_name", "site_code", "asset_id", "plant_name", "customer_name",
                    "state", "district", "location", "latitude", "longitude",
                    "ac_capacity", "dc_capacity", "inverter_capacity",
                    "vendor_name", "vendor_login_user", "vendor_email",
                    "cluster", "region", "site_status",
                    "commission_date", "om_start_date", "warranty_end_date", "remarks"]
        assert len(core_keys) == 23, f"Expected 23 core columns, got {len(core_keys)}"
        for k in expected:
            assert k in core_keys, f"Missing column {k}"

    def test_add_and_remove_custom_column(self, admin_headers):
        # cleanup if present
        requests.delete(f"{BASE_URL}/api/sites/columns/ppa_expiry", headers=admin_headers)
        r = requests.post(f"{BASE_URL}/api/sites/columns", json={"label": "PPA Expiry"}, headers=admin_headers)
        assert r.status_code == 200, r.text
        cols = requests.get(f"{BASE_URL}/api/sites/columns", headers=admin_headers).json()
        assert any(c["key"] == "ppa_expiry" for c in cols)
        # Core cannot be removed
        r2 = requests.delete(f"{BASE_URL}/api/sites/columns/site_name", headers=admin_headers)
        assert r2.status_code == 400
        # Delete custom
        r3 = requests.delete(f"{BASE_URL}/api/sites/columns/ppa_expiry", headers=admin_headers)
        assert r3.status_code == 200

    def test_template_xlsx(self, admin_headers):
        r = requests.get(f"{BASE_URL}/api/sites/template.xlsx", headers=admin_headers)
        assert r.status_code == 200
        assert "spreadsheet" in r.headers.get("content-type", "")
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) >= 2  # header + 1 example
        assert "Site Name" in rows[0]

    def test_create_update_delete_site(self, admin_headers):
        body = {"site_name": "TEST Site CRUD", "site_code": "TEST-CRUD-1", "region": "North", "ac_capacity": 10}
        r = requests.post(f"{BASE_URL}/api/sites", json=body, headers=admin_headers)
        assert r.status_code == 200, r.text
        sid = r.json()["site_id"]
        # update
        r2 = requests.put(f"{BASE_URL}/api/sites/{sid}", json={"region": "South", "ac_capacity": 20}, headers=admin_headers)
        assert r2.status_code == 200
        assert r2.json()["version"] >= 2
        # delete
        r3 = requests.delete(f"{BASE_URL}/api/sites/{sid}", headers=admin_headers)
        assert r3.status_code == 200

    def test_bulk_upsert(self, admin_headers):
        body = {"rows": [
            {"site_name": "TEST Bulk A", "site_code": "TEST-BULK-A"},
            {"site_name": "TEST Bulk B", "site_code": "TEST-BULK-B"},
        ], "delete_missing": False}
        r = requests.post(f"{BASE_URL}/api/sites/bulk", json=body, headers=admin_headers)
        assert r.status_code == 200
        assert r.json()["upserted"] == 2
        # cleanup
        all_sites = requests.get(f"{BASE_URL}/api/sites", headers=admin_headers).json()
        ids = [s["site_id"] for s in all_sites if s.get("site_code", "").startswith("TEST-BULK")]
        r2 = requests.post(f"{BASE_URL}/api/sites/bulk-delete", json={"site_ids": ids}, headers=admin_headers)
        assert r2.status_code == 200

    def test_export_xlsx_csv(self, admin_headers):
        r1 = requests.get(f"{BASE_URL}/api/sites/export.xlsx", headers=admin_headers)
        assert r1.status_code == 200
        r2 = requests.get(f"{BASE_URL}/api/sites/export.csv", headers=admin_headers)
        assert r2.status_code == 200
        assert b"Site Name" in r2.content

    def test_csv_import(self, admin_headers):
        csv_data = "Site Name,Site Code,Region\nTEST Import 1,TEST-IMP-1,East\nTEST Import 2,TEST-IMP-2,West\n"
        files = {"file": ("import.csv", csv_data, "text/csv")}
        data = {"replace": "false"}
        h = {"Authorization": admin_headers["Authorization"]}
        r = requests.post(f"{BASE_URL}/api/sites/import", files=files, data=data, headers=h)
        assert r.status_code == 200, r.text
        assert r.json()["rows"] == 2
        # cleanup
        all_sites = requests.get(f"{BASE_URL}/api/sites", headers=admin_headers).json()
        ids = [s["site_id"] for s in all_sites if s.get("site_code", "").startswith("TEST-IMP")]
        if ids:
            requests.post(f"{BASE_URL}/api/sites/bulk-delete", json={"site_ids": ids}, headers=admin_headers)


# ------------------------ Lookup ------------------------

class TestLookup:
    def test_lookup_options(self, admin_headers):
        r = requests.get(f"{BASE_URL}/api/lookup/options?source=sites&column=site_name", headers=admin_headers)
        assert r.status_code == 200
        vals = r.json()
        assert "Alpha Solar 50MW" in vals
        assert vals == sorted(vals)

    def test_lookup_resolve(self, admin_headers):
        body = {"source": "sites", "display": "site_name", "return": "asset_id",
                "value": "Alpha Solar 50MW",
                "fill": ["ac_capacity", "dc_capacity", "vendor_name", "vendor_email", "region"]}
        r = requests.post(f"{BASE_URL}/api/lookup/resolve", json=body, headers=admin_headers)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["matched"] is True
        assert d["value"] == "AST-1001"
        assert d["fill"]["ac_capacity"] == 50
        assert d["fill"]["dc_capacity"] == 65
        assert d["fill"]["vendor_name"] == "SunOps Pvt Ltd"
        assert d["fill"]["vendor_email"] == "ops@sunops.example.com"
        assert d["fill"]["region"] == "South"

    def test_lookup_columns_23(self, admin_headers):
        r = requests.get(f"{BASE_URL}/api/lookup/columns?source=sites", headers=admin_headers)
        assert r.status_code == 200
        cols = r.json()
        core = [c for c in cols if c.get("core")]
        assert len(core) == 23

    def test_public_lookup_no_auth(self):
        r = requests.get(f"{BASE_URL}/api/public/lookup/options?source=sites&column=site_name")
        assert r.status_code == 200
        vals = r.json()
        assert "Alpha Solar 50MW" in vals

    def test_public_lookup_resolve_no_auth(self):
        body = {"source": "sites", "display": "site_name", "return": "asset_id",
                "value": "Alpha Solar 50MW", "fill": ["ac_capacity", "region"]}
        r = requests.post(f"{BASE_URL}/api/public/lookup/resolve", json=body)
        assert r.status_code == 200
        d = r.json()
        assert d["matched"] is True
        assert d["value"] == "AST-1001"
        assert d["fill"]["region"] == "South"


# ------------------------ Master Data ------------------------

class TestMasterData:
    def test_list_tables_builtin(self, admin_headers):
        r = requests.get(f"{BASE_URL}/api/master-data/tables", headers=admin_headers)
        assert r.status_code == 200
        tables = {t["table"] for t in r.json()}
        for b in ["customers", "regions", "states", "departments", "products", "categories"]:
            assert b in tables

    def test_master_crud(self, admin_headers):
        # ADD
        r = requests.post(f"{BASE_URL}/api/master-data/customers",
                          json={"name": "TEST Customer Acme", "code": "ACME-T"}, headers=admin_headers)
        assert r.status_code == 200
        row_id = r.json()["row_id"]
        # UPDATE
        r2 = requests.put(f"{BASE_URL}/api/master-data/{row_id}",
                          json={"name": "TEST Customer Acme Updated", "code": "ACME-T2"},
                          headers=admin_headers)
        assert r2.status_code == 200
        assert r2.json()["version"] == 2
        # DELETE
        r3 = requests.delete(f"{BASE_URL}/api/master-data/{row_id}", headers=admin_headers)
        assert r3.status_code == 200


# ------------------------ Vendor + Vendor User + RLS ------------------------

class TestVendorAndRLS:
    vendor_id = None
    vendor_user_id = None
    vendor_email = f"TEST_vendor_{int(time.time())}@example.com"
    vendor_password = "VendorPass@123"
    vendor_token = None
    form_id = None
    other_form_id = None

    def test_create_vendor(self, admin_headers):
        body = {"name": "TEST Vendor LLC", "code": "VEN-T", "email": "vendor@example.com"}
        r = requests.post(f"{BASE_URL}/api/vendors", json=body, headers=admin_headers)
        assert r.status_code == 200
        TestVendorAndRLS.vendor_id = r.json()["vendor_id"]

    def test_list_vendors_admin(self, admin_headers):
        r = requests.get(f"{BASE_URL}/api/vendors", headers=admin_headers)
        assert r.status_code == 200
        ids = {v["vendor_id"] for v in r.json()}
        assert TestVendorAndRLS.vendor_id in ids

    def test_get_update_vendor(self, admin_headers):
        vid = TestVendorAndRLS.vendor_id
        r = requests.get(f"{BASE_URL}/api/vendors/{vid}", headers=admin_headers)
        assert r.status_code == 200
        assert "_stats" in r.json()
        r2 = requests.put(f"{BASE_URL}/api/vendors/{vid}",
                          json={"name": "TEST Vendor LLC v2", "code": "VEN-T"}, headers=admin_headers)
        assert r2.status_code == 200
        assert r2.json()["name"] == "TEST Vendor LLC v2"

    def test_create_vendor_user(self, admin_headers):
        vid = TestVendorAndRLS.vendor_id
        body = {"email": TestVendorAndRLS.vendor_email, "name": "Test Vendor User",
                "password": TestVendorAndRLS.vendor_password, "role": "vendor"}
        r = requests.post(f"{BASE_URL}/api/vendor-users/{vid}", json=body, headers=admin_headers)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["role"] == "vendor"
        assert d["vendor_id"] == vid
        assert d["assignments"] == {"forms": [], "pdf_forms": [], "sites": [], "workflows": []}
        TestVendorAndRLS.vendor_user_id = d["user_id"]

    def test_vendor_user_login(self):
        r = requests.post(f"{BASE_URL}/api/auth/login",
                          json={"email": TestVendorAndRLS.vendor_email, "password": TestVendorAndRLS.vendor_password})
        assert r.status_code == 200, r.text
        TestVendorAndRLS.vendor_token = r.json()["token"]
        assert r.json()["user"]["role"] == "vendor"

    def test_setup_forms_for_rls(self, admin_headers):
        # Create two forms as admin
        r1 = requests.post(f"{BASE_URL}/api/forms", json={"title": "TEST RLS Form A", "status": "published"},
                           headers=admin_headers)
        assert r1.status_code == 200
        TestVendorAndRLS.form_id = r1.json()["form_id"]
        r2 = requests.post(f"{BASE_URL}/api/forms", json={"title": "TEST RLS Form B (not assigned)", "status": "published"},
                           headers=admin_headers)
        assert r2.status_code == 200
        TestVendorAndRLS.other_form_id = r2.json()["form_id"]

    def test_vendor_rls_empty_initially(self):
        h = {"Authorization": f"Bearer {TestVendorAndRLS.vendor_token}"}
        r = requests.get(f"{BASE_URL}/api/forms", headers=h)
        assert r.status_code == 200
        assert r.json() == []

    def test_assign_form_and_verify_rls(self, admin_headers):
        uid = TestVendorAndRLS.vendor_user_id
        r = requests.put(f"{BASE_URL}/api/vendor-users/{uid}/assignments",
                        json={"forms": [TestVendorAndRLS.form_id], "pdf_forms": [], "sites": [], "workflows": []},
                        headers=admin_headers)
        assert r.status_code == 200
        # Now vendor should see exactly this form
        h = {"Authorization": f"Bearer {TestVendorAndRLS.vendor_token}"}
        r2 = requests.get(f"{BASE_URL}/api/forms", headers=h)
        assert r2.status_code == 200
        ids = [f["form_id"] for f in r2.json()]
        assert ids == [TestVendorAndRLS.form_id]

    def test_vendor_cannot_get_other_form(self):
        h = {"Authorization": f"Bearer {TestVendorAndRLS.vendor_token}"}
        r = requests.get(f"{BASE_URL}/api/forms/{TestVendorAndRLS.other_form_id}", headers=h)
        assert r.status_code == 403

    def test_vendor_list_vendors_only_own(self):
        h = {"Authorization": f"Bearer {TestVendorAndRLS.vendor_token}"}
        r = requests.get(f"{BASE_URL}/api/vendors", headers=h)
        assert r.status_code == 200
        rows = r.json()
        assert len(rows) == 1
        assert rows[0]["vendor_id"] == TestVendorAndRLS.vendor_id

    def test_patch_vendor_user(self, admin_headers):
        uid = TestVendorAndRLS.vendor_user_id
        r = requests.patch(f"{BASE_URL}/api/vendor-users/{uid}",
                           json={"name": "Renamed Vendor User", "is_active": True}, headers=admin_headers)
        assert r.status_code == 200
        assert r.json()["name"] == "Renamed Vendor User"

    def test_cleanup(self, admin_headers):
        # delete forms
        for fid in [TestVendorAndRLS.form_id, TestVendorAndRLS.other_form_id]:
            if fid:
                requests.delete(f"{BASE_URL}/api/forms/{fid}", headers=admin_headers)
        # delete vendor user
        if TestVendorAndRLS.vendor_user_id:
            r = requests.delete(f"{BASE_URL}/api/vendor-users/{TestVendorAndRLS.vendor_user_id}", headers=admin_headers)
            assert r.status_code == 200
        # delete vendor
        if TestVendorAndRLS.vendor_id:
            r = requests.delete(f"{BASE_URL}/api/vendors/{TestVendorAndRLS.vendor_id}", headers=admin_headers)
            assert r.status_code == 200


# ------------------------ Workflow Email Attachment smoke ------------------------

class TestWorkflowAttachment:
    def test_create_workflow_with_completed_pdf_attachment(self, admin_headers):
        # Smoke: create a workflow with action.send_email node carrying attachments=['completed_pdf']
        wf_body = {
            "name": "TEST Attachment Workflow",
            "description": "smoke",
            "trigger": {"type": "pdf_submitted", "filter": {}},
            "nodes": [
                {"id": "n1", "type": "trigger.pdf_submitted", "config": {"filter": {}},
                 "position": {"x": 0, "y": 0}, "next": ["n2"]},
                {"id": "n2", "type": "action.send_email",
                 "config": {"to": "demo@example.com", "subject": "PDF", "body": "see attached",
                            "attachments": ["completed_pdf"]},
                 "position": {"x": 200, "y": 0}, "next": []},
            ],
            "status": "draft",
        }
        r = requests.post(f"{BASE_URL}/api/workflows", json=wf_body, headers=admin_headers)
        assert r.status_code in (200, 201), r.text
        wid = r.json().get("workflow_id") or r.json().get("id")
        assert wid
        # verify node config persisted
        r2 = requests.get(f"{BASE_URL}/api/workflows/{wid}", headers=admin_headers)
        assert r2.status_code == 200
        nodes = r2.json().get("nodes", [])
        email_node = next((n for n in nodes if n["type"] == "action.send_email"), None)
        assert email_node is not None
        assert "completed_pdf" in (email_node.get("config", {}).get("attachments") or [])
        # cleanup
        requests.delete(f"{BASE_URL}/api/workflows/{wid}", headers=admin_headers)
