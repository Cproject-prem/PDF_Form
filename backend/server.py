"""
Jotform-Clone Backend (FastAPI + MongoDB)

Routes are all prefixed with /api. Auth uses JWT (HS256) for primary login
and supports Emergent Google OAuth as a secondary login path. Files are
stored via the Emergent Object Storage integration.
"""
from fastapi import FastAPI, APIRouter, HTTPException, Depends, UploadFile, File, Request, Response
from fastapi.responses import FileResponse, JSONResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr, ConfigDict
from typing import List, Optional, Any, Dict
from datetime import datetime, timezone, timedelta
from pathlib import Path
import os
import uuid
import logging
import bcrypt
import jwt
import io
import csv
import requests
import mimetypes
import re

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

logger = logging.getLogger("jotform")
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')

# ---------- Config ----------
MONGO_URL = os.environ['MONGO_URL']
DB_NAME = os.environ['DB_NAME']
JWT_SECRET = os.environ.get('JWT_SECRET', 'dev-secret')

# ---------- Startup security enforcement ----------
from security import (  # noqa: E402
    enforce_startup_security,
    LOGIN_LIMITER,
    PUBLIC_SUBMIT_LIMITER,
    UPLOAD_LIMITER,
    check_rate_limit,
    _client_ip,
    validate_password_strength,
    validate_upload,
    SecurityHeadersMiddleware,
)
enforce_startup_security(dict(os.environ))
JWT_ALGO = 'HS256'
JWT_EXPIRE_HOURS = int(os.environ.get('JWT_EXPIRE_HOURS', 168))
APP_NAME = os.environ.get('APP_NAME', 'jotform-clone')
EMERGENT_LLM_KEY = os.environ.get('EMERGENT_LLM_KEY')
MAX_UPLOAD_MB = int(os.environ.get('MAX_UPLOAD_MB', 25))
SEED_ADMIN_EMAIL = os.environ.get('SEED_ADMIN_EMAIL', 'admin@local.test')
SEED_ADMIN_PASSWORD = os.environ.get('SEED_ADMIN_PASSWORD', 'Admin@12345')
SEED_ADMIN_NAME = os.environ.get('SEED_ADMIN_NAME', 'Super Admin')

ROLES = ['super_admin', 'admin', 'vendor_admin', 'vendor_user', 'member', 'user', 'vendor']

# ---------- DB ----------
client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

# ---------- Object Storage (local disk) ----------
# Files land in /app/backend/uploads/local/tmp/ at upload time (no submission
# context yet). When a submission is created we move any referenced files to
# /app/backend/uploads/local/submissions/{sid}/ so every submission has its
# own folder on disk — see _organize_submission_files() below.
LOCAL_UPLOAD_ROOT = Path(os.environ.get("LOCAL_UPLOAD_ROOT", "/app/backend/uploads/local"))
LOCAL_TMP_DIR = LOCAL_UPLOAD_ROOT / "tmp"
LOCAL_SUB_DIR = LOCAL_UPLOAD_ROOT / "submissions"
for _d in (LOCAL_UPLOAD_ROOT, LOCAL_TMP_DIR, LOCAL_SUB_DIR):
    _d.mkdir(parents=True, exist_ok=True)


def put_object(path: str, data: bytes, content_type: str) -> dict:
    """Store bytes on local disk under LOCAL_UPLOAD_ROOT.
    `path` is a slash-separated relative path (e.g. 'tmp/abc.png' or
    'submissions/sub_x/logo.png') — we keep the same argument shape as the
    old S3-style helper so callers don't have to change.
    """
    del content_type  # not stored — we rely on the filename extension
    safe = path.lstrip("/")
    target = LOCAL_UPLOAD_ROOT / safe
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(data)
    return {"path": str(target.relative_to(LOCAL_UPLOAD_ROOT)), "size": len(data)}


def get_object(path: str) -> tuple:
    """Read bytes back from local disk. Returns (data, content_type).
    Also tolerates legacy Emergent-storage paths like
    `formforge/uploads/{user}/{uuid}.ext` — we hunt for the file's basename
    under `local/`, `pdf/` and `completed/` before giving up."""
    safe = path.lstrip("/")
    target = LOCAL_UPLOAD_ROOT / safe
    if target.exists() and target.is_file():
        ct = mimetypes.guess_type(target.name)[0] or "application/octet-stream"
        return target.read_bytes(), ct

    # Fallback: try to find by basename in known locations
    basename = os.path.basename(safe)
    if basename:
        search_dirs = [
            LOCAL_UPLOAD_ROOT / "tmp",
            LOCAL_UPLOAD_ROOT / "submissions",
            Path(str(LOCAL_UPLOAD_ROOT).replace("/local", "")) / "pdf",
            Path(str(LOCAL_UPLOAD_ROOT).replace("/local", "")) / "completed",
        ]
        for d in search_dirs:
            if not d.exists():
                continue
            # rglob so we descend into submissions/{sid}/
            for candidate in d.rglob(basename):
                if candidate.is_file():
                    ct = mimetypes.guess_type(candidate.name)[0] or "application/octet-stream"
                    return candidate.read_bytes(), ct

    raise HTTPException(status_code=404, detail=f"File missing on disk: {path}")


async def _organize_submission_files(sid: str, values: Dict[str, Any]) -> None:
    """Walk submission values, find any `{file_id, ...}` refs and physically
    move the underlying file into `local/submissions/{sid}/`. Updates the
    db.files record so later downloads read from the new location. Also
    remembers which files belong to which submission via `submission_id`
    on the file document (handy for cleanup + auditing)."""

    def _collect_ids(node: Any, out: List[str]) -> None:
        if isinstance(node, dict):
            fid = node.get("file_id")
            if isinstance(fid, str) and fid:
                out.append(fid)
            for v in node.values():
                _collect_ids(v, out)
        elif isinstance(node, list):
            for item in node:
                _collect_ids(item, out)

    file_ids: List[str] = []
    _collect_ids(values or {}, file_ids)
    if not file_ids:
        return

    dest_dir = LOCAL_SUB_DIR / sid
    dest_dir.mkdir(parents=True, exist_ok=True)

    async for rec in db.files.find({"file_id": {"$in": list(set(file_ids))}}):
        old_rel = rec.get("storage_path") or ""
        old_path = LOCAL_UPLOAD_ROOT / old_rel
        if not old_path.exists():
            continue
        # Preserve the original filename in the per-submission folder; if a
        # collision happens (rare — same name uploaded twice) fall back to
        # the file_id prefix so nothing gets clobbered.
        orig = rec.get("original_filename") or old_path.name
        safe_name = re.sub(r"[^\w.\-]+", "_", orig)[:120] or old_path.name
        target = dest_dir / safe_name
        if target.exists() and target.resolve() != old_path.resolve():
            stem, dot, ext = safe_name.rpartition(".")
            safe_name = f"{stem or safe_name}_{rec['file_id'][:6]}{dot}{ext}" if dot else f"{safe_name}_{rec['file_id'][:6]}"
            target = dest_dir / safe_name
        try:
            old_path.rename(target)
        except OSError:
            # cross-device or already-moved — fall back to copy + best-effort delete
            target.write_bytes(old_path.read_bytes())
            try:
                old_path.unlink(missing_ok=True)
            except Exception:
                pass
        new_rel = str(target.relative_to(LOCAL_UPLOAD_ROOT))
        await db.files.update_one(
            {"file_id": rec["file_id"]},
            {"$set": {"storage_path": new_rel, "submission_id": sid,
                      "organized_at": datetime.now(timezone.utc).isoformat()}},
        )

# ---------- Models ----------
class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    user_id: str
    email: EmailStr
    name: str
    role: str = "user"
    picture: Optional[str] = None
    is_active: bool = True
    created_at: str
    password_hash: Optional[str] = None  # not exposed via API
    vendor_id: Optional[str] = None  # set for vendor users (RLS scope)
    cluster_manager_name: Optional[str] = None  # for admin role — links to Site.cluster_manager_name
    region: Optional[str] = None  # for admin role — links to Site.region (regional access scope)
    assignments: Optional[Dict[str, List[str]]] = None  # {forms:[], pdf_forms:[], sites:[], workflows:[]}
    access_override: bool = False  # add-on flag → grants Super-Admin-level access

class UserOut(BaseModel):
    user_id: str
    email: str
    name: str
    role: str
    picture: Optional[str] = None
    is_active: bool = True
    created_at: str
    vendor_id: Optional[str] = None
    cluster_manager_name: Optional[str] = None
    region: Optional[str] = None
    assignments: Optional[Dict[str, List[str]]] = None
    access_override: bool = False

class RegisterIn(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str = Field(min_length=1)

class LoginIn(BaseModel):
    email: EmailStr
    password: str

class GoogleSessionIn(BaseModel):
    session_id: str

class FormField(BaseModel):
    model_config = ConfigDict(extra="allow")  # allow data_source, lookup, formula, etc.
    id: str
    type: str                 # short_text, long_text, number, email, phone, date, time,
                              # dropdown, checkbox, radio, file, url, rating, heading,
                              # paragraph, divider
    label: str = ""
    placeholder: str = ""
    description: str = ""
    required: bool = False
    read_only: bool = False
    default_value: Any = None
    options: List[str] = []   # for dropdown/checkbox/radio
    validation: Dict[str, Any] = {}  # {min,max,regex,maxLength}
    width: str = "full"       # full | half
    rich_text: str = ""       # for heading/paragraph

class FormTheme(BaseModel):
    primary_color: str = "#2563EB"
    background: str = "#FFFFFF"
    font: str = "Outfit"

class FormSettings(BaseModel):
    submission_limit: Optional[int] = None
    require_login: bool = False
    allow_multiple: bool = True
    show_progress: bool = False
    thank_you_message: str = "Thanks for your submission!"
    redirect_url: Optional[str] = None
    notify_emails: List[str] = []

class FormIn(BaseModel):
    title: str = "Untitled Form"
    description: str = ""
    fields: List[FormField] = []
    theme: FormTheme = FormTheme()
    settings: FormSettings = FormSettings()
    status: str = "draft"     # draft | published | archived
    # Optional per-form filename template used when the submitter (or an
    # admin) downloads a filled submission. Placeholders: {form_name},
    # {asset_id}, {submitter_name}, {datetime}. Left empty → resolver uses
    # the global default `{asset_id}_{submitter_name}_{datetime}`.
    filename_template: Optional[str] = ""
    # ---- Assignment fields (row-level security) ----
    assigned_site_ids: List[str] = []
    assigned_vendor_ids: List[str] = []
    assigned_vendor_user_ids: List[str] = []
    assigned_admin_ids: List[str] = []
    assigned_member_ids: List[str] = []
    assigned_department_ids: List[str] = []
    assigned_team_ids: List[str] = []
    assigned_cluster_managers: List[str] = []
    assigned_regions: List[str] = []

class Form(FormIn):
    form_id: str
    slug: str
    owner_id: str
    created_at: str
    updated_at: str
    is_favorite: bool = False
    is_archived: bool = False
    is_deleted: bool = False
    # Enrichment fields (populated by list_forms so the UI can render an
    # "owner" chip without a follow-up round-trip).
    owner_name: Optional[str] = None
    owner_email: Optional[str] = None

class SubmissionIn(BaseModel):
    values: Dict[str, Any]    # {field_id: value}

class Submission(BaseModel):
    submission_id: str
    form_id: str
    values: Dict[str, Any]
    submitted_by: Optional[str] = None
    ip: Optional[str] = None
    user_agent: Optional[str] = None
    status: str = "submitted"  # submitted | approved | rejected
    created_at: str

# ---------- App ----------
app = FastAPI(title="Jotform Clone API")
api = APIRouter(prefix="/api")

@app.on_event("startup")
async def startup():
    # indexes
    await db.users.create_index("email", unique=True)
    await db.users.create_index("user_id", unique=True)
    await db.forms.create_index("form_id", unique=True)
    await db.forms.create_index("slug", unique=True)
    await db.submissions.create_index("submission_id", unique=True)
    await db.submissions.create_index("form_id")
    # PDF Form Builder indexes
    await db.pdf_templates.create_index("template_id", unique=True)
    await db.pdf_templates.create_index("slug", unique=True)
    await db.pdf_submissions.create_index("submission_id", unique=True)
    await db.pdf_submissions.create_index("template_id")
    # Workflow indexes
    await db.workflows.create_index("workflow_id", unique=True)
    await db.workflow_executions.create_index("execution_id", unique=True)
    await db.workflow_executions.create_index("workflow_id")
    await db.approvals.create_index("approval_id", unique=True)
    await db.approval_tokens.create_index("token", unique=True)
    await db.audit_logs.create_index("created_at")
    # seed workflow templates (idempotent)
    try:
        from workflow_routes import seed_workflow_templates as _seed_wf
        await _seed_wf(db)
    except Exception as e:  # noqa: BLE001
        logger.warning(f"Workflow template seed skipped: {e}")
    # seed demo sites (idempotent — does nothing if any site exists)
    try:
        from vendor_routes import seed_demo_sites as _seed_sites
        await _seed_sites(db)
    except Exception as e:  # noqa: BLE001
        logger.warning(f"Site demo seed skipped: {e}")
    # additional indexes for vendor / site / master
    await db.vendors.create_index("vendor_id", unique=True)
    await db.sites.create_index("site_id", unique=True)
    await db.sites.create_index("site_code")
    await db.master_data.create_index([("table", 1), ("row_id", 1)])
    # seed super admin (idempotent — resets password to SEED_ADMIN_PASSWORD
    # on every boot so operators can rotate credentials by editing .env, and
    # re-activates a manually-disabled admin so lockouts are recoverable).
    existing = await db.users.find_one({"email": SEED_ADMIN_EMAIL.lower()}, {"_id": 0})
    if not existing:
        uid = f"user_{uuid.uuid4().hex[:12]}"
        await db.users.insert_one({
            "user_id": uid,
            "email": SEED_ADMIN_EMAIL.lower(),
            "name": SEED_ADMIN_NAME,
            "role": "super_admin",
            "password_hash": bcrypt.hashpw(SEED_ADMIN_PASSWORD.encode(), bcrypt.gensalt()).decode(),
            "picture": None,
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
        logger.info(f"Seeded super admin: {SEED_ADMIN_EMAIL}")
    else:
        updates: Dict[str, Any] = {}
        # Reset password_hash whenever the .env password no longer matches
        # (covers rotations, broken hashes, and lost passwords).
        stored_hash = (existing.get("password_hash") or "").encode()
        matches = False
        try:
            matches = bool(stored_hash) and bcrypt.checkpw(
                SEED_ADMIN_PASSWORD.encode(), stored_hash)
        except (ValueError, TypeError):
            matches = False
        if not matches:
            updates["password_hash"] = bcrypt.hashpw(
                SEED_ADMIN_PASSWORD.encode(), bcrypt.gensalt()).decode()
        # Re-activate disabled admin so lockouts are always recoverable.
        if not existing.get("is_active", True):
            updates["is_active"] = True
        # Guarantee super_admin role — protects against accidental demotion.
        if existing.get("role") != "super_admin":
            updates["role"] = "super_admin"
        if updates:
            await db.users.update_one(
                {"email": SEED_ADMIN_EMAIL.lower()}, {"$set": updates})
            logger.info(
                f"Refreshed super admin {SEED_ADMIN_EMAIL} fields={list(updates.keys())}"
            )

    # Seed demo accounts for the four-role permission model (idempotent —
    # resets password_hash and re-activates the account on every boot).
    async def _ensure(email: str, name: str, role: str, password: str, **extra):
        existing = await db.users.find_one({"email": email})
        pwd_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
        if existing is None:
            uid = f"user_{uuid.uuid4().hex[:12]}"
            await db.users.insert_one({
                "user_id": uid, "email": email, "name": name, "role": role,
                "password_hash": pwd_hash,
                "picture": None, "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat(),
                **extra,
            })
            logger.info(f"Seeded {role}: {email}")
            return
        # Rehash only if the stored hash no longer matches the current password.
        stored_hash = (existing.get("password_hash") or "").encode()
        matches = False
        try:
            matches = bool(stored_hash) and bcrypt.checkpw(password.encode(), stored_hash)
        except (ValueError, TypeError):
            matches = False
        updates: Dict[str, Any] = {}
        if not matches:
            updates["password_hash"] = pwd_hash
        if not existing.get("is_active", True):
            updates["is_active"] = True
        if updates:
            await db.users.update_one({"email": email}, {"$set": updates})
            logger.info(f"Refreshed {role} {email} fields={list(updates.keys())}")
    # Cluster-Manager admin scoped to Rahul Verma (sees Alpha + Bravo only)
    await _ensure("rahul.verma@example.com", "Rahul Verma (Cluster Mgr)",
                  "admin", "Admin@12345", cluster_manager_name="Rahul Verma")
    # Regional admin scoped to "South" region — sees Alpha + Bravo + any other South plants
    await _ensure("south.admin@example.com", "South Regional Admin",
                  "admin", "Admin@12345", region="South")
    # Vendor admin for SunOps (sees Alpha + Charlie only)
    sunops_vid = "ven_sunops_demo"
    if not await db.vendors.find_one({"vendor_id": sunops_vid}):
        await db.vendors.insert_one({
            "vendor_id": sunops_vid, "vendor_name": "SunOps Pvt Ltd",
            "vendor_email": "ops@sunops.example.com",
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    # Backfill sites with vendor_id where vendor_email matches
    await db.sites.update_many({"vendor_email": "ops@sunops.example.com", "vendor_id": {"$exists": False}},
                               {"$set": {"vendor_id": sunops_vid}})
    await _ensure("vendor.admin@sunops.example.com", "SunOps Vendor Admin",
                  "vendor_admin", "Vendor@12345", vendor_id=sunops_vid)
    await _ensure("vendor.user@sunops.example.com", "SunOps Vendor User",
                  "vendor_user", "Vendor@12345", vendor_id=sunops_vid)

@app.on_event("shutdown")
async def shutdown():
    client.close()

# ---------- Auth helpers ----------
def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()

def check_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False

def make_token(user_id: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRE_HOURS),
        "iat": datetime.now(timezone.utc),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)

def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")


# ---------- Download tokens (short-lived, submission-scoped) ----------
DOWNLOAD_TOKEN_HOURS = int(os.environ.get('DOWNLOAD_TOKEN_HOURS', 24))

def make_download_token(submission_id: str, kind: str) -> str:
    """Create a JWT scoped to downloading a single submission's PDF.
    kind = "form" for standard-form submissions, "pdf" for PDF-form submissions.
    """
    payload = {
        "scope": "download",
        "kind": kind,
        "sid": submission_id,
        "exp": datetime.now(timezone.utc) + timedelta(hours=DOWNLOAD_TOKEN_HOURS),
        "iat": datetime.now(timezone.utc),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)


def verify_download_token(token: str, submission_id: str, kind: str) -> None:
    """Raise 401/403 unless token is a valid download token for this submission."""
    data = decode_token(token)
    if data.get("scope") != "download":
        raise HTTPException(status_code=403, detail="Invalid download token scope")
    if data.get("kind") != kind:
        raise HTTPException(status_code=403, detail="Invalid download token kind")
    if data.get("sid") != submission_id:
        raise HTTPException(status_code=403, detail="Token does not match submission")

bearer = HTTPBearer(auto_error=False)

async def get_current_user(creds: HTTPAuthorizationCredentials = Depends(bearer)) -> User:
    if not creds or not creds.credentials:
        raise HTTPException(status_code=401, detail="Not authenticated")
    data = decode_token(creds.credentials)
    user = await db.users.find_one({"user_id": data["sub"]}, {"_id": 0, "password_hash": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    if not user.get("is_active", True):
        raise HTTPException(status_code=403, detail="User disabled")
    return User(**user)

async def get_optional_user(creds: HTTPAuthorizationCredentials = Depends(bearer)) -> Optional[User]:
    if not creds or not creds.credentials:
        return None
    try:
        data = decode_token(creds.credentials)
        user = await db.users.find_one({"user_id": data["sub"]}, {"_id": 0, "password_hash": 0})
        return User(**user) if user else None
    except Exception:
        return None

def require_role(*roles):
    async def _dep(user: User = Depends(get_current_user)):
        if user.role not in roles:
            raise HTTPException(status_code=403, detail="Forbidden")
        return user
    return _dep

# ---------- Auth routes ----------
@api.post("/auth/register")
async def register(body: RegisterIn):
    email = body.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=409, detail="Email already registered")
    uid = f"user_{uuid.uuid4().hex[:12]}"
    user_doc = {
        "user_id": uid,
        "email": email,
        "name": body.name,
        "role": "user",
        "password_hash": hash_password(body.password),
        "picture": None,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(user_doc)
    token = make_token(uid, "user")
    return {"token": token, "user": UserOut(**{k: v for k, v in user_doc.items() if k != "password_hash"})}

@api.post("/auth/login")
async def login(body: LoginIn, request: Request):
    ip = _client_ip(request)
    email = body.email.lower()
    # Rate limit BOTH by IP and by (IP + email) so distributed guessing is
    # bounded and per-account guessing is bounded.
    check_rate_limit(LOGIN_LIMITER, f"login:ip:{ip}",        "login")
    check_rate_limit(LOGIN_LIMITER, f"login:email:{email}",  "login")
    user = await db.users.find_one({"email": email})
    if not user or not user.get("password_hash") or not check_password(body.password, user["password_hash"]):
        logger.warning(f"failed_login email={email} ip={ip}")
        raise HTTPException(status_code=401, detail="Invalid credentials")
    if not user.get("is_active", True):
        raise HTTPException(status_code=403, detail="User disabled")
    # Successful login → reset that email's window so a legit user isn't
    # locked out by their own prior typos.
    LOGIN_LIMITER.reset(f"login:email:{email}")
    token = make_token(user["user_id"], user["role"])
    user.pop("password_hash", None)
    user.pop("_id", None)
    logger.info(f"login_success email={email} role={user.get('role')} ip={ip}")
    return {"token": token, "user": UserOut(**user)}

@api.post("/auth/google/session")
async def google_session(body: GoogleSessionIn):
    # Exchange session_id with Emergent Auth and create/update user
    try:
        r = requests.get(
            "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
            headers={"X-Session-ID": body.session_id}, timeout=15
        )
        r.raise_for_status()
        info = r.json()
    except Exception as e:
        raise HTTPException(status_code=401, detail=f"Google session exchange failed: {e}")
    email = info["email"].lower()
    user = await db.users.find_one({"email": email}, {"_id": 0})
    if not user:
        uid = f"user_{uuid.uuid4().hex[:12]}"
        user = {
            "user_id": uid,
            "email": email,
            "name": info.get("name", email.split("@")[0]),
            "role": "user",
            "picture": info.get("picture"),
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "password_hash": None,
        }
        await db.users.insert_one(user)
    else:
        await db.users.update_one({"user_id": user["user_id"]}, {"$set": {"picture": info.get("picture")}})
        user["picture"] = info.get("picture")
    user.pop("password_hash", None)
    token = make_token(user["user_id"], user["role"])
    return {"token": token, "user": UserOut(**user)}

@api.get("/auth/me", response_model=UserOut)
async def me(user: User = Depends(get_current_user)):
    return UserOut(**user.model_dump())


@api.get("/auth/menu")
async def auth_menu(user: User = Depends(get_current_user)):
    """Return the sidebar menu (with group segregation) + capability matrix."""
    from permissions import capabilities_for, menu_for, normalize_role, MENU_GROUPS
    items = menu_for(user)
    # Only include groups that actually have at least one visible item
    used = {i.get("group") for i in items}
    groups = [g for g in MENU_GROUPS if g["key"] in used]
    return {
        "menu": items,
        "groups": groups,
        "capabilities": capabilities_for(user),
        "role": normalize_role(user.role),
    }


# ---------- Global submissions list (admin / vendor admin views) ----------
@api.get("/submissions", response_model=List[Submission])
async def list_all_submissions(user: User = Depends(get_current_user), q: Optional[str] = None,
                               status: Optional[str] = None, form_id: Optional[str] = None):
    """List submissions visible to the current user.

      super_admin → all
      admin       → submissions whose values reference a site the admin
                    has access to (region / cluster / assigned).  If the
                    admin has no region/cluster set, they see all submissions.
      vendor_admin→ submissions from their vendor's users
      vendor_user → only their own submissions
    """
    from permissions import normalize_role, async_submission_filter
    role = normalize_role(user.role)
    query: Dict[str, Any] = await async_submission_filter(db, user)

    if role == "vendor_admin":
        vid = user.vendor_id
        if vid:
            team = await db.users.find({"vendor_id": vid}, {"_id": 0, "user_id": 1}).to_list(2000)
            query = {"submitted_by": {"$in": [u["user_id"] for u in team]}}
        else:
            return []
    if form_id:
        query = {"$and": [query, {"form_id": form_id}]} if query else {"form_id": form_id}
    if status:
        query = {"$and": [query, {"status": status}]} if query else {"status": status}
    if q:
        query = {"$and": [query, {"values": {"$regex": q, "$options": "i"}}]} if query else {"values": {"$regex": q, "$options": "i"}}
    rows = await db.submissions.find(query, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return [Submission(**r) for r in rows]


# ---------- Consolidated submissions overview (all forms + PDF forms, grouped) ----------
@api.get("/submissions/overview")
async def submissions_overview(user: User = Depends(get_current_user)):
    """Return submissions grouped form-wise across BOTH standard forms and PDF forms.

    Access model:
      * Every admin sees EVERY form & PDF template definition.
      * Submissions inside each group are further narrowed by
        `async_submission_filter` — an admin only sees rows for sites they
        can access (region / cluster / assigned). Global admins (no region,
        no cluster) see every submission.
      * vendor_admin sees vendor-team submissions; vendor_user sees own.

    Response shape:
      [{ kind: "form"|"pdf", form_id, title, slug, count, submissions: [...] }]
    """
    from permissions import (
        normalize_role, form_filter, is_super_admin, async_submission_filter,
    )
    role = normalize_role(user.role)
    groups: List[Dict[str, Any]] = []

    # --- Standard forms — every admin/super_admin sees all --------------
    if is_super_admin(user) or role == "admin":
        form_q: Dict[str, Any] = {}
    else:
        form_q = form_filter(user)
    forms = await db.forms.find(form_q, {"_id": 0}).sort("updated_at", -1).to_list(2000)

    # Restrict which subs a user can see
    sub_q_extra: Dict[str, Any] = {}
    if role == "vendor_user":
        sub_q_extra = {"submitted_by": user.user_id}
    elif role == "vendor_admin" and user.vendor_id:
        team = await db.users.find(
            {"vendor_id": user.vendor_id}, {"_id": 0, "user_id": 1},
        ).to_list(2000)
        sub_q_extra = {"submitted_by": {"$in": [u["user_id"] for u in team]}}
    elif role == "admin":
        # region/cluster-scoped submission filter (empty for global admins)
        sub_q_extra = await async_submission_filter(db, user)

    async def _enrich(subs: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Attach submitter + vendor display fields to each row so the
        admin UI can render a full 'Submitter' card. Non-admin viewers get
        the same enrichment — RLS already limits which rows they see, so
        showing the submitter's name inside those rows is fine."""
        if not subs:
            return subs
        uid_list = list({s.get("submitted_by") for s in subs if s.get("submitted_by")})
        users_by_id: Dict[str, Dict[str, Any]] = {}
        if uid_list:
            async for u in db.users.find(
                {"user_id": {"$in": uid_list}},
                {"_id": 0, "user_id": 1, "email": 1, "name": 1, "role": 1, "vendor_id": 1, "region": 1},
            ):
                users_by_id[u["user_id"]] = u
        vendor_ids = list({(users_by_id.get(u) or {}).get("vendor_id") for u in uid_list})
        vendor_ids += [s.get("vendor_id") for s in subs if s.get("vendor_id")]
        vendor_ids = [v for v in set(vendor_ids) if v]
        vendors_by_id: Dict[str, str] = {}
        if vendor_ids:
            async for v in db.vendors.find(
                {"vendor_id": {"$in": vendor_ids}},
                {"_id": 0, "vendor_id": 1, "vendor_name": 1},
            ):
                vendors_by_id[v["vendor_id"]] = v.get("vendor_name")
        for s in subs:
            u = users_by_id.get(s.get("submitted_by") or "") if s.get("submitted_by") else None
            if u:
                s.setdefault("submitted_by_email", u.get("email"))
                s.setdefault("submitted_by_name",  u.get("name"))
                s.setdefault("submitter_role",     u.get("role"))
                if u.get("vendor_id"):
                    s.setdefault("vendor_id", u["vendor_id"])
                if u.get("region") and not s.get("region"):
                    s["region"] = u["region"]
            vid = s.get("vendor_id")
            if vid and vid in vendors_by_id:
                s["vendor_name"] = vendors_by_id[vid]
        return subs

    for f in forms:
        sq = {"form_id": f["form_id"]}
        if sub_q_extra:
            sq = {"$and": [sq, sub_q_extra]}
        subs = await db.submissions.find(sq, {"_id": 0}).sort("created_at", -1).to_list(500)
        subs = await _enrich(subs)
        groups.append({
            "kind": "form",
            "id": f["form_id"],
            "title": f.get("title") or "Untitled",
            "slug": f.get("slug"),
            "status": f.get("status"),
            "field_summary": [
                {"id": fd["id"], "label": fd.get("label") or fd["id"], "type": fd.get("type")}
                for fd in (f.get("fields") or [])
                if fd.get("type") not in ("heading", "paragraph", "divider")
            ][:6],
            "count": len(subs),
            "submissions": subs,
        })

    # --- PDF templates — every admin/super_admin sees all ---------------
    if is_super_admin(user) or role == "admin":
        pdf_q: Dict[str, Any] = {"is_deleted": False}
    else:
        rls = form_filter(user)
        import json as _json
        rls = _json.loads(_json.dumps(rls).replace('"form_id"', '"template_id"'))
        pdf_q = {"$and": [{"is_deleted": False}, rls]}
    pdfs = await db.pdf_templates.find(pdf_q, {"_id": 0}).sort("updated_at", -1).to_list(2000)
    for t in pdfs:
        sq = {"template_id": t["template_id"]}
        if sub_q_extra:
            sq = {"$and": [sq, sub_q_extra]}
        subs = await db.pdf_submissions.find(sq, {"_id": 0}).sort("created_at", -1).to_list(500)
        subs = await _enrich(subs)
        groups.append({
            "kind": "pdf",
            "id": t["template_id"],
            "title": t.get("title") or "Untitled PDF",
            "slug": t.get("slug"),
            "status": t.get("status"),
            "field_summary": [
                {"id": fd.get("id"), "label": fd.get("label") or fd.get("name") or fd.get("id"), "type": fd.get("type")}
                for fd in (t.get("fields") or [])
                if fd.get("type") not in ("heading", "paragraph", "static_text", "divider", "hidden")
            ][:6],
            "count": len(subs),
            "submissions": subs,
        })
    # Sort groups: those with subs first, then alphabetical
    groups.sort(key=lambda g: (-g["count"], g["title"].lower()))
    return groups


# ---------- Excel (.xlsx) export ----------
def _xlsx_val(v):
    if v is None:
        return ""
    if isinstance(v, (list, tuple)):
        return ", ".join(str(x) for x in v)
    if isinstance(v, dict):
        # display friendly for file/upload/signature dicts
        if "filename" in v:
            return str(v.get("filename"))
        return str(v)
    if isinstance(v, str) and v.startswith("data:image"):
        return "[signature image]"
    return v


@api.get("/forms/{form_id}/submissions/export.xlsx")
async def export_submissions_xlsx(form_id: str, user: User = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    form = await _get_form_for_user(form_id, user)
    scope = await _submission_scope_query(user)
    xls_q = {"$and": [{"form_id": form_id}, scope]} if scope else {"form_id": form_id}
    rows = await db.submissions.find(xls_q, {"_id": 0}) \
        .sort("created_at", 1).to_list(5000)
    fields = [f for f in (form.get("fields") or [])
              if f.get("type") not in ("heading", "paragraph", "divider")]
    field_ids = [f["id"] for f in fields]
    field_labels = {f["id"]: (f.get("label") or f["id"]) for f in fields}

    wb = Workbook()
    ws = wb.active
    ws.title = (form.get("title") or "Submissions")[:31]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    sub_font = Font(italic=True, color="475569")
    # Row 1 = human-friendly labels
    label_row = ["Submission ID", "Status", "Submitted At", "Submitted By"] + \
                [field_labels[fid] for fid in field_ids]
    # Row 2 = machine-friendly keys/IDs
    key_row = ["submission_id", "status", "created_at", "submitted_by"] + field_ids

    ws.append(label_row)
    ws.append(key_row)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    for cell in ws[2]:
        cell.font = sub_font

    for r in rows:
        vals = r.get("values") or {}
        ws.append(
            [r.get("submission_id"), r.get("status"), r.get("created_at"),
             r.get("submitted_by") or ""] +
            [_xlsx_val(vals.get(fid, "")) for fid in field_ids],
        )
    # simple column autosize
    for i, col in enumerate(ws.columns, start=1):
        w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(w + 2, 12), 48)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_slug = form.get("slug") or form_id
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_slug}-submissions.xlsx"'},
    )


# ---------- Users (admin) ----------
def _user_scope_filter(actor: User) -> Dict[str, Any]:
    """Restrict which users an actor can list/edit.
    - super_admin: everyone
    - admin: everyone EXCEPT super_admin accounts (cannot see/manage them)
    - vendor_admin: only users in the same vendor_id
    - vendor_user: none (empty query returns their own only)
    """
    from permissions import normalize_role
    role = normalize_role(actor.role)
    if role == "super_admin":
        return {}
    if role == "admin":
        return {"role": {"$ne": "super_admin"}}
    if role == "vendor_admin":
        return {"vendor_id": actor.vendor_id, "role": {"$in": ["vendor", "vendor_admin", "vendor_user"]}}
    return {"user_id": actor.user_id}


async def _require_user_admin(user: User = Depends(get_current_user)) -> User:
    """Auth guard for user-management endpoints. Allow super_admin, admin,
    vendor_admin — each with a different scope enforced by the filter.
    vendor_user is allowed only to read (list returns just themselves)."""
    from permissions import normalize_role
    role = normalize_role(user.role)
    if role not in ("super_admin", "admin", "vendor_admin", "vendor_user"):
        raise HTTPException(403, "Forbidden")
    return user


async def _require_user_editor(user: User = Depends(get_current_user)) -> User:
    """Stricter guard for user create/update/delete — excludes vendor_user."""
    from permissions import normalize_role
    role = normalize_role(user.role)
    if role not in ("super_admin", "admin", "vendor_admin"):
        raise HTTPException(403, "Forbidden")
    return user


@api.get("/users", response_model=List[UserOut])
async def list_users(user: User = Depends(_require_user_admin)):
    q = _user_scope_filter(user)
    rows = await db.users.find(q, {"_id": 0, "password_hash": 0}).sort("created_at", -1).to_list(1000)
    return [UserOut(**r) for r in rows]

class UserCreateIn(BaseModel):
    email: EmailStr
    name: str
    password: Optional[str] = Field(default=None, min_length=6)
    role: str = "user"
    vendor_id: Optional[str] = None
    cluster_manager_name: Optional[str] = None
    region: Optional[str] = None
    access_override: bool = False
    send_welcome_email: bool = True

def _gen_temp_password(n: int = 10) -> str:
    import secrets
    import string
    alpha = string.ascii_letters + string.digits
    # ensure at least one uppercase + one digit + one symbol
    return (
        secrets.choice(string.ascii_uppercase)
        + secrets.choice(string.digits)
        + "!"
        + "".join(secrets.choice(alpha) for _ in range(max(3, n - 3)))
    )

@api.post("/users", response_model=Dict[str, Any])
async def create_user(body: UserCreateIn, user: User = Depends(_require_user_editor)):
    from permissions import normalize_role
    if body.role not in ROLES:
        raise HTTPException(status_code=400, detail="Invalid role")
    # Prevent non-super-admin from creating a super_admin
    if body.role == "super_admin" and normalize_role(user.role) != "super_admin":
        raise HTTPException(403, "Only super_admin can create super_admin accounts")
    # vendor_admin can only create vendor scope users, and only in own vendor
    if normalize_role(user.role) == "vendor_admin":
        if body.role not in ("vendor", "vendor_user"):
            raise HTTPException(403, "Vendor Admins can only create vendor users")
        body.vendor_id = user.vendor_id  # force scope
    email = body.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=409, detail="Email taken")
    temp_password = body.password or _gen_temp_password()
    generated = body.password is None
    # Strength check ONLY when the caller supplied a password.
    # Auto-generated ones already meet the policy (see _gen_temp_password).
    if not generated:
        validate_password_strength(temp_password)
    uid = f"user_{uuid.uuid4().hex[:12]}"
    # Only super_admin may grant `access_override` at creation.
    override_flag = bool(body.access_override) and normalize_role(user.role) == "super_admin"
    doc = {
        "user_id": uid, "email": email, "name": body.name, "role": body.role,
        "password_hash": hash_password(temp_password), "picture": None,
        "is_active": True, "created_at": datetime.now(timezone.utc).isoformat(),
        "vendor_id": body.vendor_id or None,
        "cluster_manager_name": body.cluster_manager_name or None,
        "region": body.region or None,
        "access_override": override_flag,
    }
    await db.users.insert_one(doc)

    # Best-effort welcome email — never blocks user creation.
    email_status = "skipped"
    email_error: Optional[str] = None
    if body.send_welcome_email:
        try:
            from workflow_routes import _send_email, EmailRequest
            settings = await db.workspace_settings.find_one({"_id": "welcome_email"}, {"_id": 0}) or {}
            base_url = os.environ.get("PUBLIC_BASE_URL") or (
                f"https://{os.environ.get('HOSTNAME','localhost')}"
            )
            subject_tpl = settings.get("subject") or "Welcome to FormForge — your account is ready"
            body_tpl = settings.get("body_html") or (
                "<p>Hi {{name}},</p>"
                "<p>Your account on FormForge has been created.</p>"
                "<ul>"
                "<li><b>Email:</b> {{email}}</li>"
                "<li><b>Temporary password:</b> {{password}}</li>"
                "</ul>"
                "<p><a href=\"{{login_url}}\">Sign in</a> and change your password on first login.</p>"
                "<p>— The FormForge team</p>"
            )
            def _fmt(s: str) -> str:
                return (s.replace("{{name}}", body.name)
                          .replace("{{email}}", email)
                          .replace("{{password}}", temp_password)
                          .replace("{{login_url}}", f"{base_url}/login"))
            res = await _send_email(
                db,
                EmailRequest(to=[email], subject=_fmt(subject_tpl), body_html=_fmt(body_tpl)),
            )
            raw_status = res.get("status", "sent") if isinstance(res, dict) else "sent"
            # Bucket transport statuses so the frontend only sees sent/skipped/failed.
            if raw_status == "sent":
                email_status = "sent"
            elif raw_status.startswith("skip"):
                email_status = "skipped"
            else:
                email_status = "failed"
        except Exception as exc:  # noqa: BLE001
            email_status = "failed"
            email_error = str(exc)[:240]

    doc.pop("password_hash", None)
    out = UserOut(**doc).model_dump()
    # Return the temp password ONLY when it was generated OR the welcome email failed;
    # this lets the creator hand it off manually if SMTP is not configured.
    if generated or email_status != "sent":
        out["temp_password"] = temp_password
    out["email_status"] = email_status
    if email_error:
        out["email_error"] = email_error
    return out

class UserUpdateIn(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    is_active: Optional[bool] = None
    password: Optional[str] = None
    vendor_id: Optional[str] = None
    cluster_manager_name: Optional[str] = None
    region: Optional[str] = None
    assignments: Optional[Dict[str, List[str]]] = None
    access_override: Optional[bool] = None

@api.patch("/users/{user_id}", response_model=UserOut)
async def update_user(user_id: str, body: UserUpdateIn, user: User = Depends(_require_user_editor)):
    from permissions import normalize_role
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(404, "User not found")
    actor_role = normalize_role(user.role)
    target_role = normalize_role(target.get("role", ""))
    # Non-super_admin cannot edit a super_admin account
    if target_role == "super_admin" and actor_role != "super_admin":
        raise HTTPException(403, "Only super_admin can edit a super_admin")
    # vendor_admin can only edit users in own vendor and cannot promote
    if actor_role == "vendor_admin":
        if target.get("vendor_id") != user.vendor_id:
            raise HTTPException(403, "Out of scope")
        if body.role and body.role not in ("vendor", "vendor_user"):
            raise HTTPException(403, "Vendor Admins can only assign vendor roles")
        # Deactivating a vendor user requires super_admin / admin approval —
        # short-circuit here so the request is queued rather than applied.
        if body.is_active is False and target.get("is_active", True) is not False:
            from datetime import datetime, timezone as _tz
            import uuid as _uuid
            existing = await db.pending_approvals.find_one({
                "type": "user_disable",
                "target_user_id": user_id,
                "status": "pending",
            })
            if existing:
                raise HTTPException(409, "A disable request is already pending for this user")
            approval = {
                "approval_id": f"apv_{_uuid.uuid4().hex[:12]}",
                "type": "user_disable",
                "target_user_id": user_id,
                "target_email": target.get("email"),
                "target_name": target.get("name"),
                "target_region": target.get("region"),
                "target_cluster_manager_name": target.get("cluster_manager_name"),
                "vendor_id": target.get("vendor_id"),
                "requested_by": {
                    "user_id": user.user_id,
                    "name": getattr(user, "name", None),
                    "email": getattr(user, "email", None),
                    "role": user.role,
                    "region": getattr(user, "region", None),
                    "cluster_manager_name": getattr(user, "cluster_manager_name", None),
                },
                "status": "pending",
                "created_at": datetime.now(_tz.utc).isoformat(),
            }
            await db.pending_approvals.insert_one(dict(approval))
            approval.pop("_id", None)
            return JSONResponse(
                status_code=202,
                content={"pending_approval": True, "approval": approval,
                         "message": "Disable request submitted for admin approval"},
            )

    updates = {}
    if body.name is not None:
        updates["name"] = body.name
    if body.role is not None:
        if body.role not in ROLES:
            raise HTTPException(400, "Invalid role")
        if body.role == "super_admin" and actor_role != "super_admin":
            raise HTTPException(403, "Only super_admin can grant super_admin")
        updates["role"] = body.role
    if body.is_active is not None:
        updates["is_active"] = body.is_active
    if body.password:
        updates["password_hash"] = hash_password(body.password)
    if body.vendor_id is not None:
        updates["vendor_id"] = body.vendor_id or None
    if body.cluster_manager_name is not None:
        updates["cluster_manager_name"] = body.cluster_manager_name or None
    if body.region is not None:
        updates["region"] = body.region or None
    if body.assignments is not None:
        updates["assignments"] = body.assignments
    if body.access_override is not None:
        # Only super_admin can toggle the access override.
        if actor_role != "super_admin":
            raise HTTPException(403, "Only super_admin can toggle Add-on access override")
        updates["access_override"] = bool(body.access_override)
    if not updates:
        raise HTTPException(400, "No fields")
    res = await db.users.update_one({"user_id": user_id}, {"$set": updates})
    if not res.matched_count:
        raise HTTPException(404, "User not found")
    doc = await db.users.find_one({"user_id": user_id}, {"_id": 0, "password_hash": 0})
    return UserOut(**doc)

@api.delete("/users/{user_id}")
async def delete_user(user_id: str, user: User = Depends(_require_user_editor)):
    from permissions import normalize_role
    if user_id == user.user_id:
        raise HTTPException(400, "Cannot delete yourself")
    target = await db.users.find_one({"user_id": user_id}, {"_id": 0, "role": 1, "vendor_id": 1})
    if not target:
        raise HTTPException(404, "User not found")
    actor_role = normalize_role(user.role)
    target_role = normalize_role(target.get("role", ""))
    if target_role == "super_admin" and actor_role != "super_admin":
        raise HTTPException(403, "Only super_admin can delete a super_admin")
    if actor_role == "vendor_admin" and target.get("vendor_id") != user.vendor_id:
        raise HTTPException(403, "Out of scope")
    await db.users.delete_one({"user_id": user_id})
    return {"ok": True}


# ---------- Region metadata (populated from Site Master) ----------
@api.get("/regions", response_model=List[str])
async def list_regions(user: User = Depends(get_current_user)):
    """Union of distinct site regions (RLS-filtered) + curated master-data
    regions. This lets admins add "possible" regions ahead of any plant
    being seeded there so the User create dropdown always has the full set.
    """
    from permissions import site_filter, is_super_admin
    q: Dict[str, Any] = {}
    if not is_super_admin(user):
        q = site_filter(user)
    live = await db.sites.distinct("region", q)
    master = await db.master_data.distinct("value", {"table": "regions"})
    combined = {r for r in live if r} | {m for m in master if m}
    return sorted(combined)


@api.post("/regions", response_model=List[str])
async def add_region(body: Dict[str, str], user: User = Depends(_require_user_editor)):
    """Register a new region in the master list.  Idempotent by (table, value)."""
    from permissions import is_super_admin, normalize_role
    if not (is_super_admin(user) or normalize_role(user.role) == "admin"):
        raise HTTPException(403, "Only Admin/Super Admin can manage regions")
    value = (body.get("value") or "").strip()
    if not value:
        raise HTTPException(400, "Region name required")
    await db.master_data.update_one(
        {"table": "regions", "value": value},
        {"$setOnInsert": {
            "table": "regions",
            "value": value,
            "row_id": f"reg_{uuid.uuid4().hex[:8]}",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "created_by": user.user_id,
        }},
        upsert=True,
    )
    return await list_regions(user)


@api.get("/cluster-managers", response_model=List[str])
async def list_cluster_managers(user: User = Depends(get_current_user)):
    from permissions import site_filter, is_super_admin
    q: Dict[str, Any] = {}
    if not is_super_admin(user):
        q = site_filter(user)
    rows = await db.sites.distinct("cluster_manager_name", q)
    return sorted([r for r in rows if r])


# ---------- Forms ----------
def _slug(title: str) -> str:
    base = "".join(c.lower() if c.isalnum() else "-" for c in title).strip("-")[:40] or "form"
    return f"{base}-{uuid.uuid4().hex[:6]}"

@api.get("/forms", response_model=List[Form])
async def list_forms(user: User = Depends(get_current_user),
                     archived: bool = False, favorite: Optional[bool] = None,
                     q: Optional[str] = None):
    from permissions import form_filter, is_super_admin
    query: Dict[str, Any] = {"is_deleted": False, "is_archived": archived}
    # Apply role-based RLS unless super admin
    if not is_super_admin(user):
        query = {"$and": [query, form_filter(user)]}
    if favorite is not None:
        # Keep the favorite filter outside the $and rewrap so we don't lose it
        if "$and" in query:
            query["$and"].append({"is_favorite": favorite})
        else:
            query["is_favorite"] = favorite
    if q:
        clause = {"title": {"$regex": q, "$options": "i"}}
        if "$and" in query:
            query["$and"].append(clause)
        else:
            query["title"] = clause["title"]
    rows = await db.forms.find(query, {"_id": 0}).sort("updated_at", -1).to_list(500)
    # Enrich with owner_name/email — cheap when the list is capped at 500
    owner_ids = {r.get("owner_id") for r in rows if r.get("owner_id")}
    owners: Dict[str, Dict[str, Any]] = {}
    if owner_ids:
        for o in await db.users.find(
            {"user_id": {"$in": list(owner_ids)}}, {"_id": 0, "user_id": 1, "name": 1, "email": 1},
        ).to_list(1000):
            owners[o["user_id"]] = o
    for r in rows:
        o = owners.get(r.get("owner_id"))
        if o:
            r["owner_name"] = o.get("name")
            r["owner_email"] = o.get("email")
    return [Form(**r) for r in rows]

@api.post("/forms", response_model=Form)
async def create_form(body: FormIn, user: User = Depends(get_current_user)):
    from permissions import require_can_create_form
    require_can_create_form(user)
    now = datetime.now(timezone.utc).isoformat()
    fid = f"form_{uuid.uuid4().hex[:12]}"
    doc = body.model_dump()
    doc.update({
        "form_id": fid, "slug": _slug(body.title), "owner_id": user.user_id,
        "created_at": now, "updated_at": now,
        "is_favorite": False, "is_archived": False, "is_deleted": False,
    })
    await db.forms.insert_one(doc)
    doc.pop("_id", None)
    return Form(**doc)

async def _get_form_for_user(form_id: str, user: User, *, write: bool = False) -> dict:
    from permissions import can_edit_form, can_view_form
    doc = await db.forms.find_one({"form_id": form_id, "is_deleted": False}, {"_id": 0})
    if not doc:
        raise HTTPException(404, "Form not found")
    if write:
        if not can_edit_form(user, doc):
            raise HTTPException(403, "You do not have permission to edit this form")
    else:
        if not can_view_form(user, doc):
            raise HTTPException(403, "You do not have permission to view this form")
    return doc

@api.get("/forms/{form_id}", response_model=Form)
async def get_form(form_id: str, user: User = Depends(get_current_user)):
    return Form(**await _get_form_for_user(form_id, user))

@api.put("/forms/{form_id}", response_model=Form)
async def update_form(form_id: str, body: FormIn, user: User = Depends(get_current_user)):
    existing = await _get_form_for_user(form_id, user, write=True)
    updates = body.model_dump()
    updates["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.forms.update_one({"form_id": form_id}, {"$set": updates})
    existing.update(updates)
    return Form(**existing)

class FormPatch(BaseModel):
    is_favorite: Optional[bool] = None
    is_archived: Optional[bool] = None
    status: Optional[str] = None
    title: Optional[str] = None
    assigned_site_ids: Optional[List[str]] = None
    assigned_vendor_ids: Optional[List[str]] = None
    assigned_vendor_user_ids: Optional[List[str]] = None
    assigned_admin_ids: Optional[List[str]] = None
    assigned_member_ids: Optional[List[str]] = None
    assigned_department_ids: Optional[List[str]] = None
    assigned_team_ids: Optional[List[str]] = None
    assigned_cluster_managers: Optional[List[str]] = None
    assigned_regions: Optional[List[str]] = None

@api.patch("/forms/{form_id}", response_model=Form)
async def patch_form(form_id: str, body: FormPatch, user: User = Depends(get_current_user)):
    existing = await _get_form_for_user(form_id, user, write=True)
    upd = {k: v for k, v in body.model_dump().items() if v is not None}
    upd["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.forms.update_one({"form_id": form_id}, {"$set": upd})
    existing.update(upd)
    return Form(**existing)

@api.post("/forms/{form_id}/duplicate", response_model=Form)
async def duplicate_form(form_id: str, user: User = Depends(get_current_user)):
    from permissions import require_can_create_form
    require_can_create_form(user)
    existing = await _get_form_for_user(form_id, user)
    new_id = f"form_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc).isoformat()
    new_doc = {**existing,
               "form_id": new_id,
               "slug": _slug(existing["title"]),
               "title": existing["title"] + " (Copy)",
               "owner_id": user.user_id,
               "created_at": now, "updated_at": now,
               "is_favorite": False, "is_archived": False, "status": "draft"}
    await db.forms.insert_one(new_doc)
    new_doc.pop("_id", None)
    return Form(**new_doc)

@api.delete("/forms/{form_id}")
async def delete_form(form_id: str, user: User = Depends(get_current_user)):
    await _get_form_for_user(form_id, user, write=True)
    await db.forms.update_one({"form_id": form_id}, {"$set": {"is_deleted": True}})
    return {"ok": True}

# ---------- Public form access ----------
@api.get("/public/forms/{slug}")
async def public_get_form(slug: str):
    doc = await db.forms.find_one({"slug": slug, "is_deleted": False, "is_archived": False},
                                  {"_id": 0, "owner_id": 0})
    if not doc:
        raise HTTPException(404, "Form not found")
    if doc.get("status") != "published":
        raise HTTPException(403, "Form is not published")
    return doc

@api.post("/public/forms/{slug}/submit")
async def public_submit(slug: str, body: SubmissionIn, request: Request,
                        viewer: User = Depends(get_current_user)):
    check_rate_limit(PUBLIC_SUBMIT_LIMITER, f"submit:ip:{_client_ip(request)}", "submit")
    form = await db.forms.find_one({"slug": slug, "is_deleted": False}, {"_id": 0})
    if not form:
        raise HTTPException(404, "Form not found")
    if form.get("status") != "published":
        raise HTTPException(403, "Form is not accepting submissions")
    # required-field validation
    for f in form.get("fields", []):
        if f.get("required") and f["type"] not in ("heading", "paragraph", "divider"):
            v = body.values.get(f["id"])
            if v is None or v == "" or (isinstance(v, list) and len(v) == 0):
                raise HTTPException(400, f"Field '{f.get('label') or f['id']}' is required")
    sid = f"sub_{uuid.uuid4().hex[:12]}"
    doc = {
        "submission_id": sid, "form_id": form["form_id"], "values": body.values,
        "submitted_by": viewer.user_id,
        "submitted_by_name": getattr(viewer, "name", None),
        "submitted_by_email": getattr(viewer, "email", None),
        "ip": request.client.host if request.client else None,
        "user_agent": request.headers.get("user-agent"),
        "status": "submitted",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.submissions.insert_one(doc)
    doc.pop("_id", None)
    # Move any referenced uploads into /submissions/{sid}/ on disk
    try:
        await _organize_submission_files(sid, body.values)
    except Exception as _e:  # noqa: BLE001
        logger.warning(f"file organization failed for {sid}: {_e}")
    # fire workflow trigger (best-effort; never blocks the response)
    try:
        from workflow_routes import fire_trigger as _ft
        # Section 8 — workflow triggers carry enough context to identify which
        # form/PDF, which site, which vendor, and the current submission status.
        site_name = body.values.get("site_name") or body.values.get("site")
        vendor_name = body.values.get("vendor_name") or body.values.get("vendor")
        await _ft(db, "form_submitted",
                  {"submission_id": sid,
                   "submission_kind": "form",
                   "form_id": form["form_id"], "form_name": form.get("title"),
                   "form_type": "form",
                   "site_name": site_name,
                   "vendor_name": vendor_name,
                   "current_status": doc["status"],
                   "values": body.values, "user_id": viewer.user_id,
                   "user_email": viewer.email, "ip": doc["ip"]})
    except Exception as _e:  # noqa: BLE001
        logger.warning(f"workflow trigger form_submitted failed: {_e}")
    # Short-lived token so the submitter can download their filled PDF
    download_token = make_download_token(sid, kind="form")
    return {**Submission(**doc).model_dump(), "download_token": download_token}

# ---------- Submissions (owner view) ----------
@api.get("/forms/{form_id}/submissions", response_model=List[Submission])
async def list_submissions(form_id: str, user: User = Depends(get_current_user)):
    await _get_form_for_user(form_id, user)
    scope = await _submission_scope_query(user)
    query = {"$and": [{"form_id": form_id}, scope]} if scope else {"form_id": form_id}
    rows = await db.submissions.find(query, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return [Submission(**r) for r in rows]

async def _submission_scope_query(user: User) -> Dict[str, Any]:
    """Return a Mongo query fragment that restricts a submissions collection
    (standard `submissions` OR `pdf_submissions`) to the rows the given user
    is allowed to see.

    * super_admin / access_override → {} (no restriction)
    * admin → async_submission_filter (region / cluster / assigned scope)
    * vendor_admin → submissions submitted by any user of the same vendor
    * vendor_user → only their own submissions
    """
    from permissions import (
        normalize_role, is_super_admin, has_access_override,
        async_submission_filter,
    )
    if is_super_admin(user) or has_access_override(user):
        return {}
    role = normalize_role(getattr(user, "role", ""))
    if role == "admin":
        return await async_submission_filter(db, user)
    if role == "vendor_admin" and user.vendor_id:
        team = await db.users.find(
            {"vendor_id": user.vendor_id}, {"_id": 0, "user_id": 1},
        ).to_list(2000)
        return {"submitted_by": {"$in": [u["user_id"] for u in team]}}
    if role == "vendor_user" and user.user_id:
        return {"submitted_by": user.user_id}
    return {"submission_id": "__none__"}


@api.get("/submissions/{submission_id}", response_model=Submission)
async def get_submission(submission_id: str, user: User = Depends(get_current_user)):
    sub = await db.submissions.find_one({"submission_id": submission_id}, {"_id": 0})
    if not sub:
        raise HTTPException(404, "Not found")
    await _get_form_for_user(sub["form_id"], user)
    if not await _submission_in_user_scope(user, sub):
        raise HTTPException(403, "You do not have permission to view this submission")
    return Submission(**sub)

class SubmissionStatusIn(BaseModel):
    status: str  # submitted | approved | rejected

@api.patch("/submissions/{submission_id}", response_model=Submission)
async def update_submission_status(submission_id: str, body: SubmissionStatusIn,
                                   user: User = Depends(get_current_user)):
    sub = await db.submissions.find_one({"submission_id": submission_id}, {"_id": 0})
    if not sub:
        raise HTTPException(404, "Not found")
    await _get_form_for_user(sub["form_id"], user)
    if not await _submission_in_user_scope(user, sub):
        raise HTTPException(403, "You do not have permission to modify this submission")
    if body.status not in ("submitted", "approved", "rejected"):
        raise HTTPException(400, "Invalid status")
    await db.submissions.update_one({"submission_id": submission_id}, {"$set": {"status": body.status}})
    sub["status"] = body.status
    return Submission(**sub)

@api.delete("/submissions/{submission_id}")
async def delete_submission(submission_id: str, user: User = Depends(get_current_user)):
    sub = await db.submissions.find_one({"submission_id": submission_id}, {"_id": 0})
    if not sub:
        raise HTTPException(404, "Not found")
    await _get_form_for_user(sub["form_id"], user)
    if not await _submission_in_user_scope(user, sub):
        raise HTTPException(403, "You do not have permission to delete this submission")
    await db.submissions.delete_one({"submission_id": submission_id})
    return {"ok": True}


async def _submission_in_user_scope(user: User, sub: Dict[str, Any]) -> bool:
    """Return True if this specific submission row is inside the user's
    row-level-scope. Uses the same logic as `_submission_scope_query` but
    evaluates it against a single already-fetched submission document."""
    from permissions import normalize_role, is_super_admin, has_access_override
    if is_super_admin(user) or has_access_override(user):
        return True
    role = normalize_role(getattr(user, "role", ""))
    if role == "admin":
        # Admins already have form-level access; region/cluster scope is
        # enforced elsewhere via async_submission_filter on list endpoints.
        return True
    uid = getattr(user, "user_id", None)
    if role == "vendor_user":
        return sub.get("submitted_by") == uid
    if role == "vendor_admin" and user.vendor_id:
        if sub.get("submitted_by") == uid:
            return True
        submitter = await db.users.find_one(
            {"user_id": sub.get("submitted_by")}, {"_id": 0, "vendor_id": 1},
        )
        return bool(submitter and submitter.get("vendor_id") == user.vendor_id)
    return False


# ---------- Submitter-facing "download filled PDF" for standard forms ----
async def _can_view_own_submission(user: User, sub: Dict[str, Any]) -> bool:
    """Submitter + their vendor_admin can view. Also anyone who has RLS
    access to the parent form (already enforced by _get_form_for_user)."""
    from permissions import normalize_role, is_super_admin
    if is_super_admin(user):
        return True
    if sub.get("submitted_by") == user.user_id:
        return True
    role = normalize_role(user.role)
    if role == "vendor_admin" and user.vendor_id:
        # look up submitter's vendor
        submitter = await db.users.find_one(
            {"user_id": sub.get("submitted_by")},
            {"_id": 0, "vendor_id": 1},
        )
        if submitter and submitter.get("vendor_id") == user.vendor_id:
            return True
    return False


@api.get("/submissions/{submission_id}/filled.pdf")
async def download_filled_pdf(submission_id: str, user: User = Depends(get_current_user)):
    """Generate a printable PDF of a standard-form submission.

    Visible to: submitter, their vendor_admin, form owners (via RLS),
    super_admin. Uses reportlab — one field per row, labeled."""
    sub = await db.submissions.find_one({"submission_id": submission_id}, {"_id": 0})
    if not sub:
        raise HTTPException(404, "Submission not found")
    # Auth: either the submitter/vendor_admin OR anyone who can view the form
    allowed = await _can_view_own_submission(user, sub)
    if not allowed:
        try:
            await _get_form_for_user(sub["form_id"], user)
            allowed = True
        except HTTPException:
            allowed = False
    if not allowed:
        raise HTTPException(403, "You do not have permission to download this submission")

    form = await db.forms.find_one({"form_id": sub["form_id"], "is_deleted": False}, {"_id": 0})
    if not form:
        raise HTTPException(404, "Parent form missing")

    return _render_filled_pdf_response(sub, form)


@api.get("/public/submissions/{submission_id}/filled.pdf")
async def public_download_filled_pdf(submission_id: str, token: str):
    """Anonymous-safe download of a filled submission PDF.
    Validates a short-lived token generated by the public submit endpoint."""
    verify_download_token(token, submission_id, kind="form")
    sub = await db.submissions.find_one({"submission_id": submission_id}, {"_id": 0})
    if not sub:
        raise HTTPException(404, "Submission not found")
    form = await db.forms.find_one({"form_id": sub["form_id"], "is_deleted": False}, {"_id": 0})
    if not form:
        raise HTTPException(404, "Parent form missing")
    return _render_filled_pdf_response(sub, form)


def _render_filled_pdf_response(sub: Dict[str, Any], form: Dict[str, Any]) -> Response:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image,
    )
    from reportlab.lib import colors
    import base64 as _b64

    # ---- image embedding helpers ------------------------------------------
    IMAGE_EXTS = {"png", "jpg", "jpeg", "gif", "webp", "bmp"}
    MAX_IMG_W = 100 * mm    # thumbnail cell width
    MAX_IMG_H = 70 * mm

    def _load_uploaded_image(file_ref: Dict[str, Any]):
        """Return a reportlab Image flowable for a `{file_id, filename, ...}`
        ref if it points to an image on disk; else return None."""
        fid = file_ref.get("file_id")
        if not fid:
            return None
        ct = (file_ref.get("content_type") or "").lower()
        fname = (file_ref.get("filename") or "").lower()
        ext = fname.rsplit(".", 1)[-1] if "." in fname else ""
        if not (ct.startswith("image/") or ext in IMAGE_EXTS):
            return None
        # Locate the file bytes on disk. We honour the same paths that
        # _organize_submission_files creates, plus the legacy fallback.
        candidates = [
            LOCAL_UPLOAD_ROOT / "submissions" / sub["submission_id"] / (file_ref.get("filename") or ""),
            LOCAL_UPLOAD_ROOT / "tmp" / f"{fid}.{ext}" if ext else None,
        ]
        for p in candidates:
            if p and p.exists():
                try:
                    return _fit_image(str(p))
                except Exception as _e:  # noqa: BLE001
                    logger.warning(f"image embed failed for {p}: {_e}")
                    return None
        # Fallback: search by basename anywhere under uploads/
        try:
            fallback_root = Path(str(LOCAL_UPLOAD_ROOT).replace("/local", ""))
            for cand in fallback_root.rglob(file_ref.get("filename") or "__none__"):
                if cand.is_file():
                    return _fit_image(str(cand))
        except Exception:  # noqa: BLE001
            pass
        return None

    def _fit_image(path_str: str) -> "Image":
        img = Image(path_str)
        iw, ih = img.imageWidth, img.imageHeight
        if iw <= 0 or ih <= 0:
            return img
        scale = min(MAX_IMG_W / iw, MAX_IMG_H / ih, 1.0)
        img.drawWidth  = iw * scale
        img.drawHeight = ih * scale
        img.hAlign = "LEFT"
        return img

    def _load_datauri_image(uri: str):
        """Signature/inline data-URLs → embed as an image cell."""
        try:
            header, b64 = uri.split(",", 1)
            data = _b64.b64decode(b64)
            tmp = LOCAL_UPLOAD_ROOT / "tmp" / f"__inline_{uuid.uuid4().hex[:10]}.png"
            tmp.write_bytes(data)
            img = _fit_image(str(tmp))
            # cleanup after reportlab has read it (we defer with a marker attr)
            img._ff_tmpfile = tmp
            return img
        except Exception as _e:  # noqa: BLE001
            logger.warning(f"data-URI decode failed: {_e}")
            return None
    # ------------------------------------------------------------------------

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=18*mm, rightMargin=18*mm,
                            topMargin=18*mm, bottomMargin=18*mm)
    styles = getSampleStyleSheet()
    h_style = ParagraphStyle("h", parent=styles["Heading1"], fontSize=18, spaceAfter=6)
    meta_style = ParagraphStyle("meta", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#64748B"))
    label_style = ParagraphStyle("lbl", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#334155"))
    val_style = ParagraphStyle("val", parent=styles["Normal"], fontSize=11, textColor=colors.HexColor("#0F172A"))
    caption_style = ParagraphStyle("cap", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#64748B"))

    elems = [
        Paragraph(form.get("title") or "Submission", h_style),
        Paragraph(
            f"Submission ID: {sub['submission_id']} · Status: {sub.get('status','submitted')}"
            f" · {sub.get('created_at','')}",
            meta_style,
        ),
        Spacer(1, 8),
    ]
    rows = []
    inline_tmps: list = []
    for f in form.get("fields") or []:
        if f.get("type") in ("heading", "paragraph", "divider"):
            continue
        v = (sub.get("values") or {}).get(f["id"], "")
        if f.get("type") == "tick":
            checked = v is True or v == "true" or v == 1 or v == "1"
            tick_label = f.get("tick_label") or "Yes"
            if checked:
                v_html = ('<font name="ZapfDingbats" size="13" color="#059669">4</font>'
                          f'&nbsp;&nbsp;{tick_label}')
            else:
                v_html = ('<font name="ZapfDingbats" size="13" color="#94A3B8">8</font>'
                          '&nbsp;&nbsp;<font color="#94A3B8">Not confirmed</font>')
            rows.append([Paragraph(str(f.get("label") or f["id"]), label_style),
                         Paragraph(v_html, val_style)])
            continue

        # File/image uploads — try to embed the actual image
        if isinstance(v, dict) and v.get("file_id"):
            img_flow = _load_uploaded_image(v)
            if img_flow is not None:
                # Stack image + a small filename caption underneath
                cap = Paragraph(v.get("filename") or "", caption_style)
                cell = [img_flow, Spacer(1, 2), cap]
                rows.append([Paragraph(str(f.get("label") or f["id"]), label_style), cell])
                continue
            # Non-image file → just show filename as text
            v = v.get("filename") or str(v)
        elif isinstance(v, str) and v.startswith("data:image"):
            img_flow = _load_datauri_image(v)
            if img_flow is not None:
                if getattr(img_flow, "_ff_tmpfile", None):
                    inline_tmps.append(img_flow._ff_tmpfile)
                rows.append([Paragraph(str(f.get("label") or f["id"]), label_style), img_flow])
                continue
            v = "[image]"
        elif isinstance(v, (list, tuple)):
            v = ", ".join(str(x) for x in v)

        rows.append([Paragraph(str(f.get("label") or f["id"]), label_style),
                     Paragraph(str(v) if v not in (None, "") else "—", val_style)])

    if rows:
        table = Table(rows, colWidths=[55*mm, 115*mm])
        table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LINEBELOW", (0, 0), (-1, -1), 0.4, colors.HexColor("#E2E8F0")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
        ]))
        elems.append(table)
    doc.build(elems)
    # Best-effort cleanup of any inline temp files we spawned
    for p in inline_tmps:
        try:
            p.unlink(missing_ok=True)
        except Exception:
            pass
    buf.seek(0)
    from filename_resolver import resolve_filename as _rf
    fname = _rf(form.get("filename_template"),
                form=form, submission=sub)
    return Response(
        content=buf.getvalue(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@api.get("/forms/{form_id}/submissions/export.csv")
async def export_submissions_csv(form_id: str, user: User = Depends(get_current_user)):
    form = await _get_form_for_user(form_id, user)
    rows = await db.submissions.find({"form_id": form_id}, {"_id": 0}).sort("created_at", 1).to_list(5000)
    field_ids = [f["id"] for f in form.get("fields", []) if f["type"] not in ("heading", "paragraph", "divider")]
    field_labels = {f["id"]: f.get("label") or f["id"] for f in form.get("fields", [])}
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["submission_id", "status", "created_at"] + [field_labels[fid] for fid in field_ids])
    for r in rows:
        vals = r.get("values", {})
        w.writerow([r["submission_id"], r["status"], r["created_at"]] +
                   [_csv_val(vals.get(fid, "")) for fid in field_ids])
    out = buf.getvalue()
    return Response(content=out, media_type="text/csv",
                    headers={"Content-Disposition": f'attachment; filename="{form["slug"]}-submissions.csv"'})

def _csv_val(v):
    if isinstance(v, (list, tuple)):
        return ", ".join(str(x) for x in v)
    if isinstance(v, dict):
        return str(v)
    return v if v is not None else ""

# ---------- Uploads ----------
ALLOWED_EXTS = {"pdf", "doc", "docx", "png", "jpg", "jpeg", "zip", "rar",
                "mp4", "xlsx", "csv", "txt", "gif", "webp"}

@api.post("/upload")
async def upload(request: Request, file: UploadFile = File(...), user: User = Depends(get_current_user)):
    check_rate_limit(UPLOAD_LIMITER, f"upload:user:{user.user_id}", "upload")
    data = await file.read()
    if len(data) > MAX_UPLOAD_MB * 1024 * 1024:
        raise HTTPException(413, f"File exceeds {MAX_UPLOAD_MB}MB limit")
    ext = (file.filename.rsplit(".", 1)[-1] if "." in file.filename else "bin").lower()
    if ext not in ALLOWED_EXTS:
        raise HTTPException(400, f"Extension '{ext}' not allowed")
    validate_upload(file.filename, data)
    file_id = uuid.uuid4().hex
    # Land in the tmp bucket first — moved into submissions/{sid}/ once the
    # submission that references this file is actually created.
    path = f"tmp/{file_id}.{ext}"
    ct = file.content_type or mimetypes.guess_type(file.filename)[0] or "application/octet-stream"
    result = put_object(path, data, ct)
    doc = {
        "file_id": file_id, "storage_path": result["path"], "original_filename": file.filename,
        "content_type": ct, "size": result.get("size", len(data)),
        "uploaded_by": user.user_id, "is_deleted": False,
        "submission_id": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.files.insert_one(doc)
    return {"file_id": file_id, "filename": file.filename, "size": doc["size"], "content_type": ct,
            "url": f"/api/files/{file_id}"}

@api.post("/public/upload")
async def public_upload(request: Request, file: UploadFile = File(...)):
    """Anonymous uploads for public form submissions."""
    check_rate_limit(UPLOAD_LIMITER, f"upload:ip:{_client_ip(request)}", "upload")
    data = await file.read()
    if len(data) > MAX_UPLOAD_MB * 1024 * 1024:
        raise HTTPException(413, f"File exceeds {MAX_UPLOAD_MB}MB limit")
    ext = (file.filename.rsplit(".", 1)[-1] if "." in file.filename else "bin").lower()
    if ext not in ALLOWED_EXTS:
        raise HTTPException(400, f"Extension '{ext}' not allowed")
    validate_upload(file.filename, data)
    file_id = uuid.uuid4().hex
    path = f"tmp/{file_id}.{ext}"
    ct = file.content_type or mimetypes.guess_type(file.filename)[0] or "application/octet-stream"
    result = put_object(path, data, ct)
    doc = {
        "file_id": file_id, "storage_path": result["path"], "original_filename": file.filename,
        "content_type": ct, "size": result.get("size", len(data)),
        "uploaded_by": None, "is_deleted": False,
        "submission_id": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.files.insert_one(doc)
    return {"file_id": file_id, "filename": file.filename, "size": doc["size"], "content_type": ct,
            "url": f"/api/files/{file_id}"}

@api.get("/files/{file_id}")
async def download_file(file_id: str):
    record = await db.files.find_one({"file_id": file_id, "is_deleted": False}, {"_id": 0})
    if not record:
        raise HTTPException(404, "File not found")
    data, ct = get_object(record["storage_path"])
    return Response(content=data, media_type=record.get("content_type", ct),
                    headers={"Content-Disposition": f'inline; filename="{record["original_filename"]}"'})

# ---------- Dashboard ----------
@api.get("/dashboard/stats")
async def dashboard_stats(user: User = Depends(get_current_user)):
    """Aggregated stats across BOTH standard forms/submissions AND PDF
    templates/submissions.  Since iter 4d the form library is shared, so
    admins see all forms — only vendor-scoped roles get filtered."""
    now = datetime.now(timezone.utc)
    today_iso = now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()

    # ---- Forms & PDF templates -------------------------------------------------
    form_q: Dict[str, Any] = {"is_deleted": False}
    pdf_q: Dict[str, Any] = {"is_deleted": False}
    # vendor_admin / vendor_user only see forms owned by them / assigned to them.
    if user.role in ("vendor_admin", "vendor_user"):
        form_q["owner_id"] = user.user_id
        pdf_q["owner_id"] = user.user_id
    total_forms = (await db.forms.count_documents(form_q)
                   + await db.pdf_templates.count_documents(pdf_q))

    form_ids = [d["form_id"] async for d in db.forms.find(form_q, {"form_id": 1})]
    pdf_ids  = [d["template_id"] async for d in db.pdf_templates.find(pdf_q, {"template_id": 1})]
    form_titles = {d["form_id"]: d.get("title", "Form")
                   async for d in db.forms.find({"form_id": {"$in": form_ids}},
                                                {"form_id": 1, "title": 1})}
    pdf_titles  = {d["template_id"]: (d.get("title") or d.get("name") or "PDF form")
                   async for d in db.pdf_templates.find({"template_id": {"$in": pdf_ids}},
                                                        {"template_id": 1, "title": 1, "name": 1})}

    # ---- RLS on submissions (super_admin & admin see all, vendors scoped) -----
    from permissions import async_submission_filter, is_super_admin, has_access_override
    if is_super_admin(user) or has_access_override(user) or user.role == "admin":
        rls: Dict[str, Any] = {}
    else:
        rls = await async_submission_filter(db, user) or {}

    def _q(base: Dict[str, Any]) -> Dict[str, Any]:
        return {"$and": [base, rls]} if rls else base

    std_q = _q({"form_id": {"$in": form_ids}}) if form_ids else {"form_id": {"$in": []}}
    pdf_sub_q = _q({"template_id": {"$in": pdf_ids}}) if pdf_ids else {"template_id": {"$in": []}}

    # ---- Totals ---------------------------------------------------------------
    total_subs = (await db.submissions.count_documents(std_q)
                  + await db.pdf_submissions.count_documents(pdf_sub_q))
    today_subs = (await db.submissions.count_documents({**std_q, "created_at": {"$gte": today_iso}})
                  + await db.pdf_submissions.count_documents({**pdf_sub_q, "created_at": {"$gte": today_iso}}))
    pending = (await db.submissions.count_documents({**std_q, "status": "submitted"})
               + await db.pdf_submissions.count_documents({**pdf_sub_q, "status": "submitted"}))

    users_count = await db.users.count_documents({}) if is_super_admin(user) else 1

    # storage
    files_q: Dict[str, Any] = {"is_deleted": False}
    if user.role in ("vendor_admin", "vendor_user"):
        files_q["uploaded_by"] = user.user_id
    storage_bytes = 0
    async for d in db.files.find(files_q, {"size": 1, "_id": 0}):
        storage_bytes += int(d.get("size") or 0)

    # 14-day series ------------------------------------------------------------
    series = []
    for i in range(13, -1, -1):
        day = now.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=i)
        next_day = day + timedelta(days=1)
        c = (await db.submissions.count_documents(
                {**std_q, "created_at": {"$gte": day.isoformat(), "$lt": next_day.isoformat()}})
             + await db.pdf_submissions.count_documents(
                {**pdf_sub_q, "created_at": {"$gte": day.isoformat(), "$lt": next_day.isoformat()}}))
        series.append({"date": day.strftime("%b %d"), "count": c})

    # recent activity (union of both, sorted, top 8) ---------------------------
    std_recent = await db.submissions.find(std_q, {"_id": 0}).sort("created_at", -1).limit(8).to_list(8)
    pdf_recent = await db.pdf_submissions.find(pdf_sub_q, {"_id": 0}).sort("created_at", -1).limit(8).to_list(8)
    activity_all: List[Dict[str, Any]] = []
    for s in std_recent:
        activity_all.append({
            "type": "submission",
            "kind": "form",
            "form_title": form_titles.get(s.get("form_id"), "Form"),
            "submission_id": s.get("submission_id"),
            "created_at": s.get("created_at"),
            "status": s.get("status"),
        })
    for s in pdf_recent:
        activity_all.append({
            "type": "submission",
            "kind": "pdf",
            "form_title": pdf_titles.get(s.get("template_id"), "PDF form"),
            "submission_id": s.get("submission_id"),
            "created_at": s.get("created_at"),
            "status": s.get("status"),
        })
    activity_all.sort(key=lambda a: a.get("created_at") or "", reverse=True)
    activity = activity_all[:8]

    # per_form top 6 ----------------------------------------------------------
    per_form: List[Dict[str, Any]] = []
    for fid in form_ids[:20]:
        cnt = await db.submissions.count_documents(_q({"form_id": fid}))
        if cnt:
            per_form.append({"form_id": fid, "title": form_titles.get(fid, ""), "count": cnt, "kind": "form"})
    for tid in pdf_ids[:20]:
        cnt = await db.pdf_submissions.count_documents(_q({"template_id": tid}))
        if cnt:
            per_form.append({"form_id": tid, "title": pdf_titles.get(tid, ""), "count": cnt, "kind": "pdf"})
    per_form.sort(key=lambda x: x["count"], reverse=True)

    return {
        "totals": {
            "forms": total_forms, "submissions": total_subs, "today": today_subs,
            "pending": pending, "users": users_count, "storage_bytes": storage_bytes,
        },
        "trend": series,
        "activity": activity,
        "per_form": per_form[:6],
    }

# ---------- Settings ----------
class SmtpIn(BaseModel):
    host: str = ""
    port: int = 587
    username: str = ""
    password: str = ""
    from_email: str = ""
    use_tls: bool = True
    enabled: bool = False

class SettingsIn(BaseModel):
    company_name: str = "FormForge"
    company_logo_url: str = ""
    primary_color: str = "#2563EB"
    smtp: SmtpIn = SmtpIn()

@api.get("/settings", response_model=SettingsIn)
async def get_settings(user: User = Depends(require_role("super_admin"))):
    doc = await db.settings.find_one({"_id": "global"}, {"_id": 0})
    if not doc:
        return SettingsIn()
    # don't leak password back to UI
    if "smtp" in doc and "password" in doc["smtp"]:
        doc["smtp"]["password"] = "********" if doc["smtp"]["password"] else ""
    return SettingsIn(**doc)

@api.put("/settings", response_model=SettingsIn)
async def update_settings(body: SettingsIn, user: User = Depends(require_role("super_admin"))):
    doc = body.model_dump()
    # if password is mask, keep existing
    if doc["smtp"]["password"] == "********":
        existing = await db.settings.find_one({"_id": "global"})
        if existing and existing.get("smtp", {}).get("password"):
            doc["smtp"]["password"] = existing["smtp"]["password"]
        else:
            doc["smtp"]["password"] = ""
    await db.settings.update_one({"_id": "global"}, {"$set": doc}, upsert=True)
    if doc["smtp"]["password"]:
        doc["smtp"]["password"] = "********"
    return SettingsIn(**doc)


# --- Branding (public + logo upload) --------------------------------------
_BRANDING_DIR = Path(__file__).parent / "uploads" / "local" / "branding"
_BRANDING_DIR.mkdir(parents=True, exist_ok=True)


@api.get("/public/branding")
async def public_branding():
    """Unauthenticated read of the workspace branding — used by the login
    screen and the initial SPA bootstrap so the app name / logo can be
    shown before the user is authenticated."""
    doc = await db.settings.find_one({"_id": "global"}, {"_id": 0}) or {}
    return {
        "app_name":       doc.get("company_name")     or "FormForge",
        "logo_url":       doc.get("company_logo_url") or "",
        "primary_color":  doc.get("primary_color")    or "#2563EB",
    }


@api.post("/settings/logo")
async def upload_logo(
    file: UploadFile = File(...),
    user: User = Depends(require_role("super_admin")),
):
    """Upload a logo image. Stores the file on disk and returns the public
    URL that the frontend can then save to `company_logo_url`."""
    data = await file.read()
    max_bytes = 2 * 1024 * 1024   # 2 MB is plenty for a logo
    if len(data) > max_bytes:
        raise HTTPException(413, "Logo must be under 2 MB")
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(400, "Upload must be an image file")
    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in ("png", "jpg", "jpeg", "webp", "svg", "gif"):
        raise HTTPException(400, "Only PNG / JPG / WEBP / SVG / GIF allowed")
    fname = f"logo-{uuid.uuid4().hex[:12]}.{ext}"
    (_BRANDING_DIR / fname).write_bytes(data)
    url = f"/api/public/branding/logo/{fname}"
    # Also persist so admins don't have to re-save Settings
    await db.settings.update_one(
        {"_id": "global"}, {"$set": {"company_logo_url": url}}, upsert=True,
    )
    return {"logo_url": url}


@api.get("/public/branding/logo/{name}")
async def serve_logo(name: str):
    """Serve the uploaded logo file (public — logos are meant to be
    shown to unauthenticated users on the login page)."""
    # Path-traversal defence: keep to just the basename we generated.
    safe = Path(name).name
    fp = _BRANDING_DIR / safe
    if not fp.exists() or not fp.is_file():
        raise HTTPException(404, "Not found")
    ext = fp.suffix.lstrip(".").lower()
    media = {
        "png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg",
        "webp": "image/webp", "svg": "image/svg+xml", "gif": "image/gif",
    }.get(ext, "application/octet-stream")
    return FileResponse(fp, media_type=media)


# ---------- Welcome email template (configurable per workspace) ----------
DEFAULT_WELCOME_EMAIL = {
    "subject": "Welcome to FormForge — your account is ready",
    "body_html": (
        "<p>Hi {{name}},</p>"
        "<p>Your account on FormForge has been created.</p>"
        "<ul>"
        "<li><b>Email:</b> {{email}}</li>"
        "<li><b>Temporary password:</b> {{password}}</li>"
        "</ul>"
        "<p><a href=\"{{login_url}}\">Sign in</a> and change your password on first login.</p>"
        "<p>— The FormForge team</p>"
    ),
}


class WelcomeEmailIn(BaseModel):
    subject: str
    body_html: str


@api.get("/settings/welcome-email", response_model=WelcomeEmailIn)
async def get_welcome_email(user: User = Depends(_require_user_editor)):
    doc = await db.workspace_settings.find_one({"_id": "welcome_email"}, {"_id": 0}) or {}
    return WelcomeEmailIn(
        subject=doc.get("subject") or DEFAULT_WELCOME_EMAIL["subject"],
        body_html=doc.get("body_html") or DEFAULT_WELCOME_EMAIL["body_html"],
    )


@api.put("/settings/welcome-email", response_model=WelcomeEmailIn)
async def put_welcome_email(body: WelcomeEmailIn, user: User = Depends(require_role("super_admin", "admin"))):
    await db.workspace_settings.update_one(
        {"_id": "welcome_email"},
        {"$set": {"subject": body.subject, "body_html": body.body_html,
                  "updated_at": datetime.now(timezone.utc).isoformat(),
                  "updated_by": user.user_id}},
        upsert=True,
    )
    return body


class WelcomeEmailPreviewIn(BaseModel):
    subject: str
    body_html: str
    name: str = "Sample User"
    email: str = "sample@example.com"
    password: str = "P4ssW0rd!"


@api.post("/settings/welcome-email/preview")
async def preview_welcome_email(body: WelcomeEmailPreviewIn, user: User = Depends(_require_user_editor)):
    base_url = os.environ.get("PUBLIC_BASE_URL") or ""
    login_url = f"{base_url}/login" if base_url else "/login"
    def _fmt(s: str) -> str:
        return (s.replace("{{name}}", body.name)
                  .replace("{{email}}", body.email)
                  .replace("{{password}}", body.password)
                  .replace("{{login_url}}", login_url))
    return {"subject": _fmt(body.subject), "body_html": _fmt(body.body_html)}


@api.get("/health")
async def health():
    return {"status": "ok", "time": datetime.now(timezone.utc).isoformat()}

# ---------- PDF Form Builder ----------
from pdf_routes import build_pdf_router
_pdf_tpl_router, _pdf_public_router, _pdf_sub_router, _pdf_pub_sub_router = build_pdf_router(
    db, get_current_user, get_optional_user,
    make_download_token=make_download_token,
    verify_download_token=verify_download_token,
    organize_submission_files=_organize_submission_files,
    uploads_root=LOCAL_UPLOAD_ROOT,
)
api.include_router(_pdf_tpl_router)
api.include_router(_pdf_public_router)
api.include_router(_pdf_sub_router)
api.include_router(_pdf_pub_sub_router)

# ---------- Workflow Automation ----------
from workflow_routes import build_workflow_routers
(_wf_router, _apv_router, _audit_router, _wfa_router, _smtp_router, _pub_apv_router) = \
    build_workflow_routers(db, get_current_user)
api.include_router(_wf_router)
api.include_router(_apv_router)
api.include_router(_audit_router)
api.include_router(_wfa_router)
api.include_router(_smtp_router)
api.include_router(_pub_apv_router)

# ---------- Vendor Management / Site Master / Master Data ----------
from vendor_routes import build_routers as _build_vendor_routers
_vendors_r, _vusers_r, _sites_r, _master_r, _lookup_r, _pub_lookup_r, _approvals_r = _build_vendor_routers(
    db, get_current_user, hash_password, get_optional_user,
)
api.include_router(_vendors_r)
api.include_router(_vusers_r)
api.include_router(_sites_r)
api.include_router(_master_r)
api.include_router(_lookup_r)
api.include_router(_pub_lookup_r)
api.include_router(_approvals_r)

# ---------- Schedule vs Actual (per-site monthly cycles) ----------
from schedule_routes import build_router as _build_schedule_router
api.include_router(_build_schedule_router(db, get_current_user))

# ---------- Formula Engine ----------
from formula_routes import build_formula_router
api.include_router(build_formula_router(db, get_current_user))

# ---------- Extended Data Sources (REST API / JSON / CSV / Excel / Another Form / Workflow Variable) ----------
from datasource_routes import build_datasource_router
api.include_router(build_datasource_router(db, get_current_user))

# ---------- Manpower Portal read-only mirror (external cmes_mp_db) ----------
from manpower_routes import build_manpower_router
api.include_router(build_manpower_router(db, get_current_user))

# ---------- In-app Notifications (bell icon + real-time WebSocket) ----------
async def _resolve_ws_user(token: str) -> Optional[User]:
    """Given a JWT (from ?token=... on the WebSocket handshake) return the
    matching User or None.  Kept intentionally quiet — we do NOT raise;
    the WebSocket route closes with code 4401 on a missing/invalid token."""
    try:
        data = decode_token(token)
    except HTTPException:
        return None
    user = await db.users.find_one({"user_id": data.get("sub")}, {"_id": 0, "password_hash": 0})
    if not user or not user.get("is_active", True):
        return None
    return User(**user)

from notifications import build_notifications_router
api.include_router(build_notifications_router(db, get_current_user, _resolve_ws_user))

# ---------- Plant Document Vault ----------
from plant_docs_routes import build_router as _build_plant_docs
_pd_router, _pd_plants = _build_plant_docs(db, get_current_user)
api.include_router(_pd_router)
api.include_router(_pd_plants)

# ---------- Backup / Restore ----------
from backup_routes import build_router as _build_backup_router, start_scheduler as _start_backup_scheduler
_bk_router, _bk_cfg = _build_backup_router(db, get_current_user)
api.include_router(_bk_router)
api.include_router(_bk_cfg)


@app.on_event("startup")
async def _kick_backup_scheduler():
    # Runs alongside the FastAPI event loop. Safe no-op if scheduler is off.
    _start_backup_scheduler(db)

# Mount router
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=[o.strip() for o in os.environ.get('CORS_ORIGINS', '*').split(',') if o.strip()],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Security headers (HSTS, CSP, X-Frame-Options, etc.).
# HSTS only when SECURITY_HTTPS=true — locally we serve HTTP.
app.add_middleware(
    SecurityHeadersMiddleware,
    enable_hsts=os.environ.get("SECURITY_HTTPS", "false").lower() == "true",
)
