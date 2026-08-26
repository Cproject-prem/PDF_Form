"""
Inventory Management Routes
Handles equipment and spare parts per site, stock movements, and summaries.
"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Response
from fastapi.responses import FileResponse
from typing import Optional, List
from datetime import datetime, timezone
from pathlib import Path
from bson import ObjectId
import logging
import os
import re
import mimetypes
import io
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from permissions import normalize_role, site_filter, has_access_override, SUPER_ADMIN, ADMIN

log = logging.getLogger("inventory_routes")

INVENTORY_DOCS_ROOT = Path(os.environ.get("LOCAL_UPLOAD_ROOT", "/app/backend/uploads/local")) / "inventory"

INVENTORY_EXCEL_COLUMNS = [
    {"key": "site_code", "label": "Plant Code (site_code)"},
    {"key": "item_type", "label": "Item Type (equipment/spare)"},
    {"key": "name", "label": "Item Name"},
    {"key": "equipment_type", "label": "Equipment Type"},
    {"key": "make", "label": "Make / Brand"},
    {"key": "model", "label": "Model"},
    {"key": "serial_number", "label": "Serial Number"},
    {"key": "quantity", "label": "Quantity"},
    {"key": "unit", "label": "Unit"},
    {"key": "status", "label": "Status (operational/faulty/under_maintenance)"},
    {"key": "min_stock_level", "label": "Min Stock Level (for spares)"},
    {"key": "location_in_plant", "label": "Location in Plant"},
    {"key": "notes", "label": "Notes / Remarks"},
]


def _sanitize_filename(name: str) -> str:
    name = os.path.basename((name or "").strip()) or "upload.bin"
    cleaned = re.sub(r"[^\w.\- ]", "_", name)
    return cleaned or "upload.bin"


def build_inventory_router(db, get_current_user):
    async def _require_admin(user=Depends(get_current_user)):
        sett = await db.settings.find_one({"_id": "global"}, {"enable_inventory": 1})
        if not (sett and sett.get("enable_inventory")):
            raise HTTPException(status_code=403, detail="Inventory module is currently disabled by Super Admin.")
        role = normalize_role(getattr(user, "role", ""))
        if role not in (SUPER_ADMIN, ADMIN):
            raise HTTPException(status_code=403, detail="Access denied: Inventory management is restricted to Administrators only.")
        return user


    router = APIRouter(prefix="/inventory", tags=["inventory"], dependencies=[Depends(_require_admin)])

    def _id(doc):
        if doc and "_id" in doc:
            doc["id"] = str(doc.pop("_id"))
        return doc

    def _clean(doc):
        if not doc:
            return None
        if "_id" in doc:
            doc["id"] = str(doc.pop("_id"))
        return doc

    def _spare_status(qty: int, min_stock: int) -> str:
        if qty <= 0:
            return "out_of_stock"
        if qty <= min_stock:
            return "low_stock"
        return "available"

    # ─────────────────────────────────────────
    # Helper: resolve site codes for current user
    # super_admin → all sites (or filtered by region param)
    # admin       → their own region/cluster sites (inner)
    #               if region_scope==outer → sites NOT in their region
    # ─────────────────────────────────────────
    async def _get_admin_site_codes(user, region_scope: str = "inner", explicit_region: str = None):
        role = normalize_role(getattr(user, "role", ""))
        if role == SUPER_ADMIN or has_access_override(user):
            # Super admin: filter by explicit_region if given
            q = {}
            if explicit_region:
                q["region"] = explicit_region
            sites = await db.sites.find(q, {"site_code": 1}).to_list(5000)
            return [s["site_code"] for s in sites if "site_code" in s], None

        # Admin: resolve their region
        sf = site_filter(user)
        all_sites = await db.sites.find({}, {"site_code": 1, "region": 1}).to_list(5000)

        # Get admin's own region
        user_region_val = getattr(user, "region", None)

        if region_scope == "outer":
            # Outer region: all sites NOT in admin's region
            if user_region_val:
                outer_codes = [s["site_code"] for s in all_sites if s.get("region") != user_region_val]
            else:
                outer_codes = [s["site_code"] for s in all_sites]
            return outer_codes, user_region_val
        else:
            # Inner region (default): admin's own region sites
            if sf == {}:
                codes = [s["site_code"] for s in all_sites]
            else:
                inner_sites = await db.sites.find(sf, {"site_code": 1}).to_list(5000)
                codes = [s["site_code"] for s in inner_sites if "site_code" in s]
            return codes, user_region_val

    # ─────────────────────────────────────────
    # GET /inventory/items  — list with filters
    # ─────────────────────────────────────────
    @router.get("/items")
    async def list_items(
        site_code: Optional[str] = None,
        item_type: Optional[str] = None,
        equipment_type: Optional[str] = None,
        status: Optional[str] = None,
        spare_availability: Optional[str] = None,
        stock_level: Optional[str] = None,
        search: Optional[str] = None,
        linked_equipment_id: Optional[str] = None,
        region_scope: Optional[str] = None,  # inner | outer (admin only)
        user=Depends(get_current_user),
    ):
        # --- Region-based site scoping ---
        allowed_site_codes, _user_region = await _get_admin_site_codes(user, region_scope or "inner")

        q = {}
        if site_code:
            # If an explicit site_code is given, honour it but check permission
            if site_code in allowed_site_codes:
                q["site_code"] = site_code
            else:
                q["site_code"] = site_code  # super_admin or explicit override
        else:
            q["site_code"] = {"$in": allowed_site_codes}

        if item_type:
            q["item_type"] = item_type
        if equipment_type:
            q["equipment_type"] = {"$regex": equipment_type, "$options": "i"}
        if status:
            q["status"] = status
        if linked_equipment_id:
            q["linked_equipment_id"] = linked_equipment_id
        if search:
            q["$or"] = [
                {"name": {"$regex": search, "$options": "i"}},
                {"serial_number": {"$regex": search, "$options": "i"}},
                {"make": {"$regex": search, "$options": "i"}},
                {"model": {"$regex": search, "$options": "i"}},
            ]

        items = await db.inventory_items.find(q).sort("created_at", -1).to_list(5000)
        result = []
        for it in items:
            it = _clean(it)
            itype = it.get("item_type", "equipment")
            if itype == "spare":
                it["spare_status"] = _spare_status(it.get("quantity", 0), it.get("min_stock_level", 0))
            if spare_availability and it.get("spare_status") != spare_availability:
                continue
            if stock_level == "low" and it.get("spare_status") not in ("low_stock", "out_of_stock"):
                continue
            result.append(it)
        return result

    # ─────────────────────────────────────────
    # GET /inventory/item-names  — distinct item names (for movement modal step 1)
    # Returns: [{name, count}] — names with total qty > 0 across accessible sites
    # ─────────────────────────────────────────
    @router.get("/item-names")
    async def get_item_names(
        region_scope: Optional[str] = "inner",
        item_type: Optional[str] = None,  # equipment | spare (optional filter)
        user=Depends(get_current_user),
    ):
        allowed_site_codes, _user_region = await _get_admin_site_codes(user, region_scope or "inner")
        q = {"site_code": {"$in": allowed_site_codes}, "quantity": {"$gt": 0}}
        if item_type:
            q["item_type"] = item_type
        items = await db.inventory_items.find(q, {"name": 1, "quantity": 1, "item_type": 1}).to_list(5000)
        name_map = {}
        for it in items:
            n = it.get("name", "").strip()
            if n:
                name_map[n] = name_map.get(n, 0) + it.get("quantity", 0)
        return sorted(
            [{"name": k, "total_qty": v} for k, v in name_map.items()],
            key=lambda x: x["name"]
        )

    # ─────────────────────────────────────────
    # GET /inventory/spare-sites  — sites having stock of a given item name
    # Returns: [{site_code, site_name, region, quantity}]
    # ─────────────────────────────────────────
    @router.get("/spare-sites")
    async def get_spare_sites(
        name: str,
        region_scope: Optional[str] = "inner",
        user=Depends(get_current_user),
    ):
        allowed_site_codes, _user_region = await _get_admin_site_codes(user, region_scope or "inner")
        q = {
            "name": {"$regex": f"^{name}$", "$options": "i"},
            "site_code": {"$in": allowed_site_codes},
            "quantity": {"$gt": 0},
        }
        items = await db.inventory_items.find(q, {"site_code": 1, "name": 1, "quantity": 1, "item_type": 1}).to_list(5000)

        # Aggregate quantity per site
        site_map = {}
        for it in items:
            sc = it.get("site_code", "")
            if sc:
                if sc not in site_map:
                    site_map[sc] = {"site_code": sc, "quantity": 0, "item_type": it.get("item_type", "spare")}
                site_map[sc]["quantity"] += it.get("quantity", 0)

        if not site_map:
            return []

        # Enrich with site names
        sites = await db.sites.find(
            {"site_code": {"$in": list(site_map.keys())}},
            {"site_code": 1, "site_name": 1, "region": 1, "cluster": 1}
        ).to_list(5000)
        for s in sites:
            sc = s.get("site_code", "")
            if sc in site_map:
                site_map[sc]["site_name"] = s.get("site_name", sc)
                site_map[sc]["region"] = s.get("region", "")
                site_map[sc]["cluster"] = s.get("cluster", "")

        return sorted(site_map.values(), key=lambda x: -x["quantity"])

    # ─────────────────────────────────────────
    # POST /inventory/items  — create
    # ─────────────────────────────────────────
    @router.post("/items")
    async def create_item(body: dict, user=Depends(get_current_user)):
        now = datetime.now(timezone.utc).isoformat()
        body["created_at"] = now
        body["updated_at"] = now
        body.setdefault("quantity", 0)
        body.setdefault("min_stock_level", 0)
        body.setdefault("status", "operational")
        body.setdefault("item_type", "equipment")
        r = await db.inventory_items.insert_one(body)
        created = await db.inventory_items.find_one({"_id": r.inserted_id})

        performer = str(getattr(user, "email", None) or getattr(user, "username", ""))
        item_id_str = str(r.inserted_id)
        item_name   = body.get("name", "")
        site_code   = body.get("site_code", "")
        item_type   = body.get("item_type", "equipment")
        qty         = int(body.get("quantity", 0))

        # ── Inventory movement log (item-level history) ──
        await db.inventory_movements.insert_one({
            "item_id":         item_id_str,
            "item_name":       item_name,
            "site_code":       site_code,
            "movement_type":   "added",
            "quantity_change": qty,
            "quantity_before": 0,
            "quantity_after":  qty,
            "reason":          f"New {item_type} added to inventory",
            "performed_by":    performer,
            "performed_at":    now,
        })

        # ── Plant-level history / audit log ──
        await db.plant_history.insert_one({
            "site_code":   site_code,
            "event_type":  "inventory_added",
            "item_id":     item_id_str,
            "item_name":   item_name,
            "item_type":   item_type,
            "quantity":    qty,
            "status":      body.get("status", "operational"),
            "description": f"New {item_type} '{item_name}' added — qty: {qty}",
            "performed_by": performer,
            "performed_at": now,
        })

        return _clean(created)

    # ─────────────────────────────────────────
    # GET /inventory/items/{item_id}
    # ─────────────────────────────────────────
    @router.get("/items/{item_id}")
    async def get_item(item_id: str, user=Depends(get_current_user)):
        try:
            doc = await db.inventory_items.find_one({"_id": ObjectId(item_id)})
        except Exception:
            raise HTTPException(400, "Invalid item ID")
        if not doc:
            raise HTTPException(404, "Item not found")
        return _clean(doc)

    # ─────────────────────────────────────────
    # PUT /inventory/items/{item_id}
    # ─────────────────────────────────────────
    @router.put("/items/{item_id}")
    async def update_item(item_id: str, body: dict, user=Depends(get_current_user)):
        try:
            oid = ObjectId(item_id)
        except Exception:
            raise HTTPException(400, "Invalid item ID")

        # Snapshot before state
        before_doc = await db.inventory_items.find_one({"_id": oid})
        qty_before  = int((before_doc or {}).get("quantity", 0))

        body.pop("_id", None)
        body.pop("id", None)
        now = datetime.now(timezone.utc).isoformat()
        body["updated_at"] = now
        await db.inventory_items.update_one({"_id": oid}, {"$set": body})
        doc = await db.inventory_items.find_one({"_id": oid})

        performer  = str(getattr(user, "email", None) or getattr(user, "username", ""))
        item_name  = doc.get("name", "") if doc else ""
        site_code  = doc.get("site_code", "") if doc else ""
        item_type  = doc.get("item_type", "equipment") if doc else "equipment"
        qty_after  = int((doc or {}).get("quantity", 0))
        qty_change = qty_after - qty_before
        status     = doc.get("status", "") if doc else ""

        # Build change summary
        changed_fields = [k for k in body if k not in ("updated_at",)]
        change_desc = f"Updated {item_type} '{item_name}'"
        if "quantity" in body:
            direction = "increased" if qty_change > 0 else "decreased" if qty_change < 0 else "unchanged"
            change_desc += f" — qty {direction}: {qty_before} → {qty_after}"
        if "status" in body:
            change_desc += f" — status set to '{status}'"

        # ── Inventory movement log (item-level history) ──
        await db.inventory_movements.insert_one({
            "item_id":         item_id,
            "item_name":       item_name,
            "site_code":       site_code,
            "movement_type":   "updated",
            "quantity_change": qty_change,
            "quantity_before": qty_before,
            "quantity_after":  qty_after,
            "changed_fields":  changed_fields,
            "reason":          change_desc,
            "performed_by":    performer,
            "performed_at":    now,
        })

        # ── Plant-level history / audit log ──
        await db.plant_history.insert_one({
            "site_code":    site_code,
            "event_type":   "inventory_updated",
            "item_id":      item_id,
            "item_name":    item_name,
            "item_type":    item_type,
            "quantity":     qty_after,
            "status":       status,
            "description":  change_desc,
            "changed_fields": changed_fields,
            "performed_by": performer,
            "performed_at": now,
        })

        return _clean(doc)

    # ─────────────────────────────────────────
    # DELETE /inventory/items/{item_id}
    # ─────────────────────────────────────────
    @router.delete("/items/{item_id}")
    async def delete_item(item_id: str, user=Depends(get_current_user)):
        try:
            oid = ObjectId(item_id)
        except Exception:
            raise HTTPException(400, "Invalid item ID")

        doc = await db.inventory_items.find_one({"_id": oid})
        if doc:
            now       = datetime.now(timezone.utc).isoformat()
            performer = str(getattr(user, "email", None) or getattr(user, "username", ""))
            item_name = doc.get("name", "")
            site_code = doc.get("site_code", "")
            item_type = doc.get("item_type", "equipment")
            qty       = int(doc.get("quantity", 0))

            # ── Inventory movement log ──
            await db.inventory_movements.insert_one({
                "item_id":         item_id,
                "item_name":       item_name,
                "site_code":       site_code,
                "movement_type":   "deleted",
                "quantity_change": -qty,
                "quantity_before": qty,
                "quantity_after":  0,
                "reason":          f"{item_type.capitalize()} '{item_name}' removed from inventory",
                "performed_by":    performer,
                "performed_at":    now,
            })

            # ── Plant history ──
            await db.plant_history.insert_one({
                "site_code":    site_code,
                "event_type":   "inventory_deleted",
                "item_id":      item_id,
                "item_name":    item_name,
                "item_type":    item_type,
                "quantity":     0,
                "description":  f"{item_type.capitalize()} '{item_name}' deleted (had qty: {qty})",
                "performed_by": performer,
                "performed_at": now,
            })

        await db.inventory_items.delete_one({"_id": oid})
        # Also remove linked spares
        await db.inventory_items.delete_many({"linked_equipment_id": item_id})
        return {"ok": True}

    # ─────────────────────────────────────────
    # POST /inventory/items/{item_id}/movement
    # ─────────────────────────────────────────
    @router.post("/items/{item_id}/movement")
    async def log_movement(item_id: str, body: dict, user=Depends(get_current_user)):
        try:
            oid = ObjectId(item_id)
        except Exception:
            raise HTTPException(400, "Invalid item ID")
        doc = await db.inventory_items.find_one({"_id": oid})
        if not doc:
            raise HTTPException(404, "Item not found")

        movement_type = body.get("movement_type", "adjustment")  # in | out | adjustment
        qty_change = int(body.get("quantity_change", 0))

        current_qty = doc.get("quantity", 0)
        if movement_type == "in":
            new_qty = current_qty + qty_change
        elif movement_type == "out":
            new_qty = max(0, current_qty - qty_change)
        else:
            new_qty = qty_change  # absolute set

        now = datetime.now(timezone.utc).isoformat()
        await db.inventory_items.update_one({"_id": oid}, {"$set": {"quantity": new_qty, "updated_at": now}})

        performer  = str(getattr(user, "email", None) or getattr(user, "user_id", ""))
        item_name  = doc.get("name", "")
        site_code  = doc.get("site_code", "")
        item_type  = doc.get("item_type", "equipment")
        reason_txt = body.get("reason", "")

        movement = {
            "item_id":         item_id,
            "item_name":       item_name,
            "site_code":       site_code,
            "movement_type":   movement_type,
            "quantity_change": qty_change,
            "quantity_before": current_qty,
            "quantity_after":  new_qty,
            "reason":          reason_txt,
            "performed_by":    performer,
            "performed_at":    now,
        }
        r = await db.inventory_movements.insert_one(movement)
        movement["id"] = str(r.inserted_id)
        movement.pop("_id", None)

        # ── Plant-level history ──
        direction_label = "restocked" if movement_type == "in" else "consumed/removed" if movement_type == "out" else "adjusted"
        await db.plant_history.insert_one({
            "site_code":    site_code,
            "event_type":   f"stock_{movement_type}",
            "item_id":      item_id,
            "item_name":    item_name,
            "item_type":    item_type,
            "quantity":     new_qty,
            "description":  f"Stock {direction_label} for '{item_name}': {current_qty} → {new_qty}" + (f" — {reason_txt}" if reason_txt else ""),
            "performed_by": performer,
            "performed_at": now,
        })

        return {"ok": True, "new_quantity": new_qty, "movement": movement}

    # ─────────────────────────────────────────
    # GET /inventory/items/{item_id}/movements
    # ─────────────────────────────────────────
    @router.get("/items/{item_id}/movements")
    async def get_movements(item_id: str, user=Depends(get_current_user)):
        docs = await db.inventory_movements.find({"item_id": item_id}).sort("performed_at", -1).to_list(200)
        return [_clean(d) for d in docs]

    # ─────────────────────────────────────────
    # GET /inventory/plant-history/{site_code}
    # Returns all inventory events for a plant (from plant_history)
    # ─────────────────────────────────────────
    @router.get("/plant-history/{site_code}")
    async def get_plant_inventory_history(
        site_code: str,
        limit: int = 100,
        user=Depends(get_current_user),
    ):
        docs = await db.plant_history.find({"site_code": site_code}).sort("performed_at", -1).to_list(limit)
        return [_clean(d) for d in docs]

    # ═══════════════════════════════════════════
    # INTER-PLANT TRANSFERS /inventory/transfer
    # ═══════════════════════════════════════════

    @router.post("/transfer")
    async def transfer_item(body: dict, user=Depends(get_current_user)):
        """Transfer stock from Plant A (from_site_code) to Plant B (to_site_code)."""
        from_site = (body.get("from_site_code") or "").strip()
        to_site = (body.get("to_site_code") or "").strip()
        item_id = (body.get("item_id") or "").strip()
        item_name_hint = (body.get("item_name") or "").strip()
        transfer_qty = int(body.get("quantity", 0))
        reason = (body.get("reason") or "").strip()
        ref_no = (body.get("reference_no") or "").strip()

        if not from_site or not to_site:
            raise HTTPException(400, "Source and Destination plants are required")
        if from_site == to_site:
            raise HTTPException(400, "Destination plant must be different from source plant")
        if transfer_qty <= 0:
            raise HTTPException(400, "Transfer quantity must be greater than 0")

        # Fetch source item — try ObjectId first, then name+site fallback
        source_doc = None
        try:
            oid = ObjectId(item_id)
            source_doc = await db.inventory_items.find_one({"_id": oid})
        except Exception:
            oid = None

        if not source_doc and (item_name_hint or item_id):
            # Fallback: look up by item name at source site
            name_lookup = item_name_hint or item_id
            source_doc = await db.inventory_items.find_one(
                {"site_code": from_site, "name": {"$regex": f"^{name_lookup}$", "$options": "i"}, "quantity": {"$gt": 0}}
            )
            if source_doc:
                oid = source_doc["_id"]

        if not source_doc:
            raise HTTPException(404, "Source item not found at the specified source plant")

        source_qty = int(source_doc.get("quantity", 0))
        if source_qty < transfer_qty:
            raise HTTPException(400, f"Insufficient stock at source plant. Available: {source_qty} {source_doc.get('unit', 'nos')}")

        now = datetime.now(timezone.utc).isoformat()

        # Resolve plant names
        from_site_obj = await db.sites.find_one({"site_code": from_site}, {"site_name": 1})
        to_site_obj   = await db.sites.find_one({"site_code": to_site}, {"site_name": 1})
        from_site_name = from_site_obj.get("site_name", from_site) if from_site_obj else from_site
        to_site_name   = to_site_obj.get("site_name", to_site) if to_site_obj else to_site

        # 1. Deduct from Source Plant (Plant A)
        new_source_qty = source_qty - transfer_qty
        await db.inventory_items.update_one({"_id": oid}, {"$set": {"quantity": new_source_qty, "updated_at": now}})

        # Log movement for Source Plant A
        await db.inventory_movements.insert_one({
            "item_id": item_id,
            "site_code": from_site,
            "movement_type": "transfer_out",
            "quantity_change": transfer_qty,
            "quantity_before": source_qty,
            "quantity_after": new_source_qty,
            "reason": f"Transferred to {to_site_name} ({to_site}) - {reason}".strip(" -"),
            "target_site_code": to_site,
            "performed_by": str(getattr(user, "email", None) or getattr(user, "user_id", "")),
            "performed_at": now,
        })

        # 2. Add / Update in Destination Plant (Plant B)
        # Search for matching item at destination
        item_name = source_doc.get("name")
        item_type = source_doc.get("item_type", "equipment")
        equip_type = source_doc.get("equipment_type", "")

        dest_doc = await db.inventory_items.find_one({
            "site_code": to_site,
            "name": item_name,
            "item_type": item_type,
        })

        if dest_doc:
            dest_id = str(dest_doc["_id"])
            dest_qty = int(dest_doc.get("quantity", 0))
            new_dest_qty = dest_qty + transfer_qty
            await db.inventory_items.update_one({"_id": dest_doc["_id"]}, {"$set": {"quantity": new_dest_qty, "updated_at": now}})
        else:
            # Create new item at Plant B copying source details
            new_dest_item = {
                "site_code": to_site,
                "item_type": item_type,
                "equipment_type": equip_type,
                "name": item_name,
                "make": source_doc.get("make", ""),
                "model": source_doc.get("model", ""),
                "serial_number": source_doc.get("serial_number", ""),
                "quantity": transfer_qty,
                "unit": source_doc.get("unit", "nos"),
                "status": "operational",
                "min_stock_level": source_doc.get("min_stock_level", 0),
                "location_in_plant": source_doc.get("location_in_plant", ""),
                "notes": f"Transferred from {from_site_name}",
                "created_at": now,
                "updated_at": now,
            }
            r_dest = await db.inventory_items.insert_one(new_dest_item)
            dest_id = str(r_dest.inserted_id)
            dest_qty = 0
            new_dest_qty = transfer_qty

        # Log movement for Destination Plant B
        await db.inventory_movements.insert_one({
            "item_id": dest_id,
            "site_code": to_site,
            "movement_type": "transfer_in",
            "quantity_change": transfer_qty,
            "quantity_before": dest_qty,
            "quantity_after": new_dest_qty,
            "reason": f"Transferred from {from_site_name} ({from_site}) - {reason}".strip(" -"),
            "source_site_code": from_site,
            "performed_by": str(getattr(user, "email", None) or getattr(user, "user_id", "")),
            "performed_at": now,
        })

        # 3. Create Transfer Audit Record
        transfer_record = {
            "from_site_code": from_site,
            "from_site_name": from_site_name,
            "to_site_code": to_site,
            "to_site_name": to_site_name,
            "source_item_id": item_id,
            "dest_item_id": dest_id,
            "item_name": item_name,
            "item_type": item_type,
            "equipment_type": equip_type,
            "quantity": transfer_qty,
            "unit": source_doc.get("unit", "nos"),
            "reason": reason,
            "reference_no": ref_no,
            "transferred_by": str(getattr(user, "name", None) or getattr(user, "email", None) or getattr(user, "user_id", "")),
            "transferred_at": now,
        }
        r_trf = await db.inventory_transfers.insert_one(transfer_record)
        transfer_record["id"] = str(r_trf.inserted_id)
        transfer_record.pop("_id", None)

        return {"ok": True, "message": f"Successfully transferred {transfer_qty} {source_doc.get('unit', 'nos')} of '{item_name}' from {from_site_name} to {to_site_name}", "transfer": transfer_record}

    @router.get("/transfers")
    async def list_transfers(
        site_code: Optional[str] = None,
        search: Optional[str] = None,
        user=Depends(get_current_user),
    ):
        """List inter-plant transfer logs."""
        q = {}
        if site_code:
            q["$or"] = [{"from_site_code": site_code}, {"to_site_code": site_code}]
        if search:
            sq = {"$regex": search, "$options": "i"}
            q["$or"] = [
                {"item_name": sq},
                {"from_site_name": sq},
                {"to_site_name": sq},
                {"reference_no": sq},
                {"reason": sq},
            ]
        docs = await db.inventory_transfers.find(q).sort("transferred_at", -1).to_list(500)
        return [_clean(d) for d in docs]

    # ─────────────────────────────────────────
    # GET /inventory/sites/{site_code}/items
    # ─────────────────────────────────────────
    @router.get("/sites/{site_code}/items")
    async def site_items(site_code: str, user=Depends(get_current_user)):
        docs = await db.inventory_items.find({"site_code": site_code}).sort("item_type", 1).to_list(5000)
        result = []
        for d in docs:
            d = _clean(d)
            if d.get("item_type") == "spare":
                d["spare_status"] = _spare_status(d.get("quantity", 0), d.get("min_stock_level", 0))
            result.append(d)
        return result

    # ─────────────────────────────────────────
    # GET /inventory/summary  — global summary
    # ─────────────────────────────────────────
    @router.get("/summary")
    async def inventory_summary(
        site_code: Optional[str] = None,
        region: Optional[str] = None,
        user=Depends(get_current_user),
    ):
        """Return per-site and aggregated inventory counts for dashboard cards."""
        role = normalize_role(getattr(user, "role", ""))

        # Build site filter respecting admin's region
        if role == SUPER_ADMIN or has_access_override(user):
            site_q = {}
            if region:
                site_q["region"] = region
        else:
            from permissions import site_filter as _sf
            site_q = _sf(user)

        sites = await db.sites.find(site_q, {"site_code": 1, "site_name": 1, "region": 1, "cluster": 1}).to_list(5000)
        site_codes = [s["site_code"] for s in sites if "site_code" in s]
        site_map = {s["site_code"]: s for s in sites}

        if site_code:
            site_codes = [site_code] if site_code in site_codes else [site_code]

        # Fetch all items for these sites
        items = await db.inventory_items.find({"site_code": {"$in": site_codes}}).to_list(5000)

        per_site = {}
        totals = {
            "total_equipment": 0,
            "operational": 0,
            "faulty": 0,
            "under_maintenance": 0,
            "total_spares": 0,
            "spares_available": 0,
            "spares_low": 0,
            "spares_out": 0,
        }

        for it in items:
            sc = it.get("site_code")
            if sc not in per_site:
                per_site[sc] = {
                    "site_code": sc,
                    "site_name": site_map.get(sc, {}).get("site_name", sc),
                    "region": site_map.get(sc, {}).get("region", ""),
                    "total_equipment": 0,
                    "operational": 0,
                    "faulty": 0,
                    "under_maintenance": 0,
                    "total_spares": 0,
                    "spares_available": 0,
                    "spares_low": 0,
                    "spares_out": 0,
                }

            itype = it.get("item_type", "equipment")
            if itype == "equipment":
                per_site[sc]["total_equipment"] += 1
                totals["total_equipment"] += 1
                st = it.get("status", "operational")
                if st == "operational":
                    per_site[sc]["operational"] += 1
                    totals["operational"] += 1
                elif st == "faulty":
                    per_site[sc]["faulty"] += 1
                    totals["faulty"] += 1
                elif st == "under_maintenance":
                    per_site[sc]["under_maintenance"] += 1
                    totals["under_maintenance"] += 1
            elif itype == "spare":
                spare_st = _spare_status(it.get("quantity", 0), it.get("min_stock_level", 0))
                per_site[sc]["total_spares"] += 1
                totals["total_spares"] += 1
                if spare_st == "available":
                    per_site[sc]["spares_available"] += 1
                    totals["spares_available"] += 1
                elif spare_st == "low_stock":
                    per_site[sc]["spares_low"] += 1
                    totals["spares_low"] += 1
                else:
                    per_site[sc]["spares_out"] += 1
                    totals["spares_out"] += 1

        return {
            "totals": totals,
            "per_site": list(per_site.values()),
        }

    # ─────────────────────────────────────────
    # GET /inventory/equipment-types — distinct types for filter
    # ─────────────────────────────────────────
    @router.get("/equipment-types")
    async def equipment_types(user=Depends(get_current_user)):
        types = await db.inventory_items.distinct("equipment_type")
        return sorted([t for t in types if t])

    # ═══════════════════════════════════════════
    # DOCUMENT MANAGEMENT  /inventory/items/{id}/docs
    # ═══════════════════════════════════════════

    def _item_doc_dir(item_id: str) -> Path:
        """Return (and ensure) the per-item document folder."""
        p = INVENTORY_DOCS_ROOT / item_id
        p.mkdir(parents=True, exist_ok=True)
        return p

    @router.get("/items/{item_id}/docs")
    async def list_docs(item_id: str, user=Depends(get_current_user)):
        """List all documents attached to an inventory item."""
        doc_dir = _item_doc_dir(item_id)
        files = []
        for f in sorted(doc_dir.iterdir()):
            if f.is_file():
                mime, _ = mimetypes.guess_type(f.name)
                files.append({
                    "name": f.name,
                    "size_bytes": f.stat().st_size,
                    "mime_type": mime or "application/octet-stream",
                    "uploaded_at": datetime.fromtimestamp(f.stat().st_mtime, tz=timezone.utc).isoformat(),
                })
        return files

    @router.post("/items/{item_id}/docs")
    async def upload_doc(
        item_id: str,
        file: UploadFile = File(...),
        user=Depends(get_current_user),
    ):
        """Upload a document to an inventory item."""
        # Verify item exists
        try:
            oid = ObjectId(item_id)
        except Exception:
            raise HTTPException(400, "Invalid item ID")
        if not await db.inventory_items.find_one({"_id": oid}):
            raise HTTPException(404, "Item not found")

        doc_dir = _item_doc_dir(item_id)
        fname = _sanitize_filename(file.filename or "document.bin")
        target = doc_dir / fname

        # Avoid clobbering existing files with same name
        if target.exists():
            stem = target.stem
            suffix = target.suffix
            i = 1
            while (doc_dir / f"{stem} ({i}){suffix}").exists():
                i += 1
            target = doc_dir / f"{stem} ({i}){suffix}"

        data = await file.read()
        target.write_bytes(data)
        mime, _ = mimetypes.guess_type(target.name)
        return {
            "name": target.name,
            "size_bytes": len(data),
            "mime_type": mime or "application/octet-stream",
        }

    @router.get("/items/{item_id}/docs/{filename}")
    async def download_doc(item_id: str, filename: str, user=Depends(get_current_user)):
        """Download / preview a document."""
        fname = _sanitize_filename(filename)
        target = _item_doc_dir(item_id) / fname
        if not target.exists():
            raise HTTPException(404, "File not found")
        mime, _ = mimetypes.guess_type(fname)
        return FileResponse(
            str(target),
            media_type=mime or "application/octet-stream",
            filename=fname,
        )

    @router.delete("/items/{item_id}/docs/{filename}")
    async def delete_doc(item_id: str, filename: str, user=Depends(get_current_user)):
        """Delete a document from an inventory item."""
        fname = _sanitize_filename(filename)
        target = _item_doc_dir(item_id) / fname
        if target.exists():
            target.unlink()
        return {"ok": True}

    # ═══════════════════════════════════════════
    # EXCEL TEMPLATE, EXPORT & BULK IMPORT
    # ═══════════════════════════════════════════

    @router.get("/template.xlsx")
    async def inventory_template(user=Depends(get_current_user)):
        """Download sample Excel template for bulk equipment/spares import."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Template"

        headers = [c["label"] for c in INVENTORY_EXCEL_COLUMNS]
        ws.append(headers)

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        sample1 = [
            "PLANT-001", "equipment", "SMA Inverter Unit 1", "Inverter",
            "SMA", "Sunny Central 2500", "SN-8829102", 1, "nos",
            "operational", 0, "Inverter Room A", "Main central inverter"
        ]
        sample2 = [
            "PLANT-001", "spare", "MC4 Connector Pairs", "DC Cable",
            "Staubli", "MC4-Evo2", "", 100, "pairs",
            "operational", 20, "Warehouse Rack 3", "Standard DC connectors"
        ]
        ws.append(sample1)
        ws.append(sample2)

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        buf = io.BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="inventory-template.xlsx"'},
        )

    @router.get("/export.xlsx")
    async def export_inventory_xlsx(
        site_code: Optional[str] = None,
        item_type: Optional[str] = None,
        user=Depends(get_current_user),
    ):
        """Export all current inventory items to Excel."""
        q = {}
        if site_code:
            q["site_code"] = site_code
        if item_type:
            q["item_type"] = item_type

        items = await db.inventory_items.find(q).sort("created_at", -1).to_list(10000)

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Export"

        headers = [c["label"] for c in INVENTORY_EXCEL_COLUMNS]
        ws.append(headers)

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for it in items:
            row = [
                it.get("site_code", ""),
                it.get("item_type", "equipment"),
                it.get("name", ""),
                it.get("equipment_type", ""),
                it.get("make", ""),
                it.get("model", ""),
                it.get("serial_number", ""),
                it.get("quantity", 0),
                it.get("unit", "nos"),
                it.get("status", "operational"),
                it.get("min_stock_level", 0),
                it.get("location_in_plant", ""),
                it.get("notes", ""),
            ]
            ws.append(row)

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        buf = io.BytesIO()
        wb.save(buf)
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="inventory-export.xlsx"'},
        )

    @router.post("/import.xlsx")
    async def import_inventory_xlsx(
        file: UploadFile = File(...),
        user=Depends(get_current_user),
    ):
        """Bulk import equipment/spares from Excel file (.xlsx)."""
        if not file.filename.endswith(".xlsx"):
            raise HTTPException(400, "File must be an Excel spreadsheet (.xlsx)")

        contents = await file.read()
        try:
            wb = load_workbook(filename=io.BytesIO(contents), data_only=True)
        except Exception as e:
            raise HTTPException(400, f"Failed to parse Excel file: {str(e)}")

        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            raise HTTPException(400, "Excel file has no data rows")

        header_row = [str(cell or "").strip().lower() for cell in rows[0]]

        def find_idx(possible_headers):
            for ph in possible_headers:
                for idx, h in enumerate(header_row):
                    if ph in h:
                        return idx
            return -1

        site_code_idx = find_idx(["plant code", "site_code", "plant"])
        item_type_idx = find_idx(["item type", "item_type", "type"])
        name_idx      = find_idx(["item name", "name"])
        eq_type_idx   = find_idx(["equipment type", "equipment_type"])
        make_idx      = find_idx(["make", "brand"])
        model_idx     = find_idx(["model"])
        serial_idx    = find_idx(["serial", "serial_number"])
        qty_idx       = find_idx(["quantity", "qty"])
        unit_idx      = find_idx(["unit"])
        status_idx    = find_idx(["status"])
        min_stock_idx = find_idx(["min stock", "min_stock_level"])
        loc_idx       = find_idx(["location", "location_in_plant"])
        notes_idx     = find_idx(["notes", "remarks"])

        if name_idx == -1 or site_code_idx == -1:
            raise HTTPException(400, "Excel file must contain 'Plant Code (site_code)' and 'Item Name' columns.")

        created_cnt = 0
        updated_cnt = 0
        skipped_cnt = 0
        now = datetime.now(timezone.utc).isoformat()

        for r in rows[1:]:
            if not r or not any(r):
                continue

            def get_val(idx, default=""):
                if idx != -1 and idx < len(r) and r[idx] is not None:
                    return str(r[idx]).strip()
                return default

            site_code = get_val(site_code_idx)
            name = get_val(name_idx)

            if not site_code or not name:
                skipped_cnt += 1
                continue

            raw_type = get_val(item_type_idx, "equipment").lower()
            item_type = "spare" if "spare" in raw_type else "equipment"

            try:
                qty = int(float(get_val(qty_idx, "0")))
            except Exception:
                qty = 0

            try:
                min_stock = int(float(get_val(min_stock_idx, "0")))
            except Exception:
                min_stock = 0

            status = get_val(status_idx, "operational").lower()
            if status not in ("operational", "faulty", "under_maintenance", "decommissioned"):
                status = "operational"

            item_data = {
                "site_code": site_code,
                "item_type": item_type,
                "name": name,
                "equipment_type": get_val(eq_type_idx, "Other"),
                "make": get_val(make_idx),
                "model": get_val(model_idx),
                "serial_number": get_val(serial_idx),
                "quantity": qty,
                "unit": get_val(unit_idx, "nos"),
                "status": status,
                "min_stock_level": min_stock,
                "location_in_plant": get_val(loc_idx),
                "notes": get_val(notes_idx),
                "updated_at": now,
            }

            existing = await db.inventory_items.find_one({
                "site_code": site_code,
                "name": name,
                "item_type": item_type,
            })

            if existing:
                await db.inventory_items.update_one({"_id": existing["_id"]}, {"$set": item_data})
                updated_cnt += 1
            else:
                item_data["created_at"] = now
                await db.inventory_items.insert_one(item_data)
                created_cnt += 1

        return {
            "ok": True,
            "message": f"Import complete: {created_cnt} created, {updated_cnt} updated, {skipped_cnt} skipped.",
            "created": created_cnt,
            "updated": updated_cnt,
            "skipped": skipped_cnt,
        }

    return router

