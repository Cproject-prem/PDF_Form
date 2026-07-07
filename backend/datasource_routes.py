"""External data-source resolvers for FormForge form fields.

The base `/api/lookup` & `/api/public/lookup` routers in `vendor_routes.py`
handle the master data sources (sites / vendors / master:*). This module
adds resolvers for *external* data sources that the form builder also
supports as "Data Source" options:

  - rest_api  : GET an HTTP endpoint and pull a list out of the JSON body
  - json      : Inline JSON array configured in the form
  - csv       : Inline CSV configured in the form
  - excel     : Inline rows from an uploaded .xlsx file
  - another_form: distinct values from a sibling form's submissions
  - workflow_variable: reads a variable from the form's workflow context

All endpoints respond with `{ items: [...] }` where items are either bare
values or `{label, value, row}` records — same shape used by the existing
lookup options endpoint so the renderer can stay generic.
"""
from __future__ import annotations

import csv as _csv
import io
import json
import logging
from typing import Any, Dict, List

import requests
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook

log = logging.getLogger("datasource")


def build_datasource_router(db, get_current_user):
    r = APIRouter(prefix="/data-source", tags=["data-source"])

    @r.post("/resolve")
    async def resolve(body: Dict[str, Any], user=Depends(get_current_user)):
        """Generic resolver — returns a list of options for any data source.

        Body:
          { type, ...source-specific args, display_column?, value_column? }
        """
        return await _resolve(body, user, db)

    @r.post("/excel-upload")
    async def excel_upload(file: UploadFile = File(...), _user=Depends(get_current_user)):
        data = await file.read()
        try:
            wb = load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            raise HTTPException(400, f"Could not read Excel file: {e}")
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(next(it, []) or [])]
        rows: List[dict] = []
        for r_row in it:
            if not r_row or all(c is None for c in r_row): continue
            rows.append({headers[i]: r_row[i] for i in range(min(len(headers), len(r_row)))})
        return {"columns": headers, "rows": rows}

    return r


async def _resolve(src: Dict[str, Any], user, db) -> Dict[str, Any]:
    t = src.get("type") or src.get("kind") or "manual"
    items: List[dict] = []

    if t == "manual":
        for opt in src.get("manual_options") or []:
            if isinstance(opt, dict):
                items.append({"label": str(opt.get("label", opt.get("value", ""))),
                              "value": opt.get("value", opt.get("label", ""))})
            else:
                items.append({"label": str(opt), "value": opt})

    elif t in ("site_management", "sites", "vendor_management", "vendors",
               "master_data", "logged_in_user", "logged_in_vendor"):
        # Re-use the existing /lookup/options endpoint logic by delegating:
        # the front-end can still call /api/lookup/options directly, but this
        # endpoint also handles these for symmetry.
        column = src.get("display_column") or src.get("display") or "site_name"
        source_key = (
            "sites" if t in ("site_management", "sites") else
            "vendors" if t in ("vendor_management", "vendors") else
            f"master:{src.get('table') or src.get('table_name', '')}" if t == "master_data" else
            t
        )
        # We'll mimic distinct() semantics here:
        if source_key == "sites":
            from vendor_routes import _site_filter_for_user
            flt = _site_filter_for_user(user, src.get("show_all_sites", False))
            rows = await db.sites.find(flt, {"_id": 0, "password_hash": 0}).to_list(100000)
            for row in rows:
                items.append({"label": str(row.get(column, "")),
                              "value": row.get(src.get("value_column") or column),
                              "row": row})
        elif source_key == "vendors":
            rows = await db.vendors.find({}, {"_id": 0}).to_list(10000)
            for row in rows:
                items.append({"label": str(row.get(column, "")),
                              "value": row.get(src.get("value_column") or column),
                              "row": row})
        elif source_key.startswith("master:"):
            table = source_key.split(":", 1)[1]
            rows = await db.master_data.find({"table": table}, {"_id": 0}).to_list(50000)
            for row in rows:
                data = row.get("data", {})
                items.append({"label": str(data.get(column, "")),
                              "value": data.get(src.get("value_column") or column),
                              "row": data})
        elif source_key == "logged_in_user":
            d = {"name": user.name, "email": user.email, "role": user.role,
                 "user_id": user.user_id, "vendor_id": getattr(user, "vendor_id", None)}
            items.append({"label": str(d.get(column, "")), "value": d.get(column), "row": d})
        elif source_key == "logged_in_vendor":
            vid = getattr(user, "vendor_id", None)
            v = await db.vendors.find_one({"vendor_id": vid}, {"_id": 0}) if vid else None
            if v:
                items.append({"label": str(v.get(column, "")), "value": v.get(column), "row": v})

    elif t == "json":
        try:
            data = json.loads(src.get("json_text") or "[]")
        except json.JSONDecodeError as e:
            raise HTTPException(400, f"Invalid JSON: {e}")
        disp = src.get("display_column"); val = src.get("value_column") or disp
        if isinstance(data, list):
            for row in data:
                if isinstance(row, dict):
                    label = row.get(disp) if disp else next(iter(row.values()), "")
                    items.append({"label": str(label),
                                  "value": row.get(val) if val else label, "row": row})
                else:
                    items.append({"label": str(row), "value": row})

    elif t == "csv":
        text = src.get("csv_text") or ""
        reader = _csv.DictReader(io.StringIO(text))
        rows = list(reader)
        disp = src.get("display_column") or (reader.fieldnames[0] if reader.fieldnames else None)
        val = src.get("value_column") or disp
        for row in rows:
            items.append({"label": str(row.get(disp, "")),
                          "value": row.get(val), "row": row})

    elif t == "rest_api":
        url = src.get("url")
        if not url: raise HTTPException(400, "REST API url is required")
        try:
            resp = requests.get(url, headers=src.get("headers") or {}, timeout=12)
            resp.raise_for_status()
            data = resp.json()
        except Exception as e:
            raise HTTPException(400, f"REST API error: {e}")
        path = src.get("json_path")
        if path:
            for part in path.split("."):
                if isinstance(data, dict): data = data.get(part)
        rows = data if isinstance(data, list) else []
        disp = src.get("display_column"); val = src.get("value_column") or disp
        for row in rows:
            if isinstance(row, dict):
                label = row.get(disp) if disp else next(iter(row.values()), "")
                items.append({"label": str(label),
                              "value": row.get(val) if val else label, "row": row})
            else:
                items.append({"label": str(row), "value": row})

    elif t == "excel":
        for row in src.get("rows") or []:
            disp = src.get("display_column"); val = src.get("value_column") or disp
            label = row.get(disp) if disp else next(iter(row.values()), "")
            items.append({"label": str(label),
                          "value": row.get(val) if val else label, "row": row})

    elif t == "another_form":
        form_id = src.get("form_id"); col = src.get("column") or "id"
        if form_id:
            subs = await db.submissions.find({"form_id": form_id}, {"_id": 0}).to_list(10000)
            seen = set()
            for s in subs:
                v = s.get("values", {}).get(col)
                if v is None or v in seen: continue
                seen.add(v)
                items.append({"label": str(v), "value": v, "row": s.get("values", {})})

    elif t == "workflow_variable":
        # Workflow variables are evaluated at runtime, not at design time —
        # we surface the variable name as a single option placeholder.
        name = src.get("variable")
        if name:
            items.append({"label": f"${{{name}}}", "value": f"${{{name}}}"})

    elif t == "sql":
        raise HTTPException(501, "SQL data source requires a relational database; "
                                  "not enabled in this build. Use REST API or Master Data.")

    else:
        raise HTTPException(400, f"Unknown data source type: {t}")

    return {"items": items}
