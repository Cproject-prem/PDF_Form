"""
FormForge — Workflow Automation Engine
======================================

A visual workflow engine modelled on Power Automate / Zapier / Jotform
Approvals. Triggers fire workflow executions; the runner walks the node graph
applying conditions and dispatching actions (email, approval, submission
update, REST call, delay, PDF generate, etc).

Public surface:
    build_workflow_routers(db, get_current_user, jwt_module, jwt_secret,
                            jwt_algo, jwt_expire_hours)
        -> (workflows_router, approvals_router, audit_router,
            workflow_analytics_router, smtp_router, public_approval_router)

    fire_trigger(db, event, payload, user_id=None)
        Awaitable hook used by the rest of the app (form submit, pdf submit,
        user login, …) to enqueue workflow executions.
"""
from __future__ import annotations

import asyncio
import logging
import os
import re
import smtplib
import ssl
import uuid
from datetime import datetime, timezone, timedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from typing import Any, Dict, List, Optional, Tuple

import httpx
from fastapi import APIRouter, Depends, HTTPException, Request
from pydantic import BaseModel, Field, ConfigDict
from simpleeval import SimpleEval, InvalidExpression, FunctionNotDefined, NameNotDefined

log = logging.getLogger("workflow")

# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------

class WorkflowNode(BaseModel):
    """A node on the visual canvas.

    `kind` is the engine category (trigger/condition/action/approval/delay/end).
    `type` is the concrete operation inside that category (e.g. action.send_email,
    trigger.form_submitted, approval.sequential).
    `config` is the user-edited parameters for that node.
    `position` is the canvas {x, y} (free-form; tolerated as dict).
    """
    model_config = ConfigDict(extra="allow")
    id: str
    kind: str
    type: str
    config: Dict[str, Any] = Field(default_factory=dict)
    position: Dict[str, float] = Field(default_factory=lambda: {"x": 0, "y": 0})
    label: Optional[str] = None


class WorkflowEdge(BaseModel):
    model_config = ConfigDict(extra="allow")
    id: str
    source: str
    target: str
    sourceHandle: Optional[str] = None  # "true" / "false" for condition nodes
    targetHandle: Optional[str] = None
    label: Optional[str] = None


class Workflow(BaseModel):
    workflow_id: str
    name: str
    description: str = ""
    status: str = "draft"  # draft | published | disabled
    version: int = 1
    nodes: List[WorkflowNode] = []
    edges: List[WorkflowEdge] = []
    triggers: List[Dict[str, Any]] = []  # cached list of {event, filter} for fast lookup
    owner_id: str
    permissions: Dict[str, List[str]] = Field(default_factory=dict)  # role -> [allow]
    is_template: bool = False
    template_slug: Optional[str] = None
    created_at: str
    updated_at: str


class WorkflowIn(BaseModel):
    name: str
    description: str = ""
    nodes: List[Dict[str, Any]] = []
    edges: List[Dict[str, Any]] = []
    status: Optional[str] = None
    permissions: Optional[Dict[str, List[str]]] = None


class ExecutionLogEntry(BaseModel):
    timestamp: str
    node_id: Optional[str] = None
    level: str = "info"  # info | warn | error | debug
    message: str
    data: Dict[str, Any] = Field(default_factory=dict)


class WorkflowExecution(BaseModel):
    execution_id: str
    workflow_id: str
    workflow_version: int
    trigger_event: str
    trigger_payload: Dict[str, Any] = Field(default_factory=dict)
    status: str = "running"  # running | success | failed | waiting_approval | cancelled
    variables: Dict[str, Any] = Field(default_factory=dict)
    current_node_id: Optional[str] = None
    logs: List[ExecutionLogEntry] = Field(default_factory=list)
    started_at: str
    finished_at: Optional[str] = None
    duration_ms: Optional[int] = None
    error: Optional[str] = None


class ApprovalStep(BaseModel):
    approval_id: str
    execution_id: str
    workflow_id: str
    node_id: str
    submission_id: Optional[str] = None
    subject: str
    description: str = ""
    approvers: List[str] = []                 # list of user_id (or emails)
    mode: str = "sequential"                  # sequential | parallel | any
    current_index: int = 0                    # for sequential mode
    status: str = "pending"                   # pending | approved | rejected | returned | escalated | timeout
    decisions: List[Dict[str, Any]] = []      # [{approver, decision, comment, at, ip}]
    due_at: Optional[str] = None
    created_at: str
    updated_at: str


class ApprovalAction(BaseModel):
    decision: str  # approve | reject | return
    comment: str = ""
    signature: Optional[str] = None  # base64 PNG data URL


class SmtpConfig(BaseModel):
    host: str = ""
    port: int = 587
    username: str = ""
    password: str = ""
    use_tls: bool = True
    from_email: str = ""
    from_name: str = "FormForge"
    enabled: bool = False


class EmailRequest(BaseModel):
    to: List[str]
    cc: List[str] = []
    bcc: List[str] = []
    subject: str
    body_html: str
    reply_to: Optional[str] = None
    attachments: List[Dict[str, Any]] = Field(default_factory=list)  # [{filename, path, content?, mimetype}]


class AuditEntry(BaseModel):
    audit_id: str
    actor_id: Optional[str] = None
    actor_email: Optional[str] = None
    action: str
    target_type: str
    target_id: Optional[str] = None
    details: Dict[str, Any] = Field(default_factory=dict)
    ip: Optional[str] = None
    created_at: str


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _gen(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:12]}"


def _clean(doc: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if doc is None:
        return None
    doc.pop("_id", None)
    return doc


def _resolve_variables(value: Any, vars: Dict[str, Any]) -> Any:
    """Replace `{{path.to.value}}` placeholders inside strings (and recursively
    inside dicts/lists) with values from `vars` using dotted-key lookup."""
    if isinstance(value, str):
        def repl(m: re.Match) -> str:
            key = m.group(1).strip()
            return str(_lookup(vars, key))
        return re.sub(r"\{\{\s*([^}]+?)\s*\}\}", repl, value)
    if isinstance(value, list):
        return [_resolve_variables(v, vars) for v in value]
    if isinstance(value, dict):
        return {k: _resolve_variables(v, vars) for k, v in value.items()}
    return value


def _lookup(obj: Any, dotted: str) -> Any:
    cur: Any = obj
    for part in dotted.split("."):
        if cur is None:
            return ""
        if isinstance(cur, dict):
            cur = cur.get(part, "")
        elif isinstance(cur, list):
            try:
                cur = cur[int(part)]
            except (ValueError, IndexError):
                return ""
        else:
            return ""
    return "" if cur is None else cur


# ---------------------------------------------------------------------------
# Formula engine
# ---------------------------------------------------------------------------

class FormulaEngine:
    """Lightweight, sandboxed formula evaluator backed by `simpleeval`."""

    def __init__(self):
        self.s = SimpleEval()
        self.s.functions.update({
            # text
            "UPPER": lambda x: str(x).upper(),
            "LOWER": lambda x: str(x).lower(),
            "LEN": lambda x: len(str(x)),
            "TRIM": lambda x: str(x).strip(),
            "CONCAT": lambda *a: "".join(str(x) for x in a),
            "REPLACE": lambda s, a, b: str(s).replace(a, b),
            "CONTAINS": lambda h, n: n in str(h),
            "STARTSWITH": lambda h, n: str(h).startswith(str(n)),
            "ENDSWITH": lambda h, n: str(h).endswith(str(n)),
            "LEFT": lambda s, n: str(s)[: int(n)],
            "RIGHT": lambda s, n: str(s)[-int(n) :] if int(n) > 0 else "",
            # numeric
            "ABS": abs,
            "ROUND": round,
            "MIN": min,
            "MAX": max,
            "SUM": lambda *a: sum(_flatten_nums(a)),
            "AVG": lambda *a: (sum(_flatten_nums(a)) / max(len(list(_flatten_nums(a))), 1)),
            "INT": int,
            "FLOAT": float,
            "NUMBER": _to_num,
            # logical
            "IF": lambda c, a, b: a if c else b,
            "AND": lambda *a: all(a),
            "OR": lambda *a: any(a),
            "NOT": lambda a: not a,
            "EQ": lambda a, b: a == b,
            "NE": lambda a, b: a != b,
            # date
            "NOW": lambda: datetime.now(timezone.utc).isoformat(),
            "TODAY": lambda: datetime.now(timezone.utc).date().isoformat(),
            "YEAR": lambda d: _parse_date(d).year,
            "MONTH": lambda d: _parse_date(d).month,
            "DAY": lambda d: _parse_date(d).day,
            "DATEDIFF": lambda a, b: (_parse_date(a) - _parse_date(b)).days,
            "ADDDAYS": lambda d, n: (_parse_date(d) + timedelta(days=int(n))).isoformat(),
            # financial
            "TAX": lambda amount, rate: float(amount) * float(rate) / 100.0,
            "DISCOUNT": lambda amount, rate: float(amount) * (1 - float(rate) / 100.0),
            "PERCENT": lambda part, whole: 0 if float(whole) == 0 else (float(part) / float(whole) * 100.0),
        })

    def evaluate(self, expr: str, context: Dict[str, Any]) -> Any:
        self.s.names = {k: _to_num_if_possible(v) for k, v in context.items()}
        try:
            return self.s.eval(expr)
        except (InvalidExpression, FunctionNotDefined, NameNotDefined, SyntaxError, ZeroDivisionError) as e:
            raise ValueError(f"Formula error: {e}")


def _flatten_nums(items):
    for it in items:
        if isinstance(it, (list, tuple)):
            for sub in _flatten_nums(it):
                yield sub
        else:
            try:
                yield float(it)
            except (TypeError, ValueError):
                continue


def _to_num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _to_num_if_possible(v):
    if isinstance(v, (int, float, bool)):
        return v
    if isinstance(v, str):
        try:
            if "." in v:
                return float(v)
            return int(v)
        except (TypeError, ValueError):
            return v
    return v


def _parse_date(d):
    if isinstance(d, datetime):
        return d
    try:
        return datetime.fromisoformat(str(d).replace("Z", "+00:00"))
    except (TypeError, ValueError):
        return datetime.now(timezone.utc)


formula = FormulaEngine()


# ---------------------------------------------------------------------------
# Condition evaluation
# ---------------------------------------------------------------------------

def _coerce(a):
    if isinstance(a, str):
        try:
            return float(a) if "." in a else int(a)
        except ValueError:
            return a
    return a


def _eval_rule(rule: Dict[str, Any], vars: Dict[str, Any]) -> bool:
    """Evaluate a single rule `{left, op, right}` against `vars`.
    `left` is a dotted variable path; `right` is a literal or `{{var}}` string.
    """
    raw_left = rule.get("left", "")
    raw_right = rule.get("right", "")
    op = (rule.get("op") or "eq").lower()

    left = _lookup(vars, raw_left) if isinstance(raw_left, str) else raw_left
    right = _resolve_variables(raw_right, vars) if isinstance(raw_right, str) else raw_right
    left_n, right_n = _coerce(left), _coerce(right)

    if op in ("eq", "="):
        return str(left) == str(right)
    if op in ("ne", "!="):
        return str(left) != str(right)
    if op == ">":
        return _to_num(left_n) > _to_num(right_n)
    if op == "<":
        return _to_num(left_n) < _to_num(right_n)
    if op == ">=":
        return _to_num(left_n) >= _to_num(right_n)
    if op == "<=":
        return _to_num(left_n) <= _to_num(right_n)
    if op == "contains":
        return str(right) in str(left)
    if op == "not_contains":
        return str(right) not in str(left)
    if op == "starts_with":
        return str(left).startswith(str(right))
    if op == "ends_with":
        return str(left).endswith(str(right))
    if op == "in":
        items = right if isinstance(right, list) else [s.strip() for s in str(right).split(",")]
        return str(left) in [str(x) for x in items]
    if op == "not_in":
        items = right if isinstance(right, list) else [s.strip() for s in str(right).split(",")]
        return str(left) not in [str(x) for x in items]
    if op == "between":
        lo, hi = (right if isinstance(right, list) else [s.strip() for s in str(right).split(",")])[:2]
        return _to_num(lo) <= _to_num(left_n) <= _to_num(hi)
    if op == "empty":
        return left in (None, "", [], {})
    if op == "not_empty":
        return left not in (None, "", [], {})
    if op == "exists":
        return left not in (None, "")
    if op == "is_true":
        return bool(left)
    if op == "is_false":
        return not bool(left)
    return False


def _eval_group(group: Dict[str, Any], vars: Dict[str, Any]) -> bool:
    """Recursively evaluate `{combinator: 'and'|'or'|'not', rules: [...] }`.
    Each rule is either another group or a leaf {left, op, right}.
    """
    if not group:
        return True
    combinator = (group.get("combinator") or "and").lower()
    rules = group.get("rules") or []
    if combinator == "not":
        if not rules:
            return True
        return not _eval_group(rules[0], vars) if isinstance(rules[0], dict) and "combinator" in rules[0] else not _eval_rule(rules[0], vars)
    results = []
    for r in rules:
        if isinstance(r, dict) and "combinator" in r:
            results.append(_eval_group(r, vars))
        else:
            results.append(_eval_rule(r, vars))
    if combinator == "or":
        return any(results) if results else True
    return all(results) if results else True


# ---------------------------------------------------------------------------
# Email
# ---------------------------------------------------------------------------

async def _load_smtp(db) -> Optional[SmtpConfig]:
    doc = _clean(await db.smtp_config.find_one({"_id": "smtp"}))
    if not doc:
        return None
    doc.pop("_id", None)
    return SmtpConfig(**doc)


async def _send_email(db, req: EmailRequest, execution_id: Optional[str] = None) -> Dict[str, Any]:
    """Queue + send an email. Always recorded in `email_queue` regardless of
    whether SMTP is enabled (so users can audit what would have been sent)."""
    eid = _gen("eml")
    queue_doc = {
        "email_id": eid,
        "to": req.to, "cc": req.cc, "bcc": req.bcc,
        "subject": req.subject, "body_html": req.body_html,
        "reply_to": req.reply_to, "execution_id": execution_id,
        "status": "queued", "created_at": _now(), "attempts": 0,
    }
    await db.email_queue.insert_one(dict(queue_doc))
    cfg = await _load_smtp(db)
    if not cfg or not cfg.enabled or not cfg.host:
        await db.email_queue.update_one({"email_id": eid}, {"$set": {"status": "skipped_no_smtp"}})
        return {"email_id": eid, "status": "skipped_no_smtp"}
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = req.subject
        msg["From"] = f"{cfg.from_name} <{cfg.from_email or cfg.username}>"
        msg["To"] = ", ".join(req.to)
        if req.cc:
            msg["Cc"] = ", ".join(req.cc)
        if req.reply_to:
            msg["Reply-To"] = req.reply_to
        msg.attach(MIMEText(req.body_html, "html", "utf-8"))
        # attachments — supports either inline `content` (bytes) or `path` (file on disk)
        for att in (req.attachments or []):
            filename = att.get("filename") or "attachment.bin"
            data: Optional[bytes] = att.get("content")
            if not data and att.get("path"):
                try:
                    with open(att["path"], "rb") as f:
                        data = f.read()
                except OSError as exc:
                    log.warning("attachment missing: %s (%s)", att.get("path"), exc)
                    continue
            if not data:
                continue
            mimetype = att.get("mimetype") or "application/octet-stream"
            maintype, _, subtype = mimetype.partition("/")
            part = MIMEBase(maintype or "application", subtype or "octet-stream")
            part.set_payload(data)
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
            msg.attach(part)
        recipients = list(req.to) + list(req.cc) + list(req.bcc)

        def _do_send():
            if cfg.use_tls:
                ctx = ssl.create_default_context()
                with smtplib.SMTP(cfg.host, cfg.port, timeout=15) as s:
                    s.starttls(context=ctx)
                    if cfg.username:
                        s.login(cfg.username, cfg.password)
                    s.sendmail(cfg.from_email or cfg.username, recipients, msg.as_string())
            else:
                with smtplib.SMTP_SSL(cfg.host, cfg.port, timeout=15) as s:
                    if cfg.username:
                        s.login(cfg.username, cfg.password)
                    s.sendmail(cfg.from_email or cfg.username, recipients, msg.as_string())

        await asyncio.to_thread(_do_send)
        await db.email_queue.update_one({"email_id": eid}, {"$set": {"status": "sent", "sent_at": _now()}})
        return {"email_id": eid, "status": "sent"}
    except Exception as e:  # noqa: BLE001 — surface to logs + queue
        log.exception("smtp send failed")
        await db.email_queue.update_one({"email_id": eid}, {"$set": {"status": "failed", "error": str(e)}})
        return {"email_id": eid, "status": "failed", "error": str(e)}


# ---------------------------------------------------------------------------
# Engine
# ---------------------------------------------------------------------------

class WorkflowEngine:
    """Executes a workflow graph for a given trigger payload."""

    def __init__(self, db):
        self.db = db

    async def fire(self, event: str, payload: Dict[str, Any], user_id: Optional[str] = None) -> List[str]:
        """Find every published workflow whose trigger matches and start it.
        Returns the list of execution_ids that were launched."""
        cursor = self.db.workflows.find({"status": "published", "is_template": {"$ne": True}})
        execution_ids: List[str] = []
        async for wf_doc in cursor:
            wf_doc.pop("_id", None)
            triggers = wf_doc.get("triggers") or []
            matched = next((t for t in triggers if t.get("event") == event and self._trigger_matches(t, payload)), None)
            if not matched:
                # also accept manual/api events when explicitly listed
                continue
            execution_ids.append(await self.start(wf_doc, event, payload, user_id))
        return execution_ids

    def _trigger_matches(self, trigger: Dict[str, Any], payload: Dict[str, Any]) -> bool:
        flt = trigger.get("filter") or {}
        for k, v in flt.items():
            if str(_lookup(payload, k)) != str(v):
                return False
        return True

    async def start(self, wf_doc: Dict[str, Any], event: str, payload: Dict[str, Any],
                    user_id: Optional[str] = None) -> str:
        execution_id = _gen("exec")
        now = _now()
        exec_doc = {
            "execution_id": execution_id,
            "workflow_id": wf_doc["workflow_id"],
            "workflow_version": wf_doc.get("version", 1),
            "trigger_event": event,
            "trigger_payload": payload,
            "status": "running",
            "variables": dict(payload),
            "started_at": now,
            "logs": [],
            "triggered_by": user_id,
        }
        await self.db.workflow_executions.insert_one(dict(exec_doc))
        # Find entry node = the trigger that matched this event
        entry = next(
            (n for n in wf_doc.get("nodes", []) if n.get("kind") == "trigger" and n.get("config", {}).get("event") == event),
            None,
        )
        if not entry:
            await self._finish(execution_id, "failed", error="No matching trigger node")
            return execution_id
        try:
            await self._walk(wf_doc, execution_id, entry["id"], dict(payload))
            await self._finish(execution_id, "success")
        except _WaitApproval as wa:
            await self.db.workflow_executions.update_one(
                {"execution_id": execution_id},
                {"$set": {"status": "waiting_approval", "current_node_id": wa.node_id}},
            )
        except Exception as e:  # noqa: BLE001
            log.exception("workflow execution failed")
            await self._finish(execution_id, "failed", error=str(e))
        return execution_id

    async def resume_from_approval(self, execution_id: str, approval_doc: Dict[str, Any]) -> None:
        ex = _clean(await self.db.workflow_executions.find_one({"execution_id": execution_id}))
        if not ex or ex.get("status") != "waiting_approval":
            return
        wf = _clean(await self.db.workflows.find_one({"workflow_id": ex["workflow_id"]}))
        if not wf:
            return
        # branch by approval outcome
        node_id = ex.get("current_node_id")
        decision_handle = approval_doc.get("status", "approved")  # approved | rejected | returned
        next_handle = "approved" if decision_handle == "approved" else "rejected"
        # merge approval result into variables
        variables = dict(ex.get("variables", {}))
        variables["approval"] = {
            "status": approval_doc.get("status"),
            "decisions": approval_doc.get("decisions", []),
            "approver": (approval_doc.get("decisions") or [{}])[-1].get("approver"),
            "comment": (approval_doc.get("decisions") or [{}])[-1].get("comment", ""),
        }

        # ---- Notify submitter (and their vendor_admin) of the decision ----
        try:
            sid = variables.get("submission_id")
            kind = variables.get("submission_kind")
            submitter_id = None
            submitter_vendor = None
            if sid:
                col = "pdf_submissions" if kind == "pdf" else "submissions"
                sub = await self.db[col].find_one({"submission_id": sid}, {"_id": 0})
                if sub:
                    submitter_id = sub.get("submitted_by")
                    if submitter_id:
                        u = await self.db.users.find_one(
                            {"user_id": submitter_id}, {"_id": 0, "vendor_id": 1},
                        )
                        if u:
                            submitter_vendor = u.get("vendor_id")
            from notifications import create_notification
            title = f"Submission {decision_handle}"
            body = variables["approval"].get("comment") or f"By {variables['approval'].get('approver') or 'reviewer'}"
            link = "/submissions"
            if submitter_id:
                await create_notification(
                    self.db, user_id=submitter_id, kind="approval_decided",
                    title=title, body=body, link=link,
                    submission_id=sid, approval_id=approval_doc.get("approval_id"),
                )
            # Also notify the vendor_admin of the submitter's vendor tree
            if submitter_vendor:
                admins = await self.db.users.find(
                    {"vendor_id": submitter_vendor, "role": "vendor_admin"},
                    {"_id": 0, "user_id": 1},
                ).to_list(50)
                for a in admins:
                    await create_notification(
                        self.db, user_id=a["user_id"], kind="approval_decided",
                        title=title, body=body, link=link,
                        submission_id=sid, approval_id=approval_doc.get("approval_id"),
                    )
        except Exception as _e:  # noqa: BLE001
            log.warning("post-decision notification failed: %s", _e)

        await self.db.workflow_executions.update_one(
            {"execution_id": execution_id},
            {"$set": {"status": "running", "variables": variables}},
        )
        try:
            edges = [WorkflowEdge(**e) for e in wf.get("edges", [])]
            outgoing = [e for e in edges if e.source == node_id and (e.sourceHandle or "approved") == next_handle]
            if not outgoing and decision_handle != "approved":
                # fall back to default outgoing edge if no specific rejected branch
                outgoing = [e for e in edges if e.source == node_id]
            for edge in outgoing:
                await self._walk(wf, execution_id, edge.target, variables)
            await self._finish(execution_id, "success")
        except _WaitApproval as wa:
            await self.db.workflow_executions.update_one(
                {"execution_id": execution_id},
                {"$set": {"status": "waiting_approval", "current_node_id": wa.node_id}},
            )
        except Exception as e:  # noqa: BLE001
            log.exception("workflow resume failed")
            await self._finish(execution_id, "failed", error=str(e))

    async def _log(self, execution_id: str, node_id: Optional[str], message: str,
                   level: str = "info", data: Optional[Dict[str, Any]] = None) -> None:
        entry = {
            "timestamp": _now(), "node_id": node_id, "level": level,
            "message": message, "data": data or {},
        }
        await self.db.workflow_executions.update_one(
            {"execution_id": execution_id}, {"$push": {"logs": entry}},
        )

    async def _finish(self, execution_id: str, status: str, error: Optional[str] = None) -> None:  # noqa: F811
        ex = _clean(await self.db.workflow_executions.find_one({"execution_id": execution_id}))
        if not ex:
            return
        started = datetime.fromisoformat(ex["started_at"])
        duration = int((datetime.now(timezone.utc) - started).total_seconds() * 1000)
        upd = {"status": status, "finished_at": _now(), "duration_ms": duration}
        if error:
            upd["error"] = error
        await self.db.workflow_executions.update_one({"execution_id": execution_id}, {"$set": upd})

    async def _walk(self, wf: Dict[str, Any], execution_id: str, node_id: str,
                    variables: Dict[str, Any]) -> None:
        """Depth-first execution. Each node may modify `variables`; conditions
        determine which outgoing edges to follow."""
        visited: set = set()
        stack: List[Tuple[str, str]] = [(node_id, "default")]  # (node_id, incoming handle)
        nodes_by_id = {n["id"]: n for n in wf.get("nodes", [])}
        edges = [WorkflowEdge(**e) for e in wf.get("edges", [])]
        while stack:
            cur_id, incoming_handle = stack.pop(0)
            if cur_id in visited:
                continue
            visited.add(cur_id)
            node = nodes_by_id.get(cur_id)
            if not node:
                continue
            await self.db.workflow_executions.update_one(
                {"execution_id": execution_id}, {"$set": {"current_node_id": cur_id}},
            )
            outgoing_handles = await self._run_node(execution_id, node, variables)
            await self.db.workflow_executions.update_one(
                {"execution_id": execution_id}, {"$set": {"variables": variables}},
            )
            if outgoing_handles is None:  # end node
                continue
            for edge in edges:
                if edge.source != cur_id:
                    continue
                handle = edge.sourceHandle or "default"
                if handle in outgoing_handles or "default" in outgoing_handles:
                    stack.append((edge.target, handle))

    async def _run_node(self, execution_id: str, node: Dict[str, Any],
                        variables: Dict[str, Any]) -> Optional[List[str]]:
        """Run a single node. Return the list of outgoing handles to follow
        (e.g. ['true'] for condition true branch, ['default'] for normal flow,
        or None to stop traversal from this node).
        """
        kind = node.get("kind")
        type_ = node.get("type")
        cfg = _resolve_variables(node.get("config", {}) or {}, variables)
        await self._log(execution_id, node["id"], f"{kind}.{type_} running", data={"config": cfg})

        if kind == "trigger":
            return ["default"]

        if kind == "condition" or type_ == "condition.if":
            ok = _eval_group(cfg.get("group") or {}, variables)
            return ["true"] if ok else ["false"]

        if kind == "end" or type_ == "logic.end":
            return None

        if type_ == "logic.delay":
            seconds = float(cfg.get("seconds", 0) or 0)
            seconds += float(cfg.get("minutes", 0) or 0) * 60
            seconds += float(cfg.get("hours", 0) or 0) * 3600
            # cap delay so we don't block the request indefinitely
            await asyncio.sleep(min(seconds, 30))
            return ["default"]

        if type_ == "action.send_email":
            to = _split_emails(cfg.get("to", ""))
            cc = _split_emails(cfg.get("cc", ""))
            bcc = _split_emails(cfg.get("bcc", ""))
            attachments_cfg = cfg.get("attachments") or []
            attachments: List[Dict[str, Any]] = []

            sid = variables.get("submission_id")
            sub_kind = variables.get("submission_kind")  # "pdf" or "form"
            form_id = variables.get("form_id")
            template_id = variables.get("template_id") or (form_id if sub_kind == "pdf" else None)

            # --- Resolve the submission + parent form/template up front ---
            sub_doc: Optional[Dict[str, Any]] = None
            parent: Optional[Dict[str, Any]] = None
            if sid:
                if sub_kind == "pdf":
                    sub_doc = await self.db.pdf_submissions.find_one({"submission_id": sid}, {"_id": 0})
                    if sub_doc:
                        parent = await self.db.pdf_templates.find_one(
                            {"template_id": sub_doc.get("template_id")}, {"_id": 0},
                        )
                else:
                    sub_doc = await self.db.submissions.find_one({"submission_id": sid}, {"_id": 0})
                    if sub_doc:
                        parent = await self.db.forms.find_one(
                            {"form_id": sub_doc.get("form_id")}, {"_id": 0},
                        )

            upload_dir = os.environ.get("UPLOAD_DIR", "/app/backend/uploads")
            slug_base = (parent or {}).get("slug") or sid or "submission"

            # 1) Completed PDF (PDF-form submissions only)
            if "completed_pdf" in attachments_cfg and sub_doc and sub_doc.get("completed_filename"):
                path = f"{upload_dir}/completed/{sub_doc['completed_filename']}"
                attachments.append({
                    "filename": f"completed-{sid}.pdf",
                    "path": path,
                    "mimetype": "application/pdf",
                })

            # 2) Original PDF template file (PDF-form submissions only)
            if "original_pdf" in attachments_cfg and parent and parent.get("storage_filename"):
                path = f"{upload_dir}/pdf/{parent['storage_filename']}"
                attachments.append({
                    "filename": parent.get("original_filename") or f"original-{template_id}.pdf",
                    "path": path,
                    "mimetype": "application/pdf",
                })

            # 3) Excel export of THIS submission's row
            if "excel_export" in attachments_cfg and sub_doc and parent:
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill
                    skip = ("heading", "paragraph", "static_text", "divider", "hidden")
                    fields = [f for f in (parent.get("fields") or []) if f.get("type") not in skip]
                    ids = [f["id"] for f in fields]
                    labels = {f["id"]: (f.get("label") or f.get("name") or f["id"]) for f in fields}
                    wb = Workbook()
                    ws = wb.active
                    ws.title = ((parent.get("title") or "Submission")[:31])
                    hf = Font(bold=True, color="FFFFFF")
                    hb = PatternFill("solid", fgColor="2563EB")
                    ws.append(["Submission ID", "Status", "Submitted At"] + [labels[i] for i in ids])
                    ws.append(["submission_id", "status", "created_at"] + ids)
                    for cell in ws[1]:
                        cell.font = hf
                        cell.fill = hb
                    for cell in ws[2]:
                        cell.font = Font(italic=True, color="475569")
                    vals = sub_doc.get("values") or {}

                    def _v(x):
                        if x is None:
                            return ""
                        if isinstance(x, (list, tuple)):
                            return ", ".join(str(y) for y in x)
                        if isinstance(x, dict):
                            return x.get("filename") or str(x)
                        if isinstance(x, str) and x.startswith("data:image"):
                            return "[signature image]"
                        return x
                    ws.append(
                        [sub_doc.get("submission_id"), sub_doc.get("status"), sub_doc.get("created_at")] +
                        [_v(vals.get(i, "")) for i in ids],
                    )
                    from io import BytesIO
                    buf = BytesIO()
                    wb.save(buf)
                    attachments.append({
                        "filename": f"{slug_base}-{sid}.xlsx",
                        "content": buf.getvalue(),
                        "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    })
                except Exception as exc:  # noqa: BLE001
                    log.warning("excel attachment generation failed: %s", exc)

            # 4) CSV export of THIS submission's row
            if "csv_export" in attachments_cfg and sub_doc and parent:
                try:
                    import csv as _csv
                    import io as _io
                    skip = ("heading", "paragraph", "static_text", "divider", "hidden")
                    fields = [f for f in (parent.get("fields") or []) if f.get("type") not in skip]
                    ids = [f["id"] for f in fields]
                    labels = {f["id"]: (f.get("label") or f.get("name") or f["id"]) for f in fields}
                    sio = _io.StringIO()
                    w = _csv.writer(sio)
                    w.writerow(["submission_id", "status", "created_at"] + [labels[i] for i in ids])
                    vals = sub_doc.get("values") or {}
                    def _csv_v(x):
                        if x is None:
                            return ""
                        if isinstance(x, (list, tuple)):
                            return ", ".join(str(y) for y in x)
                        if isinstance(x, dict):
                            return x.get("filename") or str(x)
                        return x
                    w.writerow(
                        [sub_doc.get("submission_id"), sub_doc.get("status"), sub_doc.get("created_at")] +
                        [_csv_v(vals.get(i, "")) for i in ids],
                    )
                    attachments.append({
                        "filename": f"{slug_base}-{sid}.csv",
                        "content": sio.getvalue().encode("utf-8"),
                        "mimetype": "text/csv",
                    })
                except Exception as exc:  # noqa: BLE001
                    log.warning("csv attachment generation failed: %s", exc)

            # 5) ZIP bundle of the above (Completed PDF + Original PDF + Excel + CSV)
            if "zip_archive" in attachments_cfg and attachments:
                try:
                    import zipfile
                    from io import BytesIO
                    zbuf = BytesIO()
                    with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as zf:
                        for att in attachments:
                            data = att.get("content")
                            if not data and att.get("path"):
                                try:
                                    with open(att["path"], "rb") as fh:
                                        data = fh.read()
                                except OSError:
                                    continue
                            if data:
                                zf.writestr(att["filename"], data)
                    attachments = [{
                        "filename": f"{slug_base}-{sid}.zip",
                        "content": zbuf.getvalue(),
                        "mimetype": "application/zip",
                    }]
                except Exception as exc:  # noqa: BLE001
                    log.warning("zip attachment generation failed: %s", exc)

            req = EmailRequest(
                to=to, cc=cc, bcc=bcc,
                subject=str(cfg.get("subject", "")),
                body_html=str(cfg.get("body", "")),
                reply_to=cfg.get("reply_to") or None,
                attachments=attachments,
            )
            result = await _send_email(self.db, req, execution_id=execution_id)
            await self._log(execution_id, node["id"], f"email {result['status']}", data=result)
            variables.setdefault("last_email", result)
            return ["default"]

        if type_ == "action.update_submission":
            sid = cfg.get("submission_id") or variables.get("submission_id")
            if sid:
                upd = cfg.get("set") or {}
                col = "pdf_submissions" if (variables.get("submission_kind") == "pdf") else "submissions"
                await self.db[col].update_one({"submission_id": sid}, {"$set": upd})
                await self._log(execution_id, node["id"], f"submission {sid} updated", data={"set": upd})
            return ["default"]

        if type_ == "action.set_status":
            sid = cfg.get("submission_id") or variables.get("submission_id")
            new_status = cfg.get("status", "approved")
            if sid:
                col = "pdf_submissions" if (variables.get("submission_kind") == "pdf") else "submissions"
                await self.db[col].update_one({"submission_id": sid}, {"$set": {"status": new_status}})
            return ["default"]

        if type_ == "action.formula":
            try:
                value = formula.evaluate(str(cfg.get("expression", "")), dict(variables))
                target = cfg.get("output", "formula_result")
                variables[target] = value
                await self._log(execution_id, node["id"], f"formula -> {target}={value}")
            except ValueError as e:
                await self._log(execution_id, node["id"], str(e), level="error")
            return ["default"]

        if type_ == "action.http":
            try:
                method = (cfg.get("method") or "POST").upper()
                url = cfg.get("url", "")
                headers = cfg.get("headers") or {}
                body = cfg.get("body")
                async with httpx.AsyncClient(timeout=20) as cli:
                    resp = await cli.request(method, url, headers=headers, json=body if isinstance(body, (dict, list)) else None, content=None if isinstance(body, (dict, list)) else body)
                variables["last_http"] = {"status": resp.status_code, "body": _safe_json(resp)}
                await self._log(execution_id, node["id"], f"http {method} {url} -> {resp.status_code}")
            except Exception as e:  # noqa: BLE001
                await self._log(execution_id, node["id"], f"http failed: {e}", level="error")
                variables["last_http"] = {"error": str(e)}
            return ["default"]

        if type_ == "action.set_variable":
            target = cfg.get("name", "var")
            variables[target] = cfg.get("value", "")
            return ["default"]

        if type_ == "approval.sequential" or type_ == "approval.parallel" or kind == "approval":
            approval = await self._create_approval(execution_id, node, cfg, variables)
            await self._log(execution_id, node["id"], f"approval requested: {approval['approval_id']}")
            raise _WaitApproval(node["id"], approval["approval_id"])

        if type_ == "action.audit":
            await self.db.audit_logs.insert_one({
                "audit_id": _gen("aud"),
                "actor_id": variables.get("user_id"),
                "actor_email": variables.get("user_email"),
                "action": cfg.get("action", "workflow_audit"),
                "target_type": cfg.get("target_type", "workflow"),
                "target_id": variables.get("submission_id"),
                "details": {"message": cfg.get("message", ""), **dict(variables)},
                "ip": variables.get("ip"),
                "created_at": _now(),
            })
            return ["default"]

        # Unknown — log and continue
        await self._log(execution_id, node["id"], f"unknown node type '{type_}', skipping", level="warn")
        return ["default"]

    async def _create_approval(self, execution_id: str, node: Dict[str, Any],
                               cfg: Dict[str, Any], variables: Dict[str, Any]) -> Dict[str, Any]:
        # ---- Resolve approvers ----
        # New behavior (Section 9): if the approval node is configured with
        # `auto_from_site: true` (default), we auto-populate the approver email
        # from the site's `approver_email` column, based on the site name
        # inside the triggering submission. Manual `approvers` still work as
        # a fallback (super-admin can override in the workflow config).
        approvers = cfg.get("approvers") or []
        if isinstance(approvers, str):
            approvers = [a.strip() for a in approvers.split(",") if a.strip()]
        cc_list = cfg.get("cc") or []
        if isinstance(cc_list, str):
            cc_list = [a.strip() for a in cc_list.split(",") if a.strip()]

        auto_from_site = cfg.get("auto_from_site", True)
        site_name = variables.get("site_name")
        if not approvers and auto_from_site and site_name:
            site = await self.db.sites.find_one(
                {"$or": [{"site_name": site_name}, {"asset_id": site_name}, {"site_code": site_name}]},
                {"_id": 0, "approver_email": 1, "site_name": 1},
            )
            if site and site.get("approver_email"):
                approvers = [site["approver_email"]]
                variables.setdefault("resolved_approver", site["approver_email"])

        mode = "sequential" if node.get("type") == "approval.sequential" else (cfg.get("mode") or "sequential")
        due_days = int(cfg.get("due_days", 0) or 0)
        due_at = (datetime.now(timezone.utc) + timedelta(days=due_days)).isoformat() if due_days else None
        approval = {
            "approval_id": _gen("apv"),
            "execution_id": execution_id,
            "workflow_id": (await self.db.workflow_executions.find_one({"execution_id": execution_id}, {"_id": 0, "workflow_id": 1}))["workflow_id"],
            "node_id": node["id"],
            "submission_id": variables.get("submission_id"),
            "submission_kind": variables.get("submission_kind"),
            "form_id": variables.get("form_id"),
            "form_name": variables.get("form_name"),
            "site_name": site_name,
            "subject": cfg.get("subject") or f"Approval needed: {variables.get('form_name') or 'submission'}",
            "description": cfg.get("description") or "",
            "approvers": approvers,
            "cc": cc_list,
            "mode": mode,
            "current_index": 0,
            "status": "pending",
            "decisions": [],
            "due_at": due_at,
            "created_at": _now(),
            "updated_at": _now(),
        }
        await self.db.approvals.insert_one(dict(approval))
        # Send approval emails to current approvers
        recipients = approvers if mode != "sequential" else approvers[:1]
        if recipients:
            base_url = os.environ.get("PUBLIC_BASE_URL") or ""
            for email in recipients:
                token = _gen("aptk")
                await self.db.approval_tokens.insert_one({
                    "token": token, "approval_id": approval["approval_id"],
                    "approver": email, "created_at": _now(),
                })
                # Web page (preferred — user adds a comment, then decides)
                review_url = f"{base_url}/approve/{token}" if base_url else f"/approve/{token}"
                # One-click links (no comment) as a fallback inside the email
                approve_url = f"{base_url}/api/public/approvals/{token}/approve" if base_url else f"/api/public/approvals/{token}/approve"
                reject_url = f"{base_url}/api/public/approvals/{token}/reject" if base_url else f"/api/public/approvals/{token}/reject"
                html = _approval_email_html(approval["subject"], approval["description"], review_url, approve_url, reject_url)
                await _send_email(
                    self.db,
                    EmailRequest(
                        to=[email], cc=list(cc_list),
                        subject=f"[Approval] {approval['subject']}", body_html=html,
                    ),
                    execution_id=execution_id,
                )

        # ---- In-app notifications for each approver + CC recipient ----
        try:
            from notifications import notify_users_by_email
            all_emails = list(dict.fromkeys(list(recipients) + list(cc_list)))
            await notify_users_by_email(
                self.db, all_emails,
                kind="approval_pending",
                title=f"Approval needed: {approval['subject']}",
                body=f"Site: {site_name or '—'} · Form: {variables.get('form_name') or '—'}",
                link="/approvals",
                submission_id=approval.get("submission_id"),
                approval_id=approval["approval_id"],
            )
        except Exception as _e:  # noqa: BLE001
            log.warning("approval notification failed: %s", _e)
        return approval


class _WaitApproval(Exception):
    def __init__(self, node_id: str, approval_id: str):
        self.node_id = node_id
        self.approval_id = approval_id


def _split_emails(raw: Any) -> List[str]:
    if isinstance(raw, list):
        return [str(x).strip() for x in raw if str(x).strip()]
    if not raw:
        return []
    return [s.strip() for s in str(raw).replace(";", ",").split(",") if s.strip()]


def _safe_json(resp) -> Any:
    try:
        return resp.json()
    except Exception:  # noqa: BLE001
        return resp.text[:2000]


def _approval_email_html(subject: str, description: str, review_url: str,
                         approve_url: str, reject_url: str) -> str:
    return f"""
<!doctype html>
<html><body style="font-family:Inter,Arial,sans-serif;background:#f1f5f9;padding:24px;color:#0f172a">
  <div style="max-width:560px;margin:0 auto;background:#fff;border-radius:12px;padding:28px;box-shadow:0 1px 3px rgba(15,23,42,.08)">
    <div style="font-size:11px;letter-spacing:.15em;font-weight:700;color:#2563eb;margin-bottom:8px">APPROVAL REQUEST</div>
    <h1 style="margin:0 0 12px 0;font-size:20px;line-height:1.3">{subject}</h1>
    <p style="margin:0 0 24px 0;color:#475569;font-size:14px;line-height:1.6">{description or 'You have a pending approval. Please review and respond.'}</p>
    <div>
      <a href="{review_url}" style="display:inline-block;padding:10px 20px;background:#2563eb;color:#fff;text-decoration:none;border-radius:8px;font-weight:600;font-size:14px">Open & review</a>
    </div>
    <p style="margin-top:24px;color:#94a3b8;font-size:12px">Quick actions:
      <a href="{approve_url}" style="color:#16a34a">approve</a> ·
      <a href="{reject_url}" style="color:#dc2626">reject</a>
    </p>
    <p style="margin-top:32px;color:#94a3b8;font-size:12px">Powered by FormForge Workflow Automation</p>
  </div>
</body></html>"""


# ---------------------------------------------------------------------------
# Routers
# ---------------------------------------------------------------------------

def build_workflow_routers(db, get_current_user):
    """Returns (workflows, approvals, audit, analytics, smtp, public_approvals)."""

    engine = WorkflowEngine(db)
    workflows = APIRouter(prefix="/workflows", tags=["workflows"])
    approvals = APIRouter(prefix="/approvals", tags=["approvals"])
    audit = APIRouter(prefix="/audit", tags=["audit"])
    analytics = APIRouter(prefix="/workflow-analytics", tags=["analytics"])
    smtp_r = APIRouter(prefix="/settings/smtp", tags=["smtp"])
    public_apv = APIRouter(prefix="/public/approvals", tags=["public-approval"])

    async def _require_admin(user) -> None:
        if user.role not in ("super_admin", "admin"):
            raise HTTPException(403, "Admin role required")

    # ---------- Workflows CRUD ----------

    @workflows.get("")
    async def list_workflows(user=Depends(get_current_user)):
        q = {} if user.role in ("super_admin", "admin") else {"owner_id": user.user_id}
        q["is_template"] = {"$ne": True}
        rows = await db.workflows.find(q, {"_id": 0}).sort("updated_at", -1).to_list(2000)
        return rows

    @workflows.get("/templates")
    async def list_templates(user=Depends(get_current_user)):
        rows = await db.workflows.find({"is_template": True}, {"_id": 0}).to_list(200)
        return rows

    @workflows.post("/templates/{slug}/instantiate")
    async def instantiate_template(slug: str, user=Depends(get_current_user)):
        tpl = _clean(await db.workflows.find_one({"is_template": True, "template_slug": slug}))
        if not tpl:
            raise HTTPException(404, "Template not found")
        wf_id = _gen("wf")
        now = _now()
        doc = {
            "workflow_id": wf_id,
            "name": tpl.get("name", "Untitled"),
            "description": tpl.get("description", ""),
            "status": "draft",
            "version": 1,
            "nodes": tpl.get("nodes", []),
            "edges": tpl.get("edges", []),
            "triggers": _extract_triggers(tpl.get("nodes", [])),
            "owner_id": user.user_id,
            "permissions": {},
            "is_template": False,
            "template_slug": None,
            "created_at": now, "updated_at": now,
        }
        await db.workflows.insert_one(dict(doc))
        return doc

    @workflows.post("")
    async def create_workflow(body: WorkflowIn, user=Depends(get_current_user)):
        await _require_admin(user)
        now = _now()
        wf_id = _gen("wf")
        doc = {
            "workflow_id": wf_id,
            "name": body.name or "Untitled workflow",
            "description": body.description or "",
            "status": body.status or "draft",
            "version": 1,
            "nodes": body.nodes,
            "edges": body.edges,
            "triggers": _extract_triggers(body.nodes),
            "owner_id": user.user_id,
            "permissions": body.permissions or {},
            "is_template": False,
            "created_at": now, "updated_at": now,
        }
        await db.workflows.insert_one(dict(doc))
        await _audit(db, user, "workflow.create", "workflow", wf_id, {"name": doc["name"]})
        return doc

    @workflows.get("/{workflow_id}")
    async def get_workflow(workflow_id: str, user=Depends(get_current_user)):
        wf = _clean(await db.workflows.find_one({"workflow_id": workflow_id}))
        if not wf:
            raise HTTPException(404, "Workflow not found")
        if wf["owner_id"] != user.user_id and user.role not in ("super_admin", "admin"):
            raise HTTPException(403, "Not allowed")
        return wf

    @workflows.put("/{workflow_id}")
    async def update_workflow(workflow_id: str, body: WorkflowIn, user=Depends(get_current_user)):
        wf = _clean(await db.workflows.find_one({"workflow_id": workflow_id}))
        if not wf:
            raise HTTPException(404, "Workflow not found")
        if wf["owner_id"] != user.user_id and user.role not in ("super_admin", "admin"):
            raise HTTPException(403, "Not allowed")
        # snapshot previous version
        await db.workflow_versions.insert_one({
            "snapshot_id": _gen("wfv"),
            "workflow_id": workflow_id, "version": wf.get("version", 1),
            "nodes": wf.get("nodes", []), "edges": wf.get("edges", []),
            "name": wf.get("name"), "description": wf.get("description"),
            "created_at": _now(), "created_by": user.user_id,
        })
        upd = {
            "name": body.name, "description": body.description,
            "nodes": body.nodes, "edges": body.edges,
            "triggers": _extract_triggers(body.nodes),
            "updated_at": _now(),
            "version": int(wf.get("version", 1)) + 1,
        }
        if body.status:
            upd["status"] = body.status
        if body.permissions is not None:
            upd["permissions"] = body.permissions
        await db.workflows.update_one({"workflow_id": workflow_id}, {"$set": upd})
        await _audit(db, user, "workflow.update", "workflow", workflow_id, {"version": upd["version"]})
        merged = {**wf, **upd}
        return merged

    @workflows.patch("/{workflow_id}/status")
    async def set_status(workflow_id: str, body: Dict[str, Any], user=Depends(get_current_user)):
        if body.get("status") not in ("draft", "published", "disabled"):
            raise HTTPException(400, "Invalid status")
        wf = _clean(await db.workflows.find_one({"workflow_id": workflow_id}))
        if not wf:
            raise HTTPException(404, "Not found")
        if wf["owner_id"] != user.user_id and user.role not in ("super_admin", "admin"):
            raise HTTPException(403, "Not allowed")
        await db.workflows.update_one({"workflow_id": workflow_id}, {"$set": {"status": body["status"], "updated_at": _now()}})
        await _audit(db, user, f"workflow.{body['status']}", "workflow", workflow_id, {})
        wf["status"] = body["status"]
        return wf

    @workflows.delete("/{workflow_id}")
    async def delete_workflow(workflow_id: str, user=Depends(get_current_user)):
        wf = _clean(await db.workflows.find_one({"workflow_id": workflow_id}))
        if not wf:
            raise HTTPException(404, "Not found")
        if wf["owner_id"] != user.user_id and user.role not in ("super_admin", "admin"):
            raise HTTPException(403, "Not allowed")
        await db.workflows.delete_one({"workflow_id": workflow_id})
        await _audit(db, user, "workflow.delete", "workflow", workflow_id, {})
        return {"ok": True}

    @workflows.post("/{workflow_id}/duplicate")
    async def duplicate_workflow(workflow_id: str, user=Depends(get_current_user)):
        wf = _clean(await db.workflows.find_one({"workflow_id": workflow_id}))
        if not wf:
            raise HTTPException(404, "Not found")
        new_id = _gen("wf")
        now = _now()
        copy = {**wf, "workflow_id": new_id, "name": (wf.get("name") or "") + " (copy)",
                "status": "draft", "version": 1, "owner_id": user.user_id,
                "created_at": now, "updated_at": now}
        await db.workflows.insert_one(dict(copy))
        return copy

    @workflows.post("/{workflow_id}/test")
    async def test_workflow(workflow_id: str, body: Dict[str, Any], user=Depends(get_current_user)):
        wf = _clean(await db.workflows.find_one({"workflow_id": workflow_id}))
        if not wf:
            raise HTTPException(404, "Not found")
        event = body.get("event") or ((wf.get("triggers") or [{}])[0].get("event") or "manual")
        payload = body.get("payload") or {"test": True, "user_id": user.user_id}
        execution_id = await engine.start(wf, event, payload, user_id=user.user_id)
        await _audit(db, user, "workflow.test", "workflow", workflow_id, {"execution_id": execution_id})
        return {"execution_id": execution_id}

    @workflows.get("/{workflow_id}/executions")
    async def list_executions(workflow_id: str, user=Depends(get_current_user)):
        rows = await db.workflow_executions.find(
            {"workflow_id": workflow_id}, {"_id": 0}
        ).sort("started_at", -1).to_list(500)
        return rows

    @workflows.get("/{workflow_id}/versions")
    async def list_versions(workflow_id: str, user=Depends(get_current_user)):
        rows = await db.workflow_versions.find(
            {"workflow_id": workflow_id}, {"_id": 0}
        ).sort("created_at", -1).to_list(100)
        return rows

    @workflows.post("/formula/evaluate")
    async def eval_formula(body: Dict[str, Any], user=Depends(get_current_user)):
        try:
            result = formula.evaluate(str(body.get("expression", "")), dict(body.get("context") or {}))
            return {"ok": True, "result": result}
        except ValueError as e:
            return {"ok": False, "error": str(e)}

    # ---------- Executions detail ----------

    @workflows.get("/executions/{execution_id}")
    async def get_execution(execution_id: str, user=Depends(get_current_user)):
        ex = _clean(await db.workflow_executions.find_one({"execution_id": execution_id}))
        if not ex:
            raise HTTPException(404, "Execution not found")
        return ex

    # ---------- Approvals ----------

    @approvals.get("")
    async def list_my_approvals(status_filter: Optional[str] = None, user=Depends(get_current_user)):
        identifiers = [user.user_id, user.email]
        q: Dict[str, Any] = {"approvers": {"$in": identifiers}}
        if status_filter:
            q["status"] = status_filter
        rows = await db.approvals.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
        return rows

    @approvals.get("/{approval_id}")
    async def get_approval(approval_id: str, user=Depends(get_current_user)):
        apv = _clean(await db.approvals.find_one({"approval_id": approval_id}))
        if not apv:
            raise HTTPException(404, "Not found")
        return apv

    @approvals.post("/{approval_id}/decide")
    async def decide(approval_id: str, body: ApprovalAction, request: Request, user=Depends(get_current_user)):
        apv = _clean(await db.approvals.find_one({"approval_id": approval_id}))
        if not apv:
            raise HTTPException(404, "Not found")
        if apv.get("status") != "pending":
            raise HTTPException(400, "Approval already finalised")
        identifiers = {user.user_id, user.email}
        if not (identifiers & set(apv.get("approvers", []))):
            raise HTTPException(403, "Not an approver on this step")
        await _record_decision(db, apv, user.email or user.user_id, body, request)
        await _maybe_resume(db, engine, apv["approval_id"])
        return _clean(await db.approvals.find_one({"approval_id": approval_id}))

    # ---------- Public approval links ----------

    @public_apv.get("/{token}")
    async def public_get_approval(token: str):
        tok = _clean(await db.approval_tokens.find_one({"token": token}))
        if not tok:
            raise HTTPException(404, "Invalid link")
        apv = _clean(await db.approvals.find_one({"approval_id": tok["approval_id"]}))
        if not apv:
            raise HTTPException(404, "Approval not found")
        return {"approval": apv, "approver": tok["approver"]}

    @public_apv.post("/{token}/decide")
    async def public_decide(token: str, body: ApprovalAction, request: Request):
        tok = _clean(await db.approval_tokens.find_one({"token": token}))
        if not tok:
            raise HTTPException(404, "Invalid link")
        apv = _clean(await db.approvals.find_one({"approval_id": tok["approval_id"]}))
        if not apv:
            raise HTTPException(404, "Not found")
        if apv.get("status") != "pending":
            return {"ok": True, "message": "Approval already finalised", "status": apv.get("status")}
        await _record_decision(db, apv, tok["approver"], body, request)
        await _maybe_resume(db, engine, apv["approval_id"])
        await db.approval_tokens.delete_one({"token": token})
        return {"ok": True, "status": (await db.approvals.find_one({"approval_id": apv["approval_id"]}, {"_id": 0, "status": 1}))["status"]}

    @public_apv.get("/{token}/approve")
    async def public_quick_approve(token: str, request: Request):
        return await public_decide(token, ApprovalAction(decision="approve"), request)

    @public_apv.get("/{token}/reject")
    async def public_quick_reject(token: str, request: Request):
        return await public_decide(token, ApprovalAction(decision="reject"), request)

    # ---------- Audit logs ----------

    @audit.get("")
    async def list_audit(limit: int = 200, user=Depends(get_current_user)):
        await _require_admin(user)
        rows = await db.audit_logs.find({}, {"_id": 0}).sort("created_at", -1).to_list(min(limit, 1000))
        return rows

    @audit.post("")
    async def write_audit(body: Dict[str, Any], request: Request, user=Depends(get_current_user)):
        await _audit(db, user, body.get("action", "manual"), body.get("target_type", "other"),
                     body.get("target_id"), body.get("details") or {}, ip=request.client.host if request.client else None)
        return {"ok": True}

    # ---------- Analytics ----------

    @analytics.get("")
    async def overview(user=Depends(get_current_user)):
        await _require_admin(user)
        total = await db.workflow_executions.count_documents({})
        ok = await db.workflow_executions.count_documents({"status": "success"})
        failed = await db.workflow_executions.count_documents({"status": "failed"})
        waiting = await db.workflow_executions.count_documents({"status": "waiting_approval"})
        running = await db.workflow_executions.count_documents({"status": "running"})
        pipeline = [{"$group": {"_id": "$status", "avg": {"$avg": "$duration_ms"}}}]
        avg_rows = await db.workflow_executions.aggregate(pipeline).to_list(10)
        avg_ms = next((r["avg"] for r in avg_rows if r["_id"] == "success"), 0) or 0
        # email stats
        em_total = await db.email_queue.count_documents({})
        em_sent = await db.email_queue.count_documents({"status": "sent"})
        em_fail = await db.email_queue.count_documents({"status": "failed"})
        # most-used
        top = await db.workflow_executions.aggregate([
            {"$group": {"_id": "$workflow_id", "count": {"$sum": 1}}},
            {"$sort": {"count": -1}}, {"$limit": 5},
        ]).to_list(5)
        names = {}
        for row in top:
            wf = await db.workflows.find_one({"workflow_id": row["_id"]}, {"_id": 0, "name": 1})
            names[row["_id"]] = (wf or {}).get("name", row["_id"])
        return {
            "total_executions": total,
            "successful": ok,
            "failed": failed,
            "waiting_approval": waiting,
            "running": running,
            "avg_duration_ms": int(avg_ms),
            "email": {"total": em_total, "sent": em_sent, "failed": em_fail,
                      "success_rate": (em_sent / em_total * 100.0) if em_total else 0},
            "most_used": [{"workflow_id": r["_id"], "name": names.get(r["_id"], r["_id"]), "count": r["count"]} for r in top],
        }

    # ---------- SMTP settings ----------

    @smtp_r.get("")
    async def get_smtp(user=Depends(get_current_user)):
        await _require_admin(user)
        doc = _clean(await db.smtp_config.find_one({"_id": "smtp"}))
        if not doc:
            return SmtpConfig().model_dump()
        doc.pop("_id", None)
        # never return password back
        doc["password"] = "*****" if doc.get("password") else ""
        return doc

    @smtp_r.put("")
    async def put_smtp(body: SmtpConfig, user=Depends(get_current_user)):
        await _require_admin(user)
        existing = _clean(await db.smtp_config.find_one({"_id": "smtp"})) or {}
        merged = {**existing, **body.model_dump()}
        if body.password in ("", "*****"):
            merged["password"] = existing.get("password", "")
        await db.smtp_config.update_one({"_id": "smtp"}, {"$set": merged}, upsert=True)
        await _audit(db, user, "smtp.update", "settings", "smtp", {"host": body.host, "enabled": body.enabled})
        merged["password"] = "*****" if merged.get("password") else ""
        return merged

    @smtp_r.post("/test")
    async def smtp_test(body: Dict[str, Any], user=Depends(get_current_user)):
        await _require_admin(user)
        # short-circuit when SMTP is disabled — no point requiring 'to'
        cfg = await _load_smtp(db)
        if not cfg or not cfg.enabled or not cfg.host:
            return {"status": "skipped_no_smtp"}
        to = body.get("to")
        if not to:
            raise HTTPException(400, "Missing 'to'")
        result = await _send_email(
            db,
            EmailRequest(
                to=[to], subject="FormForge SMTP test",
                body_html="<h2>It works!</h2><p>If you're reading this, your SMTP config is good.</p>",
            ),
        )
        return result

    return workflows, approvals, audit, analytics, smtp_r, public_apv


# ---------------------------------------------------------------------------
# Module-level helpers used by routers
# ---------------------------------------------------------------------------

async def _audit(db, user, action: str, target_type: str, target_id: Optional[str],
                 details: Dict[str, Any], ip: Optional[str] = None) -> None:
    await db.audit_logs.insert_one({
        "audit_id": _gen("aud"),
        "actor_id": getattr(user, "user_id", None),
        "actor_email": getattr(user, "email", None),
        "action": action,
        "target_type": target_type,
        "target_id": target_id,
        "details": details or {},
        "ip": ip,
        "created_at": _now(),
    })


async def _record_decision(db, apv: Dict[str, Any], approver: str,
                           body: ApprovalAction, request: Request) -> None:
    decision_norm = (body.decision or "").lower()
    if decision_norm not in ("approve", "reject", "return"):
        raise HTTPException(400, "Invalid decision")
    decision_doc = {
        "approver": approver,
        "decision": decision_norm,
        "comment": body.comment or "",
        "signature": body.signature or None,
        "at": _now(),
        "ip": request.client.host if request.client else None,
    }
    decisions = list(apv.get("decisions", [])) + [decision_doc]
    mode = apv.get("mode", "sequential")
    new_status = apv.get("status", "pending")
    new_index = apv.get("current_index", 0)
    if decision_norm == "reject":
        new_status = "rejected"
    elif decision_norm == "return":
        new_status = "returned"
    elif decision_norm == "approve":
        if mode == "parallel":
            approved_set = {d["approver"] for d in decisions if d["decision"] == "approve"}
            if approved_set.issuperset(set(apv.get("approvers", []))):
                new_status = "approved"
        elif mode == "any":
            new_status = "approved"
        else:  # sequential
            new_index = apv.get("current_index", 0) + 1
            if new_index >= len(apv.get("approvers", [])):
                new_status = "approved"
    await db.approvals.update_one(
        {"approval_id": apv["approval_id"]},
        {"$set": {"decisions": decisions, "status": new_status,
                  "current_index": new_index, "updated_at": _now()}},
    )
    if new_status == "pending" and mode == "sequential":
        # send next approver an invite
        next_approver = apv["approvers"][new_index]
        token = _gen("aptk")
        await db.approval_tokens.insert_one({
            "token": token, "approval_id": apv["approval_id"],
            "approver": next_approver, "created_at": _now(),
        })
        base_url = os.environ.get("PUBLIC_BASE_URL") or ""
        review_url = f"{base_url}/approve/{token}" if base_url else f"/approve/{token}"
        approve_url = f"{base_url}/api/public/approvals/{token}/approve" if base_url else f"/api/public/approvals/{token}/approve"
        reject_url = f"{base_url}/api/public/approvals/{token}/reject" if base_url else f"/api/public/approvals/{token}/reject"
        html = _approval_email_html(
            apv.get("subject", "Approval needed"),
            apv.get("description", ""),
            review_url, approve_url, reject_url,
        )
        await _send_email(
            db,
            EmailRequest(to=[next_approver], subject=f"[Approval] {apv.get('subject','')}", body_html=html),
            execution_id=apv.get("execution_id"),
        )


async def _maybe_resume(db, engine: "WorkflowEngine", approval_id: str) -> None:
    apv = _clean(await db.approvals.find_one({"approval_id": approval_id}))
    if not apv or apv.get("status") == "pending":
        return
    await engine.resume_from_approval(apv["execution_id"], apv)


def _extract_triggers(nodes: List[Any]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for n in nodes:
        if isinstance(n, dict) and n.get("kind") == "trigger":
            cfg = n.get("config", {}) or {}
            event = cfg.get("event") or n.get("type", "").replace("trigger.", "")
            out.append({"event": event, "filter": cfg.get("filter") or {}, "node_id": n.get("id")})
    return out


# ---------------------------------------------------------------------------
# Public trigger hook for the rest of the backend
# ---------------------------------------------------------------------------

async def fire_trigger(db, event: str, payload: Dict[str, Any], user_id: Optional[str] = None) -> List[str]:
    """Fire `event` against every published workflow that listens for it.

    Safe to call from request handlers: failures inside workflow execution are
    swallowed and logged, so they never break the original request.
    """
    try:
        engine = WorkflowEngine(db)
        return await engine.fire(event, payload, user_id=user_id)
    except Exception:  # noqa: BLE001
        log.exception("fire_trigger failed for event=%s", event)
        return []


# ---------------------------------------------------------------------------
# Seed built-in workflow templates
# ---------------------------------------------------------------------------

WORKFLOW_TEMPLATES = [
    {
        "slug": "leave-approval",
        "name": "Leave Approval",
        "description": "Routes a leave request to the employee's manager. Auto-approves <=2 days, escalates to HR for >5 days.",
        "nodes": [
            {"id": "n1", "kind": "trigger", "type": "trigger.form_submitted", "label": "When leave form submitted",
             "config": {"event": "form_submitted"}, "position": {"x": 60, "y": 80}},
            {"id": "n2", "kind": "condition", "type": "condition.if", "label": "More than 5 days?",
             "config": {"group": {"combinator": "and", "rules": [{"left": "values.days", "op": ">", "right": 5}]}},
             "position": {"x": 360, "y": 80}},
            {"id": "n3", "kind": "approval", "type": "approval.sequential", "label": "HR + Manager approval",
             "config": {"subject": "Leave > 5 days requires HR + manager approval",
                        "approvers": ["hr@example.com", "manager@example.com"], "due_days": 3},
             "position": {"x": 660, "y": 20}},
            {"id": "n4", "kind": "approval", "type": "approval.sequential", "label": "Manager approval",
             "config": {"subject": "Leave request awaiting your approval",
                        "approvers": ["manager@example.com"], "due_days": 2},
             "position": {"x": 660, "y": 180}},
            {"id": "n5", "kind": "action", "type": "action.set_status", "label": "Mark approved",
             "config": {"status": "approved"}, "position": {"x": 960, "y": 100}},
            {"id": "n6", "kind": "action", "type": "action.send_email", "label": "Notify employee",
             "config": {"to": "{{values.email}}", "subject": "Your leave request was {{approval.status}}",
                        "body": "<p>Hi {{values.name}},</p><p>Your leave request has been <b>{{approval.status}}</b>.</p>"},
             "position": {"x": 1260, "y": 100}},
        ],
        "edges": [
            {"id": "e1", "source": "n1", "target": "n2"},
            {"id": "e2", "source": "n2", "target": "n3", "sourceHandle": "true", "label": "yes"},
            {"id": "e3", "source": "n2", "target": "n4", "sourceHandle": "false", "label": "no"},
            {"id": "e4", "source": "n3", "target": "n5", "sourceHandle": "approved"},
            {"id": "e5", "source": "n4", "target": "n5", "sourceHandle": "approved"},
            {"id": "e6", "source": "n5", "target": "n6"},
        ],
    },
    {
        "slug": "expense-approval",
        "name": "Expense Approval",
        "description": "Approval routing based on amount: <$1k auto-approve, $1k–10k manager, >$10k CFO.",
        "nodes": [
            {"id": "n1", "kind": "trigger", "type": "trigger.form_submitted", "label": "Expense submitted",
             "config": {"event": "form_submitted"}, "position": {"x": 60, "y": 100}},
            {"id": "n2", "kind": "condition", "type": "condition.if", "label": "Amount > 10000?",
             "config": {"group": {"combinator": "and", "rules": [{"left": "values.amount", "op": ">", "right": 10000}]}},
             "position": {"x": 320, "y": 100}},
            {"id": "n3", "kind": "approval", "type": "approval.sequential", "label": "CFO approval",
             "config": {"subject": "Large expense requires CFO approval", "approvers": ["cfo@example.com"]},
             "position": {"x": 600, "y": 30}},
            {"id": "n4", "kind": "approval", "type": "approval.sequential", "label": "Manager approval",
             "config": {"subject": "Expense awaiting approval", "approvers": ["manager@example.com"]},
             "position": {"x": 600, "y": 200}},
            {"id": "n5", "kind": "action", "type": "action.set_status", "config": {"status": "approved"}, "label": "Approve",
             "position": {"x": 900, "y": 120}},
        ],
        "edges": [
            {"id": "e1", "source": "n1", "target": "n2"},
            {"id": "e2", "source": "n2", "target": "n3", "sourceHandle": "true"},
            {"id": "e3", "source": "n2", "target": "n4", "sourceHandle": "false"},
            {"id": "e4", "source": "n3", "target": "n5", "sourceHandle": "approved"},
            {"id": "e5", "source": "n4", "target": "n5", "sourceHandle": "approved"},
        ],
    },
    {
        "slug": "purchase-request",
        "name": "Purchase Request",
        "description": "Sends purchase requests to procurement and emails finance once approved.",
        "nodes": [
            {"id": "n1", "kind": "trigger", "type": "trigger.form_submitted",
             "config": {"event": "form_submitted"}, "position": {"x": 60, "y": 120}, "label": "Submitted"},
            {"id": "n2", "kind": "approval", "type": "approval.sequential",
             "config": {"subject": "PO awaiting approval", "approvers": ["procurement@example.com"]},
             "position": {"x": 360, "y": 120}, "label": "Procurement approval"},
            {"id": "n3", "kind": "action", "type": "action.send_email",
             "config": {"to": "finance@example.com",
                        "subject": "PO {{submission_id}} approved",
                        "body": "<p>PO {{submission_id}} approved by {{approval.approver}}.</p>"},
             "position": {"x": 660, "y": 120}, "label": "Notify finance"},
        ],
        "edges": [
            {"id": "e1", "source": "n1", "target": "n2"},
            {"id": "e2", "source": "n2", "target": "n3", "sourceHandle": "approved"},
        ],
    },
    {
        "slug": "medical-certificate",
        "name": "Medical Certificate",
        "description": "Sends an acknowledgement email when a medical certificate is uploaded.",
        "nodes": [
            {"id": "n1", "kind": "trigger", "type": "trigger.pdf_submitted",
             "config": {"event": "pdf_submitted"}, "position": {"x": 60, "y": 100}, "label": "PDF submitted"},
            {"id": "n2", "kind": "action", "type": "action.send_email",
             "config": {"to": "{{values.email}}",
                        "subject": "Medical certificate received",
                        "body": "<p>We've received your medical certificate. Reference: {{submission_id}}</p>"},
             "position": {"x": 360, "y": 100}, "label": "Acknowledge"},
        ],
        "edges": [{"id": "e1", "source": "n1", "target": "n2"}],
    },
]


async def seed_workflow_templates(db) -> None:
    for tpl in WORKFLOW_TEMPLATES:
        existing = await db.workflows.find_one({"is_template": True, "template_slug": tpl["slug"]})
        if existing:
            continue
        now = _now()
        await db.workflows.insert_one({
            "workflow_id": _gen("wftpl"),
            "name": tpl["name"], "description": tpl["description"],
            "status": "published", "version": 1,
            "nodes": tpl["nodes"], "edges": tpl["edges"],
            "triggers": _extract_triggers(tpl["nodes"]),
            "owner_id": "system",
            "permissions": {},
            "is_template": True, "template_slug": tpl["slug"],
            "created_at": now, "updated_at": now,
        })
