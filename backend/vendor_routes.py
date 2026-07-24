"""
FormForge — Vendor Management, Site Master, Master Data
=======================================================

Three tightly-related modules that share a common pattern: a flexible,
schema-light data store ("master_collections") backed by import/export,
inline editing and a lookup engine that the rest of the app (forms, PDF
forms, workflows, emails) can pull values from.

Public surface:
    build_routers(db, get_current_user, hash_password_fn)
        -> (vendors_router, vendor_users_router, sites_router,
            master_data_router, lookup_router)

    apply_vendor_scope(user, query)
        Mutates a Mongo filter dict to include the vendor RLS clause when
        the caller is a vendor user. Used by forms / submissions / pdf_*
        routes to enforce row-level security without touching their code.
"""
from __future__ import annotations

import csv
import io
import logging
import os
import re
import uuid
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, ConfigDict
from openpyxl import Workbook, load_workbook

log = logging.getLogger("vendor")

# ---------------------------------------------------------------------------
# Models
# ---------------------------------------------------------------------------

class Vendor(BaseModel):
    model_config = ConfigDict(extra="allow")
    vendor_id: str
    name: str
    code: str = ""
    contact_person: str = ""
    email: str = ""
    phone: str = ""
    address: str = ""
    logo_url: Optional[str] = None
    status: str = "active"
    notes: str = ""
    created_at: str
    updated_at: str
    created_by: Optional[str] = None


class VendorIn(BaseModel):
    name: str
    code: str = ""
    contact_person: str = ""
    email: str = ""
    phone: str = ""
    address: str = ""
    logo_url: Optional[str] = None
    status: str = "active"
    notes: str = ""


class VendorUserIn(BaseModel):
    email: str
    name: str
    password: Optional[str] = None  # auto-generated if missing
    role: str = "vendor"  # vendor | vendor_admin
    is_active: bool = True


class VendorUserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[str] = None
    is_active: Optional[bool] = None
    password: Optional[str] = None  # reset


class Assignment(BaseModel):
    forms: List[str] = []
    pdf_forms: List[str] = []
    sites: List[str] = []
    workflows: List[str] = []


# Site master rows are intentionally schema-less so admins can add columns
# without touching code. We only define a canonical column list for import,
# export and the "download template" button — extras are accepted verbatim.
SITE_COLUMNS = [
    "site_name", "site_code", "asset_id", "plant_name", "customer_name",
    "state", "district", "location", "latitude", "longitude",
    "ac_capacity", "dc_capacity", "inverter_capacity",
    "vendor_name", "vendor_login_user", "vendor_email",
    "approver_email", "cluster_manager_name",
    "cluster", "region", "site_status",
    "commission_date", "om_start_date", "warranty_end_date",
    "cycles_per_month", "pm_cycles_per_quarter",
    "remarks",
]

SITE_COLUMN_LABELS = {
    "site_name": "Site Name", "site_code": "Site Code", "asset_id": "Asset ID",
    "plant_name": "Plant Name", "customer_name": "Customer Name",
    "state": "State", "district": "District", "location": "Location",
    "latitude": "Latitude", "longitude": "Longitude",
    "ac_capacity": "AC Capacity (MW)", "dc_capacity": "DC Capacity (MW)",
    "inverter_capacity": "Inverter Capacity (MW)",
    "vendor_name": "Vendor Name", "vendor_login_user": "Vendor Login User",
    "vendor_email": "Vendor Email",
    "approver_email": "Approver Email",
    "cluster_manager_name": "Cluster Manager",
    "cluster": "Cluster", "region": "Region",
    "site_status": "Site Status", "commission_date": "Commission Date",
    "om_start_date": "O&M Start Date", "warranty_end_date": "Warranty End Date",
    "cycles_per_month": "Cleaning cycles / Month",
    "pm_cycles_per_quarter": "PM cycles / Quarter",
    "remarks": "Remarks",
}


# Master Data is a generic key-value store for lookup tables (customers,
# regions, departments, etc.). Each row lives in `db.master_data` and is
# identified by `{table, row_id}`.

class MasterRowIn(BaseModel):
    table: str
    data: Dict[str, Any]


class BulkSitesIn(BaseModel):
    rows: List[Dict[str, Any]]
    delete_missing: bool = False  # if True, remove sites not in this payload


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _gen(prefix: str) -> str:
    return f"{prefix}_{uuid.uuid4().hex[:12]}"


def _clean(doc: Optional[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    if doc is None:
        return None
    doc.pop("_id", None)
    return doc


def _slug(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (s or "").lower()).strip("_")


# ---------------------------------------------------------------------------
# Manpower integration helper — a subset of what `manpower_routes.py` does,
# reused here so form/PDF Data-Source dropdowns can pull manpower directly.
# The manpower directory lives in a separate DB (default `cmes_mp_db`) on
# the SAME Mongo instance; env vars decide the exact names.
# ---------------------------------------------------------------------------
def _mp_enabled() -> bool:
    v = (os.environ.get("MANPOWER_ENABLED", "true") or "").strip().strip('"').strip("'")
    return v.lower() not in ("0", "false", "no", "off")


def _manpower_coll(db):
    """Returns a Motor Collection handle for the external manpower table,
    or None if the integration is disabled or the client isn't available."""
    if not _mp_enabled():
        return None
    client = getattr(db, "client", None)
    if client is None:
        return None
    db_name = (os.environ.get("MANPOWER_DB_NAME", "cmes_mp_db") or "").strip().strip('"').strip("'")
    coll_name = (os.environ.get("MANPOWER_COLLECTION", "manpower") or "").strip().strip('"').strip("'")
    return client[db_name][coll_name]


# Columns exposed to the form/PDF designers when the Data Source is Manpower.
# Order matters — this is the order they appear in the dropdown.
MANPOWER_COLUMNS = [
    {"key": "manpower_id",  "label": "Manpower ID", "core": True},
    {"key": "full_name",    "label": "Full Name",   "core": True},
    {"key": "status",       "label": "Status",      "core": False},
    {"key": "company_name", "label": "Company",     "core": False},
    {"key": "work_state",   "label": "Work State",  "core": False},
    {"key": "location",     "label": "Location",    "core": False},
    {"key": "city",         "label": "City",        "core": False},
    {"key": "state",        "label": "Home State",  "core": False},
    {"key": "phone",        "label": "Phone",       "core": False},
    {"key": "blood_group",  "label": "Blood Group", "core": False},
    {"key": "subvendor",    "label": "Subvendor",   "core": False},
    {"key": "reporting_cluster_manager", "label": "Reporting Manager", "core": False},
    {"key": "reporting_manager_email",   "label": "Manager Email",     "core": False},
    {"key": "postal_code",  "label": "Postal Code", "core": False},
    {"key": "reference",    "label": "Reference",   "core": False},
]


def apply_vendor_scope(user, query: Dict[str, Any]) -> Dict[str, Any]:
    """If `user` is a vendor, mutate `query` so it only matches docs owned
    by, or assigned to, that vendor.

    Safe no-op for super_admin / admin.
    """
    if not user or user.role not in ("vendor", "vendor_admin"):
        return query
    vid = getattr(user, "vendor_id", None)
    if not vid:
        # Vendor user with no vendor_id — return an impossible filter so they
        # see nothing instead of crashing.
        query["$and"] = (query.get("$and") or []) + [{"_impossible": True}]
        return query
    query["vendor_id"] = vid
    return query


def list_assigned_ids(user, kind: str) -> Optional[List[str]]:
    """Return the assigned ids of a given kind ('forms','pdf_forms','sites',
    'workflows') for a vendor user, or None if no scoping is needed."""
    if not user or user.role not in ("vendor", "vendor_admin"):
        return None
    ass = getattr(user, "assignments", None) or {}
    return list(ass.get(kind) or [])


def _site_filter_for_user(user, show_all: bool = False) -> Dict[str, Any]:
    """Build a Mongo filter for `sites` based on caller role.

    Delegates to the centralised `permissions.site_filter()` which implements
    the 4-tier RBAC (super_admin, admin, vendor_admin, vendor_user).
    `show_all=True` is honoured only for super_admin (admins always see the
    restricted set determined by cluster_manager / assigned_admin_ids).
    """
    if not user:
        return {"_impossible": True}
    try:
        from permissions import site_filter, is_super_admin
    except Exception:
        site_filter = None  # type: ignore
        is_super_admin = None  # type: ignore
    if is_super_admin and is_super_admin(user) and show_all:
        return {}
    if site_filter:
        return site_filter(user)
    # legacy fallback (used only if permissions module fails to import)
    if user.role in ("super_admin", "admin") or show_all:
        return {}
    if user.role in ("vendor", "vendor_admin", "vendor_user"):
        or_clauses: List[Dict[str, Any]] = []
        if user.email:
            or_clauses.append({"vendor_email": user.email})
        vid = getattr(user, "vendor_id", None)
        if vid:
            or_clauses.append({"vendor_id": vid})
        assigned = list_assigned_ids(user, "sites") or []
        if assigned:
            or_clauses.append({"site_id": {"$in": assigned}})
        if not or_clauses:
            return {"_impossible": True}
        return {"$or": or_clauses}
    return {"_impossible": True}


async def _vendor_scoped_site_lookup(db, user, display: str, chosen: Any) -> Optional[Dict[str, Any]]:
    """Find a single site row matching `display`==chosen, with vendor scope
    applied — so vendor users cannot resolve a site they don't own/assign to.
    """
    flt = _site_filter_for_user(user, show_all=False)
    flt[display] = chosen
    return await db.sites.find_one(flt)


# ---------------------------------------------------------------------------
# Routers
# ---------------------------------------------------------------------------

def build_routers(db, get_current_user, hash_password_fn, get_optional_user=None):
    vendors = APIRouter(prefix="/vendors", tags=["vendors"])
    vusers = APIRouter(prefix="/vendor-users", tags=["vendor-users"])
    sites = APIRouter(prefix="/sites", tags=["sites"])
    master = APIRouter(prefix="/master-data", tags=["master-data"])
    lookup = APIRouter(prefix="/lookup", tags=["lookup"])
    public_lookup = APIRouter(prefix="/public/lookup", tags=["public-lookup"])
    approvals = APIRouter(prefix="/admin-approvals", tags=["admin-approvals"])

    # Fallback dependency: if the host app forgot to inject `get_optional_user`
    # (older deployments), just resolve to `None` so the public endpoints
    # still work anonymously.
    if get_optional_user is None:
        async def _get_optional_user() -> Optional[Any]:  # type: ignore[valid-type]
            return None
        get_optional_user = _get_optional_user

    async def _require_admin(user) -> None:
        if user.role not in ("super_admin", "admin"):
            raise HTTPException(403, "Admin role required")

    async def _require_master_data_editor(user) -> None:
        """Only Super Admin may write to master data (sites, vendors, master tables).

        This matches the spec: Admin can view but not edit Site/Vendor/Master Data.
        """
        if user.role != "super_admin":
            raise HTTPException(403, "Only Super Admin can edit master data records.")

    async def _require_site_editor(user) -> None:
        """Super Admin OR Admin may edit sites/site columns.

        Site-level data is now part of the day-to-day admin workflow (Plant View
        supports inline editing), so admins are trusted to write to `sites`.
        Vendor/master-data tables remain super-admin only.
        """
        if user.role not in ("super_admin", "admin"):
            raise HTTPException(403, "Admin role required")

    async def _require_vendor_user_editor(user, vendor_id: str) -> None:
        """Super Admin OR the Vendor Admin of THAT vendor.

        Vendor Admins can manage their own team users (add/remove/reset password/
        activate/deactivate). Super Admin can manage anyone.
        """
        if user.role == "super_admin":
            return
        if user.role == "vendor_admin" and getattr(user, "vendor_id", None) == vendor_id:
            return
        raise HTTPException(403, "Only Super Admin or this vendor's Vendor Admin may edit team users.")

    # ---------- Vendors CRUD ----------

    def _normalize_vendor(row: Dict[str, Any]) -> Dict[str, Any]:
        """Alias `name` ↔ `vendor_name` and `email` ↔ `vendor_email` so both
        the legacy schema (seeded `SunOps` doc) and the new schema (from
        `VendorIn`) look identical to the front-end.  Prefer whichever value
        is present so nothing gets clobbered."""
        if not row:
            return row
        n = row.get("name") or row.get("vendor_name") or ""
        e = row.get("email") or row.get("vendor_email") or ""
        row["name"] = n
        row["vendor_name"] = n
        row["email"] = e
        row["vendor_email"] = e
        return row

    @vendors.get("")
    async def list_vendors(user=Depends(get_current_user)):
        if user.role in ("vendor", "vendor_admin"):
            # vendor sees only itself
            rows = await db.vendors.find({"vendor_id": user.vendor_id}, {"_id": 0}).to_list(10)
            return [_normalize_vendor(r) for r in rows]
        if user.role not in ("super_admin", "admin"):
            return []  # regular users do not see any vendors
        rows = await db.vendors.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)
        return [_normalize_vendor(r) for r in rows]

    @vendors.post("")
    async def create_vendor(body: VendorIn, user=Depends(get_current_user)):
        await _require_master_data_editor(user)
        now = _now()
        doc = {
            "vendor_id": _gen("ven"),
            **body.model_dump(),
            "created_at": now, "updated_at": now, "created_by": user.user_id,
        }
        await db.vendors.insert_one(dict(doc))
        return _normalize_vendor(doc)

    @vendors.get("/{vid}")
    async def get_vendor(vid: str, user=Depends(get_current_user)):
        if user.role in ("vendor", "vendor_admin") and user.vendor_id != vid:
            raise HTTPException(403, "Not allowed")
        v = _clean(await db.vendors.find_one({"vendor_id": vid}))
        if not v:
            raise HTTPException(404, "Vendor not found")
        _normalize_vendor(v)
        # also fetch counts
        v["_stats"] = {
            "users": await db.users.count_documents({"vendor_id": vid}),
            "sites": await db.sites.count_documents({"vendor_id": vid}),
        }
        return v

    @vendors.put("/{vid}")
    async def update_vendor(vid: str, body: VendorIn, user=Depends(get_current_user)):
        await _require_master_data_editor(user)
        existing = _clean(await db.vendors.find_one({"vendor_id": vid}))
        if not existing:
            raise HTTPException(404, "Vendor not found")
        upd = {**body.model_dump(), "updated_at": _now()}
        await db.vendors.update_one({"vendor_id": vid}, {"$set": upd})
        return _normalize_vendor({**existing, **upd})

    @vendors.delete("/{vid}")
    async def delete_vendor(vid: str, user=Depends(get_current_user)):
        await _require_master_data_editor(user)
        # also deactivate vendor users
        await db.vendors.delete_one({"vendor_id": vid})
        await db.users.update_many({"vendor_id": vid}, {"$set": {"is_active": False}})
        return {"ok": True}

    # ---------- Vendor users ----------

    @vusers.get("/{vid}")
    async def list_vendor_users(vid: str, user=Depends(get_current_user)):
        if user.role in ("vendor", "vendor_admin") and user.vendor_id != vid:
            raise HTTPException(403, "Not allowed")
        elif user.role not in ("super_admin", "admin", "vendor", "vendor_admin"):
            raise HTTPException(403, "Not allowed")
        rows = await db.users.find({"vendor_id": vid}, {"_id": 0, "password_hash": 0}).to_list(500)
        return rows

    @vusers.post("/{vid}")
    async def create_vendor_user(vid: str, body: VendorUserIn, user=Depends(get_current_user)):
        await _require_vendor_user_editor(user, vid)
        vendor = _clean(await db.vendors.find_one({"vendor_id": vid}))
        if not vendor:
            raise HTTPException(404, "Vendor not found")
        # check uniqueness
        if await db.users.find_one({"email": body.email.lower()}):
            raise HTTPException(400, "Email already registered")
        pwd = body.password or _gen("pwd")
        doc = {
            "user_id": _gen("usr"),
            "email": body.email.lower(),
            "name": body.name,
            "role": body.role if body.role in ("vendor", "vendor_admin") else "vendor",
            "is_active": body.is_active,
            "created_at": _now(),
            "password_hash": hash_password_fn(pwd),
            "vendor_id": vid,
            "assignments": {"forms": [], "pdf_forms": [], "sites": [], "workflows": []},
        }
        await db.users.insert_one(dict(doc))
        out = {k: v for k, v in doc.items() if k != "password_hash"}
        out["initial_password"] = pwd if not body.password else None
        return out

    @vusers.patch("/{user_id}")
    @vusers.put("/{user_id}")
    async def update_vendor_user(user_id: str, body: VendorUserUpdate, user=Depends(get_current_user)):
        target = _clean(await db.users.find_one({"user_id": user_id}))
        if not target or not target.get("vendor_id"):
            raise HTTPException(404, "Vendor user not found")
        await _require_vendor_user_editor(user, target["vendor_id"])
        # Vendor-admin deactivations require super_admin / admin approval.
        # We short-circuit here so the deactivation is queued rather than
        # applied immediately.  Everything else is applied normally, and
        # `is_active=True` (re-enable) is left alone since re-enabling is
        # not destructive per the current product decision.
        if (user.role == "vendor_admin" and body.is_active is False
                and target.get("is_active", True) is not False):
            existing = await db.pending_approvals.find_one({
                "type": "user_disable",
                "target_user_id": user_id,
                "status": "pending",
            })
            if existing:
                raise HTTPException(409, "A disable request is already pending for this user")
            approval = {
                "approval_id": _gen("apv"),
                "type": "user_disable",
                "target_user_id": user_id,
                "target_email": target.get("email"),
                "target_name": target.get("name"),
                "target_region": target.get("region"),
                "target_cluster_manager_name": target.get("cluster_manager_name"),
                "vendor_id": target.get("vendor_id"),
                "requested_by": {
                    "user_id": user.user_id,
                    "name": getattr(user, "name", None),
                    "email": getattr(user, "email", None),
                    "role": user.role,
                    "region": getattr(user, "region", None),
                    "cluster_manager_name": getattr(user, "cluster_manager_name", None),
                },
                "status": "pending",
                "created_at": _now(),
            }
            await db.pending_approvals.insert_one(dict(approval))
            approval.pop("_id", None)
            return {"pending_approval": True, "approval": approval,
                    "message": "Disable request submitted for admin approval"}
        upd: Dict[str, Any] = {}
        if body.name is not None:
            upd["name"] = body.name
        if body.role is not None and body.role in ("vendor", "vendor_admin"):
            upd["role"] = body.role
        if body.is_active is not None:
            upd["is_active"] = body.is_active
        if body.password:
            upd["password_hash"] = hash_password_fn(body.password)
        if upd:
            await db.users.update_one({"user_id": user_id}, {"$set": upd})
        target.update(upd)
        target.pop("password_hash", None)
        return target

    @vusers.put("/{user_id}/assignments")
    async def set_assignments(user_id: str, body: Assignment, user=Depends(get_current_user)):
        target = _clean(await db.users.find_one({"user_id": user_id}))
        if not target or not target.get("vendor_id"):
            raise HTTPException(404, "Vendor user not found")
        await _require_vendor_user_editor(user, target["vendor_id"])
        await db.users.update_one(
            {"user_id": user_id},
            {"$set": {"assignments": body.model_dump(),
                      "updated_at": _now()}},
        )
        target["assignments"] = body.model_dump()
        target.pop("password_hash", None)
        return target

    @vusers.delete("/{user_id}")
    async def delete_vendor_user(user_id: str, user=Depends(get_current_user)):
        target = _clean(await db.users.find_one({"user_id": user_id, "vendor_id": {"$exists": True}}))
        if not target:
            raise HTTPException(404, "Vendor user not found")
        await _require_vendor_user_editor(user, target["vendor_id"])
        await db.users.delete_one({"user_id": user_id, "vendor_id": {"$exists": True}})
        return {"ok": True}

    # ---------- Sites (master) ----------

    @sites.get("")
    async def list_sites(user=Depends(get_current_user), q: Optional[str] = None,
                          limit: int = 5000, show_all: bool = False):
        # Vendor users default to seeing only their own / assigned sites.
        # Admins can pass show_all=true if they want — they already see all by default.
        flt = _site_filter_for_user(user, show_all=show_all)
        if flt.get("_impossible"):
            return []
        rows = await db.sites.find(flt, {"_id": 0}).sort("site_name", 1).to_list(min(limit, 50000))
        if q:
            ql = q.lower()
            rows = [r for r in rows if any(ql in str(v).lower() for v in r.values() if v)]
        return rows

    @sites.get("/columns")
    async def list_columns(user=Depends(get_current_user)):
        cfg = _clean(await db.site_columns.find_one({"_id": "default"})) or {}
        custom = cfg.get("custom") or []
        cols = [{"key": k, "label": SITE_COLUMN_LABELS[k], "core": True} for k in SITE_COLUMNS]
        cols.extend([{"key": c["key"], "label": c["label"], "core": False} for c in custom])
        return cols

    @sites.post("/columns")
    async def add_column(body: Dict[str, str], user=Depends(get_current_user)):
        await _require_site_editor(user)
        label = (body.get("label") or "").strip()
        if not label:
            raise HTTPException(400, "Label required")
        key = _slug(label)
        if key in SITE_COLUMNS:
            raise HTTPException(400, "Column already exists")
        cfg = _clean(await db.site_columns.find_one({"_id": "default"})) or {"custom": []}
        if any(c["key"] == key for c in cfg["custom"]):
            raise HTTPException(400, "Custom column already exists")
        cfg["custom"].append({"key": key, "label": label})
        await db.site_columns.update_one({"_id": "default"}, {"$set": {"custom": cfg["custom"]}}, upsert=True)
        return cfg["custom"]

    @sites.delete("/columns/{key}")
    async def del_column(key: str, user=Depends(get_current_user)):
        await _require_site_editor(user)
        if key in SITE_COLUMNS:
            raise HTTPException(400, "Core columns cannot be removed")
        cfg = _clean(await db.site_columns.find_one({"_id": "default"})) or {"custom": []}
        cfg["custom"] = [c for c in cfg["custom"] if c["key"] != key]
        await db.site_columns.update_one({"_id": "default"}, {"$set": {"custom": cfg["custom"]}}, upsert=True)
        return cfg["custom"]

    @sites.post("")
    async def create_site(body: Dict[str, Any], user=Depends(get_current_user)):
        await _require_master_data_editor(user)
        return await _upsert_site(db, body, user)

    @sites.put("/{site_id}")
    async def update_site(site_id: str, body: Dict[str, Any], user=Depends(get_current_user)):
        await _require_site_editor(user)
        existing = _clean(await db.sites.find_one({"site_id": site_id}))
        if not existing:
            raise HTTPException(404, "Site not found")
        # version snapshot
        await db.site_versions.insert_one({
            "snapshot_id": _gen("siv"),
            "site_id": site_id, "version": existing.get("version", 1),
            "row": existing, "saved_at": _now(), "saved_by": user.user_id,
        })
        upd = {k: v for k, v in body.items() if k not in ("site_id", "_id")}
        # Normalise `vendor_email` string with ; , or newline separators →
        # first one becomes vendor_email, all of them become allowed_emails.
        raw_email = upd.get("vendor_email")
        if isinstance(raw_email, str) and raw_email.strip():
            emails = _split_emails(raw_email)
            if emails:
                upd["vendor_email"] = emails[0]
                upd["allowed_emails"] = emails
        elif isinstance(upd.get("allowed_emails"), str):
            upd["allowed_emails"] = _split_emails(upd["allowed_emails"])
        upd["updated_at"] = _now()
        upd["updated_by"] = user.user_id
        upd["version"] = int(existing.get("version", 1)) + 1
        await db.sites.update_one({"site_id": site_id}, {"$set": upd})
        await _audit_master(db, user, "site.update", site_id, {"changes": list(upd.keys())})
        return {**existing, **upd}

    @sites.post("/relink-vendors")
    async def relink_vendors(user=Depends(get_current_user)):
        """Best-effort repair: for every site missing `vendor_id`, try to
        match by `vendor_name` or `vendor_email` against the vendors
        collection and link the site to the matching vendor doc."""
        await _require_site_editor(user)
        vendors_rows = await db.vendors.find(
            {}, {"_id": 0, "vendor_id": 1, "name": 1, "vendor_name": 1,
                 "email": 1, "vendor_email": 1}).to_list(5000)
        by_name = {}
        by_email = {}
        for v in vendors_rows:
            for n in (v.get("name"), v.get("vendor_name")):
                if n:
                    by_name[str(n).strip().lower()] = v["vendor_id"]
            for e in (v.get("email"), v.get("vendor_email")):
                if e:
                    by_email[str(e).strip().lower()] = v["vendor_id"]
        cursor = db.sites.find({"$or": [{"vendor_id": None}, {"vendor_id": {"$exists": False}}]},
                               {"_id": 0, "site_id": 1, "vendor_name": 1, "vendor_email": 1})
        fixed = 0
        async for s in cursor:
            vid = None
            n = (s.get("vendor_name") or "").strip().lower()
            e = (s.get("vendor_email") or "").strip().lower()
            if n and n in by_name:
                vid = by_name[n]
            elif e and e in by_email:
                vid = by_email[e]
            if vid:
                await db.sites.update_one({"site_id": s["site_id"]},
                                          {"$set": {"vendor_id": vid}})
                fixed += 1
        return {"ok": True, "relinked": fixed}

    @sites.delete("/{site_id}")
    async def delete_site(site_id: str, user=Depends(get_current_user)):
        await _require_master_data_editor(user)
        await db.sites.delete_one({"site_id": site_id})
        await _audit_master(db, user, "site.delete", site_id, {})
        return {"ok": True}

    @sites.post("/bulk")
    async def bulk_upsert(body: BulkSitesIn, user=Depends(get_current_user)):
        await _require_master_data_editor(user)
        upserted = 0
        for row in body.rows:
            await _upsert_site(db, row, user)
            upserted += 1
        if body.delete_missing:
            keep_ids = [r.get("site_id") for r in body.rows if r.get("site_id")]
            await db.sites.delete_many({"site_id": {"$nin": keep_ids}})
        await _audit_master(db, user, "site.bulk", None, {"upserted": upserted})
        return {"ok": True, "upserted": upserted}

    @sites.post("/bulk-delete")
    async def bulk_delete(body: Dict[str, List[str]], user=Depends(get_current_user)):
        await _require_master_data_editor(user)
        ids = body.get("site_ids", [])
        n = await db.sites.delete_many({"site_id": {"$in": ids}})
        await _audit_master(db, user, "site.bulk_delete", None, {"count": n.deleted_count})
        return {"deleted": n.deleted_count}

    @sites.get("/template.xlsx")
    async def site_template(user=Depends(get_current_user)):
        await _require_admin(user)
        wb = Workbook()
        ws = wb.active
        ws.title = "Sites"
        cols = await list_columns(user=user)
        ws.append([c["label"] for c in cols])
        # add one example row
        example = {
            "site_name": "Solar Farm Alpha", "site_code": "SF-001", "asset_id": "AST-1001",
            "plant_name": "Alpha Plant", "customer_name": "Acme Power",
            "state": "Karnataka", "district": "Tumkur", "location": "Pavagada",
            "latitude": 14.099, "longitude": 77.275,
            "ac_capacity": 50, "dc_capacity": 65, "inverter_capacity": 50,
            "vendor_name": "SunOps Pvt Ltd", "vendor_login_user": "sunops@example.com",
            "vendor_email": "ops@sunops.example.com",
            "cluster": "South-1", "region": "South",
            "site_status": "operational",
            "commission_date": "2023-04-01", "om_start_date": "2023-05-01",
            "warranty_end_date": "2028-04-01", "remarks": "Reference row — replace with your data",
        }
        ws.append([example.get(c["key"], "") for c in cols])
        # header styling
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="sites-template.xlsx"'},
        )

    @sites.get("/export.xlsx")
    async def export_xlsx(user=Depends(get_current_user)):
        rows = await list_sites(user=user, limit=50000)
        cols = await list_columns(user=user)
        wb = Workbook()
        ws = wb.active
        ws.title = "Sites"
        ws.append([c["label"] for c in cols])
        for r in rows:
            ws.append([r.get(c["key"], "") for c in cols])
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": 'attachment; filename="sites-export.xlsx"'},
        )

    @sites.get("/export.csv")
    async def export_csv(user=Depends(get_current_user)):
        rows = await list_sites(user=user, limit=50000)
        cols = await list_columns(user=user)
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow([c["label"] for c in cols])
        for r in rows:
            w.writerow([r.get(c["key"], "") for c in cols])
        return StreamingResponse(
            io.BytesIO(buf.getvalue().encode("utf-8")),
            media_type="text/csv",
            headers={"Content-Disposition": 'attachment; filename="sites-export.csv"'},
        )

    @sites.post("/import")
    async def import_xlsx_or_csv(
        file: UploadFile = File(...),
        replace: bool = Form(False),
        user=Depends(get_current_user),
    ):
        await _require_master_data_editor(user)
        cols = await list_columns(user=user)
        key_by_label = {c["label"].lower(): c["key"] for c in cols}
        # also accept the raw keys as headers
        for c in cols:
            key_by_label[c["key"].lower()] = c["key"]
        content = await file.read()
        rows: List[Dict[str, Any]] = []
        name = (file.filename or "").lower()
        if name.endswith(".csv"):
            text = content.decode("utf-8-sig", errors="ignore")
            reader = csv.reader(io.StringIO(text))
            header = next(reader, []) or []
            keys = [key_by_label.get((h or "").strip().lower()) for h in header]
            for r in reader:
                rows.append({keys[i]: r[i] for i in range(len(keys)) if keys[i] and i < len(r)})
        else:
            try:
                wb = load_workbook(io.BytesIO(content), data_only=True)
            except Exception as e:
                raise HTTPException(400, f"Unable to read file: {e}")
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            header = next(it, []) or []
            keys = [key_by_label.get((str(h or "")).strip().lower()) for h in header]
            for r in it:
                row = {}
                for i, val in enumerate(r):
                    if i < len(keys) and keys[i]:
                        row[keys[i]] = val
                if any(v not in (None, "") for v in row.values()):
                    rows.append(row)
        # Persist
        if replace:
            await db.sites.delete_many({})
        upserted = 0
        for row in rows:
            await _upsert_site(db, row, user)
            upserted += 1
        # import history
        await db.site_imports.insert_one({
            "import_id": _gen("imp"), "filename": file.filename,
            "rows": upserted, "replace": replace,
            "by": user.user_id, "at": _now(),
        })
        await _audit_master(db, user, "site.import", None, {"rows": upserted, "replace": replace})
        return {"ok": True, "rows": upserted}

    @sites.get("/{site_id}/history")
    async def site_history(site_id: str, user=Depends(get_current_user)):
        await _require_admin(user)
        rows = await db.site_versions.find({"site_id": site_id}, {"_id": 0}).sort("saved_at", -1).to_list(50)
        return rows

    @sites.get("/by-code/{site_code}/history")
    async def site_history_by_code(site_code: str, user=Depends(get_current_user)):
        """Plant View — edit history timeline for the plant with this code.
        RLS-aware: any user who can view the plant can see its history.
        Returns snapshots newest→oldest with saved_by user resolved to name/email
        and a diff (list of changed fields) versus the immediately-newer snapshot
        (or the *current* live row for the newest snapshot)."""
        from permissions import site_filter as _sf, is_super_admin
        q: Dict[str, Any] = {"site_code": site_code}
        if not is_super_admin(user):
            rls = _sf(user)
            q = {"$and": [q, rls]} if rls else q
        site = await db.sites.find_one(q, {"_id": 0})
        if not site:
            raise HTTPException(404, "Plant not found or out of scope")

        versions = await db.site_versions.find(
            {"site_id": site.get("site_id")}, {"_id": 0},
        ).sort("saved_at", -1).to_list(50)

        # Resolve saved_by → user email/name
        user_ids = list({v.get("saved_by") for v in versions if v.get("saved_by")})
        users_by_id: Dict[str, Dict[str, Any]] = {}
        if user_ids:
            async for u in db.users.find(
                {"user_id": {"$in": user_ids}}, {"_id": 0, "user_id": 1, "email": 1, "name": 1},
            ):
                users_by_id[u["user_id"]] = u

        _skip_diff_keys = {
            "_id", "updated_at", "updated_by", "version", "created_at",
            "is_deleted", "assigned_admin_ids", "assigned_vendor_ids",
        }

        def _diff(old_row: Dict[str, Any], new_row: Dict[str, Any]) -> List[Dict[str, Any]]:
            keys = set(old_row.keys()) | set(new_row.keys())
            out = []
            for k in keys:
                if k in _skip_diff_keys:
                    continue
                ov, nv = old_row.get(k), new_row.get(k)
                if ov != nv:
                    out.append({"field": k, "from": ov, "to": nv})
            return out

        # Each snapshot stores the row *before* an edit. So version[i]["row"]
        # was replaced by version[i-1]["row"] (or the current live row if i==0).
        enriched: List[Dict[str, Any]] = []
        for idx, v in enumerate(versions):
            snap_row = v.get("row") or {}
            next_row = versions[idx - 1]["row"] if idx > 0 else site
            changed = _diff(snap_row, next_row)
            saved_by = v.get("saved_by")
            u = users_by_id.get(saved_by) if saved_by else None
            enriched.append({
                "snapshot_id": v.get("snapshot_id"),
                "version": v.get("version"),
                "saved_at": v.get("saved_at"),
                "saved_by": saved_by,
                "saved_by_email": (u or {}).get("email"),
                "saved_by_name": (u or {}).get("name"),
                "changes": changed,
                "change_count": len(changed),
            })
        return {
            "site_id": site.get("site_id"),
            "site_code": site.get("site_code"),
            "current_version": site.get("version", 1),
            "history": enriched,
        }

    @sites.get("/by-code/{site_code}")
    async def get_site_by_code(site_code: str, user=Depends(get_current_user)):
        """Plant View endpoint — fetch one site (with RLS) and enrich with
        recent submissions summary for the Plant page.  Non-privileged
        users hit a 404 when the site is outside their access scope."""
        from permissions import site_filter as _sf, is_super_admin
        q: Dict[str, Any] = {"site_code": site_code}
        if not is_super_admin(user):
            rls = _sf(user)
            q = {"$and": [q, rls]} if rls else q
        site = await db.sites.find_one(q, {"_id": 0})
        if not site:
            raise HTTPException(404, "Plant not found or out of scope")

        # Recent submissions that mention this plant.  We can NOT rely on the
        # submission having a specific key like `values.site_name` because
        # form-designers use arbitrary field IDs (e.g. `dropdown-1738`,
        # `plant`, `Plant Name` …).  Instead scan every value in each
        # submission's `values` object for a match against any of the
        # plant's canonical identifiers.
        match_values = [
            site.get("site_name"), site.get("site_code"),
            site.get("asset_id"), site.get("site_id"),
        ]
        match_values = [m for m in match_values if m]
        if not match_values:
            return {"site": site, "recent_submissions": []}

        value_scan = {
            "$expr": {
                "$anyElementTrue": {
                    "$map": {
                        "input": {"$objectToArray": {"$ifNull": ["$values", {}]}},
                        "as": "kv",
                        "in": {"$in": ["$$kv.v", match_values]},
                    }
                }
            }
        }

        recent: List[Dict[str, Any]] = []
        for col in ("submissions", "pdf_submissions"):
            rows = await db[col].find(
                value_scan,
                {"_id": 0, "submission_id": 1, "form_id": 1, "template_id": 1,
                 "created_at": 1, "status": 1, "submitted_by": 1, "values": 1},
            ).sort("created_at", -1).limit(50).to_list(50)
            for r in rows:
                r["kind"] = "pdf" if col == "pdf_submissions" else "form"
                # drop the heavy `values` blob from the wire payload — only
                # kept above so the $expr can run server-side.
                r.pop("values", None)
                recent.append(r)
        recent.sort(key=lambda r: r.get("created_at", ""), reverse=True)
        return {"site": site, "recent_submissions": recent[:20]}

    @sites.put("/by-code/{site_code}")
    async def update_site_by_code(site_code: str, body: Dict[str, Any],
                                  user=Depends(get_current_user)):
        """Plant View editor — patch a site by its `site_code`.

        Any admin can update sites they can see (Plant View is now editable).
        Fields not in the current column schema are silently accepted so the
        UI can add ad-hoc data alongside adding a new column.
        """
        await _require_site_editor(user)
        existing = _clean(await db.sites.find_one({"site_code": site_code}))
        if not existing:
            raise HTTPException(404, "Site not found")
        # version snapshot for audit
        await db.site_versions.insert_one({
            "snapshot_id": _gen("siv"),
            "site_id": existing.get("site_id"),
            "version": existing.get("version", 1),
            "row": existing, "saved_at": _now(), "saved_by": user.user_id,
        })
        upd = {k: v for k, v in body.items() if k not in ("site_id", "_id", "site_code")}
        upd["updated_at"] = _now()
        upd["updated_by"] = user.user_id
        upd["version"] = int(existing.get("version", 1)) + 1
        await db.sites.update_one({"site_code": site_code}, {"$set": upd})
        await _audit_master(db, user, "site.update", existing.get("site_id"),
                            {"via": "by-code", "changes": list(upd.keys())})
        return {**existing, **upd}

    @sites.get("/_imports")
    async def import_history(user=Depends(get_current_user)):
        await _require_admin(user)
        rows = await db.site_imports.find({}, {"_id": 0}).sort("at", -1).to_list(100)
        return rows

    # ---------- Generic Master Data ----------

    @master.get("/tables")
    async def list_tables(user=Depends(get_current_user)):
        rows = await db.master_data.aggregate([
            {"$group": {"_id": "$table", "count": {"$sum": 1}}},
            {"$project": {"_id": 0, "table": "$_id", "count": 1}},
        ]).to_list(200)
        # include the always-present built-ins
        builtins = ["customers", "regions", "states", "departments", "products", "categories"]
        present = {r["table"] for r in rows}
        for b in builtins:
            if b not in present:
                rows.append({"table": b, "count": 0})
        return sorted(rows, key=lambda r: r["table"])

    @master.get("/{table}")
    async def list_master(table: str, q: Optional[str] = None, user=Depends(get_current_user)):
        rows = await db.master_data.find({"table": table}, {"_id": 0}).sort("created_at", -1).to_list(20000)
        if q:
            ql = q.lower()
            rows = [r for r in rows if any(ql in str(v).lower() for v in (r.get("data") or {}).values() if v)]
        return rows

    @master.post("/{table}")
    async def add_master(table: str, body: Dict[str, Any], user=Depends(get_current_user)):
        await _require_master_data_editor(user)
        doc = {
            "row_id": _gen("mdr"), "table": table,
            "data": body, "created_at": _now(), "created_by": user.user_id, "version": 1,
        }
        await db.master_data.insert_one(dict(doc))
        await _audit_master(db, user, "master.add", doc["row_id"], {"table": table})
        return doc

    @master.put("/{row_id}")
    async def update_master(row_id: str, body: Dict[str, Any], user=Depends(get_current_user)):
        await _require_master_data_editor(user)
        existing = _clean(await db.master_data.find_one({"row_id": row_id}))
        if not existing:
            raise HTTPException(404, "Row not found")
        await db.master_data_versions.insert_one({
            "snapshot_id": _gen("mdv"), "row_id": row_id, "version": existing.get("version", 1),
            "data": existing.get("data"), "saved_at": _now(), "saved_by": user.user_id,
        })
        upd = {
            "data": body, "updated_at": _now(),
            "updated_by": user.user_id,
            "version": int(existing.get("version", 1)) + 1,
        }
        await db.master_data.update_one({"row_id": row_id}, {"$set": upd})
        await _audit_master(db, user, "master.update", row_id, {})
        return {**existing, **upd}

    @master.delete("/{row_id}")
    async def delete_master(row_id: str, user=Depends(get_current_user)):
        await _require_master_data_editor(user)
        await db.master_data.delete_one({"row_id": row_id})
        await _audit_master(db, user, "master.delete", row_id, {})
        return {"ok": True}

    # ---------- Excel import / export ----------

    @master.get("/{table}/export.xlsx")
    async def export_master(table: str, user=Depends(get_current_user)):
        """Download all rows of a master-data table as an Excel workbook.
        Column order follows the union of keys across all rows (stable)."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from fastapi.responses import StreamingResponse
        from io import BytesIO
        rows = await db.master_data.find(
            {"table": table}, {"_id": 0}).sort("created_at", 1).to_list(20000)
        # Union of keys in a stable order — first-seen order wins.
        cols: List[str] = []
        seen = set()
        for r in rows:
            for k in (r.get("data") or {}).keys():
                if k not in seen:
                    seen.add(k)
                    cols.append(k)
        wb = Workbook()
        ws = wb.active
        ws.title = table[:30] or "master"
        header_fill = PatternFill("solid", fgColor="1F2937")
        header_font = Font(color="FFFFFF", bold=True)
        for i, c in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=i, value=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", horizontal="left")
            ws.column_dimensions[chr(ord("A") + i - 1)].width = max(14, len(c) + 2)
        ws.row_dimensions[1].height = 22
        for ri, r in enumerate(rows, start=2):
            d = r.get("data") or {}
            for ci, c in enumerate(cols, start=1):
                v = d.get(c)
                # Convert lists/dicts to a JSON-ish string so nothing breaks in Excel
                if isinstance(v, (list, dict)):
                    import json as _json
                    v = _json.dumps(v, ensure_ascii=False)
                ws.cell(row=ri, column=ci, value=v)
        ws.freeze_panes = "A2"
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{table}.xlsx"'},
        )

    @master.get("/{table}/template.xlsx")
    async def template_master(table: str, user=Depends(get_current_user)):
        """Blank template — headers only, taken from the union of existing
        columns.  When the table has no rows yet, returns a minimal
        placeholder ("column_1") so the user can start from scratch."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from fastapi.responses import StreamingResponse
        from io import BytesIO
        rows = await db.master_data.find({"table": table}, {"_id": 0}).limit(200).to_list(200)
        cols: List[str] = []
        seen = set()
        for r in rows:
            for k in (r.get("data") or {}).keys():
                if k not in seen:
                    seen.add(k)
                    cols.append(k)
        if not cols:
            cols = ["column_1", "column_2"]
        wb = Workbook(); ws = wb.active; ws.title = "template"
        for i, c in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=i, value=c)
            cell.fill = PatternFill("solid", fgColor="1F2937")
            cell.font = Font(color="FFFFFF", bold=True)
            ws.column_dimensions[chr(ord("A") + i - 1)].width = max(14, len(c) + 2)
        buf = BytesIO(); wb.save(buf); buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{table}-template.xlsx"'},
        )

    @master.post("/{table}/import")
    async def import_master(table: str, file: UploadFile = File(...),
                            mode: str = "append",
                            user=Depends(get_current_user)):
        """Bulk-import rows from an uploaded xlsx / csv.

        First row = header (column names).  Subsequent rows become
        `master_data` documents under `table`.  `mode`:
          * append (default) — insert every row as new
          * replace          — delete all existing rows of this table first
        """
        await _require_master_data_editor(user)
        if mode not in ("append", "replace"):
            raise HTTPException(400, "mode must be 'append' or 'replace'")
        raw = await file.read()
        if not raw:
            raise HTTPException(400, "Empty file")
        fname = (file.filename or "").lower()
        headers: List[str] = []
        rows: List[Dict[str, Any]] = []
        if fname.endswith(".csv"):
            import csv, io
            text = raw.decode("utf-8-sig", errors="replace")
            reader = csv.reader(io.StringIO(text))
            for i, row in enumerate(reader):
                if i == 0:
                    headers = [str(c).strip() for c in row]
                else:
                    if not any(str(c).strip() for c in row):
                        continue
                    d = {headers[j]: row[j] for j in range(min(len(row), len(headers)))
                         if headers[j] and str(row[j]).strip() != ""}
                    if d: rows.append(d)
        elif fname.endswith(".xlsx") or fname.endswith(".xlsm"):
            from openpyxl import load_workbook
            from io import BytesIO
            wb = load_workbook(BytesIO(raw), data_only=True, read_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            try:
                header_row = next(it)
            except StopIteration:
                raise HTTPException(400, "Workbook is empty")
            headers = [str(c).strip() if c is not None else "" for c in header_row]
            for r in it:
                if r is None or not any((c not in (None, "") for c in r)):
                    continue
                d = {}
                for j, val in enumerate(r):
                    if j >= len(headers) or not headers[j]:
                        continue
                    if val in (None, ""):
                        continue
                    # Excel gives datetime objects for date cells — stringify
                    if hasattr(val, "isoformat"):
                        val = val.isoformat()[:10]
                    d[headers[j]] = val
                if d: rows.append(d)
        else:
            raise HTTPException(400, "Unsupported file type — upload .xlsx or .csv")

        if not rows:
            return {"inserted": 0, "replaced": 0, "columns": headers}

        replaced_count = 0
        if mode == "replace":
            r = await db.master_data.delete_many({"table": table})
            replaced_count = r.deleted_count

        docs = []
        now = _now()
        for r in rows:
            docs.append({
                "row_id": _gen("mdr"),
                "table": table,
                "data": r,
                "created_at": now,
                "created_by": user.user_id,
                "version": 1,
            })
        if docs:
            await db.master_data.insert_many(docs)
        await _audit_master(db, user, "master.import", table,
                            {"inserted": len(docs), "replaced": replaced_count, "mode": mode})
        return {"inserted": len(docs), "replaced": replaced_count, "columns": headers}

    # ---------- Lookup engine ----------

    @lookup.post("/resolve")
    async def lookup_resolve(body: Dict[str, Any], user=Depends(get_current_user)):
        """Resolve a lookup configured by the form/PDF designer.

        Body:
            source: "sites" | "master:<table>" | "vendors" | "logged_in_user" | "logged_in_vendor"
            display: column to display in dropdown (e.g. "site_name")
            return: column whose value should be selected (defaults to display)
            value: the chosen `display` value
            fill: list of columns to also return from the matched row (auto-fill)

        Returns: {value, fill: {col: value}}
        """
        source = body.get("source", "sites")
        display = body.get("display") or "site_name"
        return_col = body.get("return") or display
        chosen = body.get("value")
        fill_cols: List[str] = list(body.get("fill") or [])
        if source == "logged_in_user":
            data = {"name": user.name, "email": user.email, "role": user.role,
                    "user_id": user.user_id, "vendor_id": getattr(user, "vendor_id", None)}
            return {"value": data.get(return_col, chosen), "fill": {c: data.get(c) for c in fill_cols}, "matched": True, "row": data}
        if source == "logged_in_vendor":
            vid = getattr(user, "vendor_id", None)
            v = _clean(await db.vendors.find_one({"vendor_id": vid})) if vid else None
            if not v:
                return {"value": None, "fill": {}, "matched": False}
            return {"value": v.get(return_col, chosen), "fill": {c: v.get(c) for c in fill_cols}, "matched": True, "row": v}
        if source == "vendors":
            row = _clean(await db.vendors.find_one({display: chosen}))
            if not row:
                return {"value": None, "fill": {}, "matched": False}
            return {"value": row.get(return_col, chosen), "fill": {c: row.get(c) for c in fill_cols}, "matched": True, "row": row}
        if source == "sites":
            row = _clean(await _vendor_scoped_site_lookup(db, user, display, chosen))
            if not row:
                return {"value": None, "fill": {}, "matched": False}
            return {
                "value": row.get(return_col, chosen),
                "fill": {c: row.get(c) for c in fill_cols},
                "matched": True,
                "row": {k: v for k, v in row.items() if k != "password_hash"},
            }
        if source.startswith("master:"):
            table = source.split(":", 1)[1]
            row = _clean(await db.master_data.find_one({
                "table": table, f"data.{display}": chosen,
            }))
            if not row:
                return {"value": None, "fill": {}, "matched": False}
            data = row.get("data", {})
            return {
                "value": data.get(return_col, chosen),
                "fill": {c: data.get(c) for c in fill_cols},
                "matched": True,
                "row": data,
            }
        if source == "manpower":
            # Manpower Portal integration — resolves against the external
            # cmes_mp_db.manpower collection (see manpower_routes.py).
            coll = _manpower_coll(db)
            if coll is None:
                return {"value": None, "fill": {}, "matched": False}
            row = _clean(await coll.find_one({display: chosen}, {"_id": 0}))
            if not row:
                return {"value": None, "fill": {}, "matched": False}
            return {
                "value": row.get(return_col, chosen),
                "fill": {c: row.get(c) for c in fill_cols},
                "matched": True,
                "row": row,
            }
        raise HTTPException(400, f"Unknown lookup source '{source}'")

    @lookup.get("/options")
    async def lookup_options(source: str = "sites", column: str = "site_name",
                              q: Optional[str] = None, limit: int = 500,
                              show_all: bool = False,
                              user=Depends(get_current_user)):
        """Returns distinct values of `column` from the source — used to
        populate dropdown options.

        For Site Management dropdowns, results are automatically scoped to the
        sites assigned to the calling vendor user (by `vendor_email` ==
        user.email OR `vendor_id` == user.vendor_id OR site_id in
        assignments.sites). Pass `show_all=true` (admins only) to bypass this.
        """
        out: List[str] = []
        if source == "logged_in_user":
            data = {"name": user.name, "email": user.email, "role": user.role}
            return [str(data.get(column, ""))] if data.get(column) else []
        if source == "logged_in_vendor":
            vid = getattr(user, "vendor_id", None)
            if not vid:
                return []
            v = await db.vendors.find_one({"vendor_id": vid}, {"_id": 0})
            return [str(v.get(column, ""))] if v and v.get(column) else []
        if source == "vendors":
            flt: Dict[str, Any] = {}
            if user.role in ("vendor", "vendor_admin"):
                flt["vendor_id"] = user.vendor_id  # vendor user sees only own vendor
            if q:
                flt[column] = {"$regex": re.escape(q), "$options": "i"}
            vals = await db.vendors.distinct(column, flt)
            out = [v for v in vals if v not in (None, "")]
        elif source == "sites":
            flt = _site_filter_for_user(user, show_all)
            if q:
                flt[column] = {"$regex": re.escape(q), "$options": "i"}
            vals = await db.sites.distinct(column, flt)
            out = [v for v in vals if v not in (None, "")]
        elif source.startswith("master:"):
            table = source.split(":", 1)[1]
            flt = {"table": table}
            if q:
                flt[f"data.{column}"] = {"$regex": re.escape(q), "$options": "i"}
            vals = await db.master_data.distinct(f"data.{column}", flt)
            out = [v for v in vals if v not in (None, "")]
        elif source == "manpower":
            coll = _manpower_coll(db)
            if coll is None:
                return []
            flt = {}
            if q:
                flt[column] = {"$regex": re.escape(q), "$options": "i"}
            vals = await coll.distinct(column, flt)
            out = [v for v in vals if v not in (None, "")]
        else:
            raise HTTPException(400, "Unknown source")
        return sorted([str(v) for v in out])[:limit]

    @lookup.get("/columns")
    async def lookup_columns(source: str = "sites", user=Depends(get_current_user)):
        """Returns the available columns for a given lookup source."""
        if source == "sites":
            cols = await list_columns(user=user)
            return cols
        if source == "manpower":
            return MANPOWER_COLUMNS
        if source.startswith("master:"):
            table = source.split(":", 1)[1]
            # discover columns by union of `data` keys across the table
            rows = await db.master_data.find({"table": table}, {"_id": 0, "data": 1}).to_list(500)
            keys: List[str] = []
            for r in rows:
                for k in (r.get("data") or {}).keys():
                    if k not in keys:
                        keys.append(k)
            return [{"key": k, "label": k.replace("_", " ").title(), "core": False} for k in keys]
        raise HTTPException(400, "Unknown source")

    # ---------- Public lookup (anonymous form fillers) ----------
    # These endpoints are deliberately unauthenticated because public web forms
    # need to fetch dropdown options and resolve auto-fills without a token.
    # They expose only Site Master + Master Data (which is reference data) and
    # never return password_hash / private fields.

    @public_lookup.get("/options")
    async def public_options(source: str = "sites", column: str = "site_name",
                              q: Optional[str] = None, limit: int = 500,
                              show_all: bool = False,
                              user=Depends(get_optional_user)):
        # Public endpoint — anonymous access returns the full set. If the
        # caller sent a valid Authorization token (e.g. a vendor logged in
        # while filling a shared /p/:slug link), scope by their role so
        # vendors only see their own plants and admins their region/cluster.
        if source == "sites":
            if user:
                flt = _site_filter_for_user(user, show_all=False)
            else:
                flt = {}
            if q:
                flt[column] = {"$regex": re.escape(q), "$options": "i"}
            vals = await db.sites.distinct(column, flt)
        elif source.startswith("master:"):
            table = source.split(":", 1)[1]
            flt = {"table": table}
            if q:
                flt[f"data.{column}"] = {"$regex": re.escape(q), "$options": "i"}
            vals = await db.master_data.distinct(f"data.{column}", flt)
        else:
            raise HTTPException(400, "Unknown source")
        return sorted([str(v) for v in vals if v not in (None, "")])[:limit]

    @public_lookup.post("/resolve")
    async def public_resolve(body: Dict[str, Any], user=Depends(get_optional_user)):
        source = body.get("source", "sites")
        display = body.get("display") or "site_name"
        return_col = body.get("return") or display
        chosen = body.get("value")
        fill_cols: List[str] = list(body.get("fill") or [])
        if source == "sites":
            # When the caller is a logged-in vendor, force the lookup to
            # match only sites they can access — prevents cross-vendor leak
            # when a shared /p/:slug link is opened from a vendor session.
            base_flt = _site_filter_for_user(user, show_all=False) if user else {}
            base_flt[display] = chosen
            row = _clean(await db.sites.find_one(base_flt))
        elif source.startswith("master:"):
            table = source.split(":", 1)[1]
            md = _clean(await db.master_data.find_one({"table": table, f"data.{display}": chosen}))
            row = (md or {}).get("data") if md else None
        else:
            raise HTTPException(400, "Unknown source")
        if not row:
            return {"value": None, "fill": {}, "matched": False}
        return {
            "value": row.get(return_col, chosen),
            "fill": {c: row.get(c) for c in fill_cols},
            "matched": True,
        }

    # ---------- Pending Approvals (admin approval workflow) ----------
    def _import_perm():
        from permissions import can_approve as _ca, approval_level as _al
        return _ca, _al

    @approvals.get("")
    async def list_approvals(user=Depends(get_current_user),
                             status: str = "pending", limit: int = 200):
        # Super/admin see items within their scope; vendor_admin sees only
        # their own requests. Regular vendor users see nothing.
        flt: Dict[str, Any] = {"status": status} if status else {}
        if user.role == "vendor_admin":
            flt["requested_by.user_id"] = user.user_id
        elif user.role not in ("super_admin", "admin"):
            return []
        rows = await db.pending_approvals.find(flt, {"_id": 0}) \
            .sort("created_at", -1).to_list(min(limit, 1000))
        if user.role in ("super_admin", "admin"):
            can_approve, _ = _import_perm()
            # Only surface rows the caller could actually act on. Super
            # Admin (or override) always passes; region/cluster admins get
            # only their scope. Vendor Admin's `.get("")` branch already
            # scoped by user_id above.
            rows = [r for r in rows if can_approve(user, r)]
        return rows

    @approvals.post("/{approval_id}/approve")
    async def approve(approval_id: str, user=Depends(get_current_user)):
        row = _clean(await db.pending_approvals.find_one({"approval_id": approval_id}))
        if not row:
            raise HTTPException(404, "Approval not found")
        if row.get("status") != "pending":
            raise HTTPException(400, f"Approval already {row.get('status')}")
        can_approve, _ = _import_perm()
        if not can_approve(user, row):
            raise HTTPException(403, "You are not in the approval scope for this request")
        # Apply the action
        if row["type"] == "user_disable":
            await db.users.update_one(
                {"user_id": row["target_user_id"]},
                {"$set": {"is_active": False,
                          "deactivated_at": _now(),
                          "deactivated_by": user.user_id}},
            )
        await db.pending_approvals.update_one(
            {"approval_id": approval_id},
            {"$set": {"status": "approved",
                      "resolved_at": _now(),
                      "resolved_by": {"user_id": user.user_id,
                                      "name": getattr(user, "name", None),
                                      "role": user.role}}},
        )
        return {"ok": True}

    @approvals.post("/{approval_id}/reject")
    async def reject(approval_id: str, body: Optional[Dict[str, str]] = None,
                     user=Depends(get_current_user)):
        row = _clean(await db.pending_approvals.find_one({"approval_id": approval_id}))
        if not row:
            raise HTTPException(404, "Approval not found")
        if row.get("status") != "pending":
            raise HTTPException(400, f"Approval already {row.get('status')}")
        can_approve, _ = _import_perm()
        if not can_approve(user, row):
            raise HTTPException(403, "You are not in the approval scope for this request")
        note = (body or {}).get("reason", "")
        await db.pending_approvals.update_one(
            {"approval_id": approval_id},
            {"$set": {"status": "rejected",
                      "resolved_at": _now(),
                      "resolved_by": {"user_id": user.user_id,
                                      "name": getattr(user, "name", None),
                                      "role": user.role},
                      "reject_reason": note}},
        )
        return {"ok": True}

    return vendors, vusers, sites, master, lookup, public_lookup, approvals


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

async def _upsert_site(db, row: Dict[str, Any], user) -> Dict[str, Any]:
    """Upsert a site by either `site_id` (existing row) or `site_code`.
    If the row carries `vendor_email` but no `vendor_id`, try to resolve and
    attach `vendor_id` so vendor users automatically see their own sites.
    """
    row = {k: _coerce_value(k, v) for k, v in row.items() if v is not None}
    # Normalise `vendor_email`: accept multiple emails separated by ; , or newline.
    # Keep the first one in `vendor_email` (main contact / display value) and put
    # every one of them into `allowed_emails` so the RLS filter grants access
    # to all listed users under the same `vendor_id`.
    raw_email = row.get("vendor_email")
    if isinstance(raw_email, str) and raw_email.strip():
        emails = _split_emails(raw_email)
        if emails:
            row["vendor_email"] = emails[0]
            row["allowed_emails"] = emails
    elif isinstance(row.get("allowed_emails"), str):
        row["allowed_emails"] = _split_emails(row["allowed_emails"])
    # auto-link vendor when an email OR a vendor_name is provided.  We match
    # against BOTH the users collection (legacy behaviour) and the vendors
    # collection (new Vendor Management doc), preferring the exact vendor
    # doc when present.  This ensures sites entered via Site Master with a
    # freeform `vendor_name` still get a real `vendor_id` on the row.
    if not row.get("vendor_id"):
        vendor_doc = None
        vname = (row.get("vendor_name") or "").strip()
        vemail = (row.get("vendor_email") or "").strip().lower()
        if vname:
            vendor_doc = await db.vendors.find_one(
                {"$or": [
                    {"name":        {"$regex": f"^{re.escape(vname)}$", "$options": "i"}},
                    {"vendor_name": {"$regex": f"^{re.escape(vname)}$", "$options": "i"}},
                ]},
                {"_id": 0, "vendor_id": 1},
            )
        if vendor_doc is None and vemail:
            vendor_doc = await db.vendors.find_one(
                {"$or": [{"email": vemail}, {"vendor_email": vemail}]},
                {"_id": 0, "vendor_id": 1},
            )
        if vendor_doc and vendor_doc.get("vendor_id"):
            row["vendor_id"] = vendor_doc["vendor_id"]
        elif vemail:
            # Final fallback — legacy behaviour: link via a user email.
            vendor_user = await db.users.find_one(
                {"email": vemail, "vendor_id": {"$exists": True}},
                {"_id": 0, "vendor_id": 1},
            )
            if vendor_user and vendor_user.get("vendor_id"):
                row["vendor_id"] = vendor_user["vendor_id"]
    site_id = row.get("site_id")
    if site_id:
        existing = _clean(await db.sites.find_one({"site_id": site_id}))
        if existing:
            upd = {**row, "updated_at": _now(), "updated_by": user.user_id,
                   "version": int(existing.get("version", 1)) + 1}
            await db.site_versions.insert_one({
                "snapshot_id": _gen("siv"), "site_id": site_id,
                "version": existing.get("version", 1), "row": existing,
                "saved_at": _now(), "saved_by": user.user_id,
            })
            await db.sites.update_one({"site_id": site_id}, {"$set": upd})
            await _audit_master(db, user, "site.update", site_id, {"changes": list(row.keys())})
            return {**existing, **upd}
    # by site_code
    if row.get("site_code"):
        existing = _clean(await db.sites.find_one({"site_code": row["site_code"]}))
        if existing:
            upd = {**row, "updated_at": _now(), "updated_by": user.user_id,
                   "version": int(existing.get("version", 1)) + 1}
            await db.sites.update_one({"site_id": existing["site_id"]}, {"$set": upd})
            return {**existing, **upd}
    # otherwise insert
    new = {
        "site_id": _gen("site"),
        **row,
        "version": 1, "created_at": _now(), "created_by": user.user_id,
        "updated_at": _now(),
    }
    await db.sites.insert_one(dict(new))
    await _audit_master(db, user, "site.create", new["site_id"], {"name": new.get("site_name")})
    # Auto-provision the plant's document vault folders on disk so admins
    # find a pre-organised structure the first time they open the tab.
    try:
        from plant_docs_routes import bootstrap_new_plant
        await bootstrap_new_plant(db, new["site_id"])
    except Exception:
        pass
    return new


def _coerce_value(key: str, v: Any) -> Any:
    """Best-effort type conversion for known numeric / date columns."""
    if v is None or v == "":
        return v
    numeric = {"latitude", "longitude", "ac_capacity", "dc_capacity", "inverter_capacity"}
    if key in numeric:
        try:
            return float(v)
        except (TypeError, ValueError):
            return v
    return v


_EMAIL_SPLIT_RE = re.compile(r"[;,\s]+")


def _split_emails(raw: str) -> List[str]:
    """Split a `;`/`,`/whitespace-separated string of emails into a
    deduplicated, lower-cased list.  Bad tokens are silently dropped."""
    out: List[str] = []
    seen: set = set()
    for token in _EMAIL_SPLIT_RE.split(raw or ""):
        e = token.strip().lower()
        if not e or "@" not in e or e in seen:
            continue
        seen.add(e)
        out.append(e)
    return out


async def _audit_master(db, user, action: str, target_id: Optional[str], details: Dict[str, Any]) -> None:
    await db.audit_logs.insert_one({
        "audit_id": _gen("aud"),
        "actor_id": getattr(user, "user_id", None),
        "actor_email": getattr(user, "email", None),
        "action": action,
        "target_type": "master_data",
        "target_id": target_id,
        "details": details or {},
        "ip": None,
        "created_at": _now(),
    })


# ---------------------------------------------------------------------------
# Seed a tiny demo Site Master
# ---------------------------------------------------------------------------

DEMO_SITES = [
    {"site_name": "Alpha Solar 50MW", "site_code": "ALPHA-50", "asset_id": "AST-1001",
     "plant_name": "Alpha", "customer_name": "Acme Power",
     "state": "Karnataka", "district": "Tumkur", "location": "Pavagada",
     "latitude": 14.099, "longitude": 77.275,
     "ac_capacity": 50, "dc_capacity": 65, "inverter_capacity": 50,
     "vendor_name": "SunOps Pvt Ltd", "vendor_email": "ops@sunops.example.com",
     "approver_email": "approver.alpha@example.com",
     "cluster": "South-1", "region": "South", "site_status": "operational",
     "cluster_manager_name": "Rahul Verma",
     "commission_date": "2023-04-01"},
    {"site_name": "Bravo Wind 30MW", "site_code": "BRAVO-30", "asset_id": "AST-1002",
     "plant_name": "Bravo", "customer_name": "Helios Energy",
     "state": "Tamil Nadu", "district": "Tirunelveli", "location": "Kayathar",
     "latitude": 8.95, "longitude": 77.79,
     "ac_capacity": 30, "dc_capacity": 0, "inverter_capacity": 30,
     "vendor_name": "WindWorks", "vendor_email": "ops@windworks.example.com",
     "approver_email": "approver.bravo@example.com",
     "cluster": "South-2", "region": "South", "site_status": "operational",
     "cluster_manager_name": "Rahul Verma",
     "commission_date": "2022-09-10"},
    {"site_name": "Charlie Hybrid 25MW", "site_code": "CHARLIE-25", "asset_id": "AST-1003",
     "plant_name": "Charlie", "customer_name": "Acme Power",
     "state": "Gujarat", "district": "Kutch", "location": "Mundra",
     "latitude": 22.83, "longitude": 69.69,
     "ac_capacity": 25, "dc_capacity": 32, "inverter_capacity": 25,
     "vendor_name": "SunOps Pvt Ltd", "vendor_email": "ops@sunops.example.com",
     "approver_email": "approver.charlie@example.com",
     "cluster": "West-1", "region": "West", "site_status": "commissioning",
     "cluster_manager_name": "Priya Sharma",
     "commission_date": "2024-11-15"},
]


async def seed_demo_sites(db) -> None:
    """Idempotent demo data — does nothing if any site already exists.

    Also performs a one-time backfill of `cluster_manager_name` on existing
    demo sites for the new admin-by-cluster-manager RLS feature.
    """
    # Backfill cluster_manager_name and approver_email on demo rows.
    # NOTE: for the three seed site_codes we always overwrite approver_email
    # because an earlier iteration accidentally copied cluster_manager into
    # this column.  Non-demo rows are left untouched.
    for r in DEMO_SITES:
        if r.get("cluster_manager_name"):
            await db.sites.update_many(
                {"site_code": r["site_code"], "cluster_manager_name": {"$in": [None, ""]}},
                {"$set": {"cluster_manager_name": r["cluster_manager_name"]}},
            )
        if r.get("approver_email"):
            await db.sites.update_one(
                {"site_code": r["site_code"]},
                {"$set": {"approver_email": r["approver_email"]}},
            )
    if await db.sites.count_documents({}) > 0:
        return
    now = _now()
    for r in DEMO_SITES:
        await db.sites.insert_one({
            "site_id": _gen("site"), **r,
            "version": 1, "created_at": now, "updated_at": now,
        })
